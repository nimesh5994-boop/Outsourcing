"""Writes canonical DataFrames onto raw "DATA_*" sheets in the workbook, and
hands back cell-range references for each column - the foundation every
formula-linked schedule builds on. Every schedule cell should be a formula
referencing these ranges (or another schedule's own cells), not a
Python-computed literal, so the workbook recalculates like a manually-built
working paper and a reviewer can trace the formula chain.

A synthetic RowID column is added to nominal activity so a matrix-style
schedule (one row per transaction) can key a SUMIFS to exactly one source
row without relying on text criteria that might not be unique.
"""
from dataclasses import dataclass, field

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.xlformulas import range_ref

HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FONT = Font(bold=True)


@dataclass
class SheetRefs:
    """Column letter lookup + row bounds for one DATA_* sheet."""
    sheet_name: str
    columns: dict[str, str]  # canonical column -> Excel column letter
    first_row: int
    last_row: int  # first_row - 1 if the sheet has no data rows

    def col_range(self, column: str) -> str:
        return range_ref(self.sheet_name, self.columns[column], self.first_row, max(self.last_row, self.first_row))

    def is_empty(self) -> bool:
        return self.last_row < self.first_row


@dataclass
class DataRefs:
    tb_current: SheetRefs | None = None
    tb_comparative: SheetRefs | None = None
    nominal_current: SheetRefs | None = None
    aged_debtors: SheetRefs | None = None
    aged_creditors: SheetRefs | None = None
    pl_current: SheetRefs | None = None
    bs_current: SheetRefs | None = None
    fixed_asset_register: SheetRefs | None = None


def _cell_value(val):
    if isinstance(val, bool):
        return val
    if isinstance(val, pd.Timestamp):
        return val.date() if pd.notna(val) else None
    if isinstance(val, (int, float, str)):
        return val
    if val is None or pd.isna(val):
        return None
    return val


def _write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> SheetRefs:
    ws: Worksheet = wb.create_sheet(sheet_name)
    columns = list(df.columns)
    col_letters = {col: get_column_letter(i + 1) for i, col in enumerate(columns)}

    for i, col in enumerate(columns):
        cell = ws.cell(row=1, column=i + 1, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    r = 2
    for _, row in df.iterrows():
        for i, col in enumerate(columns):
            ws.cell(row=r, column=i + 1, value=_cell_value(row[col]))
        r += 1
    last_row = r - 1

    ws.sheet_state = "hidden"
    return SheetRefs(sheet_name=sheet_name, columns=col_letters, first_row=2, last_row=last_row)


def write_data_sheets(wb: Workbook, data: dict) -> DataRefs:
    refs = DataRefs()

    if data.get("tb_current") is not None and not data["tb_current"].empty:
        df = data["tb_current"][["account_code", "account_name", "account_type", "debit", "credit", "balance"]]
        refs.tb_current = _write_sheet(wb, "DATA_TB_Current", df)

    if data.get("tb_comparative") is not None and not data["tb_comparative"].empty:
        df = data["tb_comparative"][["account_code", "account_name", "account_type", "debit", "credit", "balance"]]
        refs.tb_comparative = _write_sheet(wb, "DATA_TB_Comparative", df)

    if data.get("nominal_current") is not None and not data["nominal_current"].empty:
        df = data["nominal_current"].copy().reset_index(drop=True)
        df.insert(0, "row_id", range(1, len(df) + 1))
        cols = ["row_id", "date", "account_code", "account_name", "reference", "description", "contact", "source_type", "debit", "credit"]
        cols = [c for c in cols if c in df.columns]
        refs.nominal_current = _write_sheet(wb, "DATA_Nominal", df[cols])

    if data.get("aged_debtors") is not None and not data["aged_debtors"].empty:
        refs.aged_debtors = _write_sheet(wb, "DATA_AgedDebtors", data["aged_debtors"])

    if data.get("aged_creditors") is not None and not data["aged_creditors"].empty:
        refs.aged_creditors = _write_sheet(wb, "DATA_AgedCreditors", data["aged_creditors"])

    if data.get("pl_current") is not None and not data["pl_current"].empty:
        refs.pl_current = _write_sheet(wb, "DATA_PL", data["pl_current"])

    if data.get("bs_current") is not None and not data["bs_current"].empty:
        refs.bs_current = _write_sheet(wb, "DATA_BS", data["bs_current"])

    if data.get("fixed_asset_register") is not None and not data["fixed_asset_register"].empty:
        refs.fixed_asset_register = _write_sheet(wb, "DATA_FixedAssetRegister", data["fixed_asset_register"])

    return refs
