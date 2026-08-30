"""Builds the final Excel working paper pack from canonical data + recon
results, in the house style observed in the firm's real working paper files:
a CLIENT NAME / PERIOD / SCHEDULE TITLE header block on every sheet, and a
numbered index of schedules with their status and any queries raised.
"""
import io
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet

from app.control_accounts import ControlAccountResult
from app.corporation_tax import CTComputation
from app.data_sheets import DataRefs, with_row_ids, write_data_sheets
from app.financial_statements import (
    _BS_CURRENT_ASSETS,
    _BS_CURRENT_LIABILITIES,
    _BS_EQUITY,
    _BS_FIXED_ASSETS,
    _BS_LONG_TERM_LIABILITIES,
    _PL_DIRECT_COSTS,
    _PL_TURNOVER,
    StatementResult,
    build_bs_statement,
    build_pl_statement,
)
from app.fixed_assets import AssetRegisterResult, FixedAssetResult, group_fixed_asset_codes
from app.nominal_matrix import MatrixResult, build_matrix_row_groups
from app.recon import MATERIALITY_AMOUNT, VARIANCE_PCT_THRESHOLD, ReconResult
from app.xlformulas import cell_ref, literal, quote, sum_of_values, sumifs_exact

NAVY = "1F3864"
GREEN = "C6EFCE"
GREEN_TEXT = "006100"
AMBER = "FFEB9C"
AMBER_TEXT = "9C6500"
RED = "FFC7CE"
RED_TEXT = "9C0006"
GREY = "F2F2F2"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
CLIENT_FONT = Font(name="Calibri", size=13, bold=True, color=NAVY)
PERIOD_FONT = Font(name="Calibri", size=11, color="595959")
SCHEDULE_FONT = Font(name="Calibri", size=12, bold=True, color="000000")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")
BOLD = Font(name="Calibri", size=11, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CURRENCY_FMT = '#,##0.00;[RED](#,##0.00)'


def _style_header_row(ws: Worksheet, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _column_widths(df: pd.DataFrame) -> list[int]:
    widths = []
    for col in df.columns:
        content_max = int(df[col].astype(str).str.len().max()) if len(df) else 0
        width = max(content_max, len(str(col))) + 2
        widths.append(max(12, min(60, width)))
    return widths


def _autosize(ws: Worksheet, df: pd.DataFrame, start_col: int = 1):
    for i, width in enumerate(_column_widths(df)):
        ws.column_dimensions[get_column_letter(start_col + i)].width = width


DEFAULT_HEADER_CELLS = {"client_name_cell": "A1", "period_cell": "A2", "schedule_title_cell": "A3"}


def _write_title(
    ws: Worksheet, client_name: str, period_label: str, schedule_title: str, ref: str = "",
    header_cells: dict | None = None,
) -> int:
    """House-style header block: CLIENT NAME / PERIOD / SCHEDULE TITLE,
    written into whichever cells a template's config specifies
    (storage.DEFAULT_TEMPLATE_CONFIG's header_cells) - defaults to A1/A2/A3
    (the generic-layout convention) when no template config applies.
    Returns the first free row below the header block, computed from
    whichever of the three configured cells sits lowest rather than
    hardcoded, since a template's own convention might not put them on
    rows 1-3."""
    cells = header_cells or DEFAULT_HEADER_CELLS
    client_cell = cells.get("client_name_cell") or DEFAULT_HEADER_CELLS["client_name_cell"]
    period_cell = cells.get("period_cell") or DEFAULT_HEADER_CELLS["period_cell"]
    title_cell = cells.get("schedule_title_cell") or DEFAULT_HEADER_CELLS["schedule_title_cell"]

    ws[client_cell] = client_name.upper()
    ws[client_cell].font = CLIENT_FONT
    ws[period_cell] = period_label.upper()
    ws[period_cell].font = PERIOD_FONT
    ws[title_cell] = f"{ref + '  ' if ref else ''}{schedule_title}"
    ws[title_cell].font = SCHEDULE_FONT

    last_row = max(coordinate_from_string(c)[1] for c in (client_cell, period_cell, title_cell))
    return last_row + 2


def _status_fill(status: str) -> PatternFill:
    return {
        "ok": PatternFill("solid", fgColor=GREEN),
        "review": PatternFill("solid", fgColor=AMBER),
        "error": PatternFill("solid", fgColor=RED),
        "n/a": PatternFill("solid", fgColor=GREY),
    }.get(status, PatternFill("solid", fgColor=GREY))


def _status_font(status: str) -> Font:
    color = {"ok": GREEN_TEXT, "review": AMBER_TEXT, "error": RED_TEXT, "n/a": "595959"}.get(status, "000000")
    return Font(bold=True, color=color)


def _write_dataframe(ws: Worksheet, df: pd.DataFrame, start_row: int, start_col: int = 1) -> int:
    """Writes header + rows starting at start_row, returns the next free row.
    Columns are sized to their content (capped at 60 - wider than that reads
    worse than wrapping), and any cell whose text still doesn't fit is
    wrapped with the row height grown to fit rather than left to truncate."""
    widths = _column_widths(df)
    for i, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = width

    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=str(col))
    _style_header_row(ws, start_row, len(df.columns))

    r = start_row + 1
    for _, row in df.iterrows():
        max_lines = 1
        for j, col in enumerate(df.columns):
            val = row[col]
            if isinstance(val, pd.Timestamp):
                val = val.date() if not pd.isna(val) else ""
            cell = ws.cell(row=r, column=start_col + j, value=val)
            cell.border = BORDER
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = CURRENCY_FMT
            elif isinstance(val, str) and len(val) > widths[j]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                max_lines = max(max_lines, -(-len(val) // widths[j]))  # ceil division
        if max_lines > 1:
            ws.row_dimensions[r].height = min(120, 15 * max_lines)
        r += 1
    return r + 1


class _RefCounter:
    def __init__(self, start: int = 1):
        self.n = start - 1

    def next(self) -> str:
        self.n += 1
        return str(self.n)


def build_index_sheet(ws: Worksheet, client_name: str, current_label: str, comparative_label: str, entries: list[dict]):
    """Writes the index onto whatever worksheet the caller hands it -
    build_workbook passes wb.active (a fresh workbook's only sheet, so
    reusing it as "Index" is free), template-based generation passes a
    freshly created sheet instead, since wb.active there is the practice's
    own real first sheet and must not be overwritten."""
    ws["A1"] = client_name.upper()
    ws["A1"].font = CLIENT_FONT
    ws["A2"] = f"WORKING PAPERS - {current_label.upper()}"
    ws["A2"].font = PERIOD_FONT
    ws["A3"] = f"Comparative period: {comparative_label or 'none'}"
    ws["A3"].font = SUBTITLE_FONT
    ws["A4"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws["A4"].font = SUBTITLE_FONT

    headers = ["Ref", "Schedule", "Status", "Comment"]
    start = 6
    for j, h in enumerate(headers):
        ws.cell(row=start, column=1 + j, value=h)
    _style_header_row(ws, start, len(headers))

    r = start + 1
    for e in entries:
        ws.cell(row=r, column=1, value=e["ref"]).border = BORDER
        ws.cell(row=r, column=2, value=e["title"]).border = BORDER
        status_cell = ws.cell(row=r, column=3, value=e["status"].upper())
        status_cell.fill = _status_fill(e["status"])
        status_cell.font = _status_font(e["status"])
        status_cell.alignment = Alignment(horizontal="center")
        status_cell.border = BORDER
        comment_cell = ws.cell(row=r, column=4, value=e["message"])
        comment_cell.border = BORDER
        comment_cell.alignment = Alignment(wrap_text=True)
        r += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 130
    ws.freeze_panes = "A7"


def build_tb_lead_schedule(wb: Workbook, client_name: str, current_label: str, ref: str, variance_detail: pd.DataFrame, header_cells: dict | None = None):
    ws = wb.create_sheet(f"{ref} TB Lead Schedule"[:31])
    row = _write_title(ws, client_name, current_label, "TRIAL BALANCE LEAD SCHEDULE", ref, header_cells=header_cells)
    if variance_detail is None or variance_detail.empty:
        ws.cell(row=row, column=1, value="No trial balance data available.")
        return

    df = variance_detail.copy()
    df["Reviewed / reallocation required?"] = ""
    df = df.rename(columns={
        "account_code": "Account Code", "account_name": "Account Name",
        "current_year": "Current Year", "comparative_year": "Comparative Year",
        "variance_amount": "Variance (£)", "variance_pct": "Variance (%)", "flag": "Flagged?",
    })
    df["Variance (%)"] = df["Variance (%)"].apply(lambda x: round(x * 100, 1))
    next_row = _write_dataframe(ws, df, start_row=row)

    flag_col = list(df.columns).index("Flagged?") + 1
    for r in range(row + 1, next_row - 1):
        if ws.cell(row=r, column=flag_col).value is True:
            for c in range(1, len(df.columns) + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=AMBER)
    ws.freeze_panes = f"A{row + 1}"


def build_tb_lead_schedule_formulas(wb: Workbook, client_name: str, current_label: str, ref: str, variance_detail: pd.DataFrame, refs: DataRefs, header_cells: dict | None = None):
    """Same schedule as build_tb_lead_schedule, but every numeric cell is a
    live formula against the DATA_TB_* sheets instead of a Python-computed
    literal - which accounts appear and their sort order still come from
    variance_detail (a listing question, not a value one)."""
    ws = wb.create_sheet(f"{ref} TB Lead Schedule"[:31])
    row = _write_title(ws, client_name, current_label, "TRIAL BALANCE LEAD SCHEDULE", ref, header_cells=header_cells)
    if variance_detail is None or variance_detail.empty or refs.tb_current is None:
        ws.cell(row=row, column=1, value="No trial balance data available.")
        return

    headers = ["Account Code", "Account Name", "Current Year", "Comparative Year", "Variance (£)", "Variance (%)", "Flagged?", "Reviewed / reallocation required?"]
    for j, h in enumerate(headers):
        ws.cell(row=row, column=1 + j, value=h)
    _style_header_row(ws, row, len(headers))

    cur_range = refs.tb_current.col_range("balance")
    cur_code_range = refs.tb_current.col_range("account_code")
    comp_range = refs.tb_comparative.col_range("balance") if refs.tb_comparative else None
    comp_code_range = refs.tb_comparative.col_range("account_code") if refs.tb_comparative else None

    r = row + 1
    for _, acct in variance_detail.iterrows():
        code = str(acct["account_code"])
        ws.cell(row=r, column=1, value=code).border = BORDER
        ws.cell(row=r, column=2, value=acct["account_name"]).border = BORDER

        cur_cell = ws.cell(row=r, column=3, value=sumifs_exact(cur_range, (cur_code_range, quote(code))))
        cur_cell.number_format = CURRENCY_FMT
        cur_cell.border = BORDER

        if comp_range:
            comp_cell = ws.cell(row=r, column=4, value=sumifs_exact(comp_range, (comp_code_range, quote(code))))
        else:
            comp_cell = ws.cell(row=r, column=4, value=0)
        comp_cell.number_format = CURRENCY_FMT
        comp_cell.border = BORDER

        var_cell = ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
        var_cell.number_format = CURRENCY_FMT
        var_cell.border = BORDER

        pct_cell = ws.cell(row=r, column=6, value=f"=IF(D{r}=0,IF(E{r}<>0,1,0),E{r}/ABS(D{r}))")
        pct_cell.number_format = "0.0%"
        pct_cell.border = BORDER

        flag_cell = ws.cell(row=r, column=7, value=f"=AND(ABS(E{r})>={MATERIALITY_AMOUNT!r},ABS(F{r})>={VARIANCE_PCT_THRESHOLD!r})")
        flag_cell.border = BORDER

        ws.cell(row=r, column=8, value="").border = BORDER
        r += 1

    widths = [14, 40, 14, 16, 14, 12, 10, 40]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    ws.freeze_panes = f"A{row + 1}"


def build_statement_sheet(wb: Workbook, client_name: str, period_label: str, ref: str, sheet_name: str, title: str, df: pd.DataFrame, statement: StatementResult | None = None, header_cells: dict | None = None):
    ws = wb.create_sheet(sheet_name[:31])
    row = _write_title(ws, client_name, period_label, title, ref, header_cells=header_cells)
    if df is None or df.empty:
        ws.cell(row=row, column=1, value="No data uploaded for this statement.")
        return

    if statement is not None and not statement.statement.empty:
        if statement.status in ("ok", "review"):
            status_cell = ws.cell(row=row, column=1, value=f"Status: {statement.status.upper()} - {statement.message}")
            status_cell.font = _status_font(statement.status)
            status_cell.fill = _status_fill(statement.status)
            status_cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = 30
            row += 2
        summary = statement.statement.copy()
        bold_lines = {"GROSS PROFIT", "NET PROFIT / (LOSS) FOR THE YEAR", "TOTAL ASSETS", "TOTAL LIABILITIES", "NET ASSETS", "TOTAL EQUITY"}
        next_row = _write_dataframe(ws, summary, start_row=row)
        for r_offset, line in enumerate(summary["Line"]):
            r = row + 1 + r_offset
            if line in bold_lines or line.startswith("CHECK") or line.startswith("NET PROFIT"):
                for c in (1, 2):
                    ws.cell(row=r, column=c).font = BOLD
            if line.startswith("CHECK") and statement.status != "ok":
                for c in (1, 2):
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=AMBER)
        row = next_row + 1
        ws.cell(row=row, column=1, value="DETAIL BY ACCOUNT").font = SCHEDULE_FONT
        row += 1

    display = df.rename(columns={"account_code": "Account Code", "account_name": "Account Name", "category": "Category", "amount": "Amount"})
    _write_dataframe(ws, display, start_row=row)
    ws.freeze_panes = f"A{row + 1}"


def build_recon_sheet(wb: Workbook, client_name: str, period_label: str, ref: str, sheet_name: str, result: ReconResult, header_cells: dict | None = None):
    ws = wb.create_sheet(sheet_name[:31])
    row = _write_title(ws, client_name, period_label, result.name, ref, header_cells=header_cells)
    status_cell = ws.cell(row=row, column=1, value=f"Status: {result.status.upper()} - {result.message}")
    status_cell.font = _status_font(result.status)
    status_cell.fill = _status_fill(result.status)
    status_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 30

    detail_row = row + 2
    next_row = detail_row
    if result.detail is None or result.detail.empty:
        ws.cell(row=detail_row, column=1, value="No supporting detail available.")
        next_row = detail_row + 2
    else:
        next_row = _write_dataframe(ws, result.detail, start_row=detail_row)
        ws.freeze_panes = f"A{detail_row + 1}"

    if result.extra_detail is not None and not result.extra_detail.empty:
        label_row = next_row + 1
        ws.cell(row=label_row, column=1, value=result.extra_detail_label.upper()).font = SCHEDULE_FONT
        next_row = _write_dataframe(ws, result.extra_detail, start_row=label_row + 1)

    if getattr(result, "ai_note", ""):
        note_row = next_row + 1
        note_cell = ws.cell(row=note_row, column=1, value=f"\U0001F916 AI-ASSISTED NOTE (not a finding - verify before relying on it): {result.ai_note}")
        note_cell.font = Font(italic=True, color="5B4B8A", size=10)
        note_cell.fill = PatternFill("solid", fgColor="F1EDFB")
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
        ws.row_dimensions[note_row].height = 45


def build_control_account_sheet(wb: Workbook, client_name: str, current_label: str, ref: str, result: ControlAccountResult, header_cells: dict | None = None):
    sheet_name = f"{ref} {result.account_name}"[:31]
    ws = wb.create_sheet(sheet_name)
    suffix = "" if "control account" in result.account_name.lower() else " CONTROL ACCOUNT"
    title = f"{result.account_name.upper()}{suffix} (nominal {result.account_code})"
    row = _write_title(ws, client_name, current_label, title, ref, header_cells=header_cells)

    status_cell = ws.cell(row=row, column=1, value=f"Status: {result.status.upper()} - {result.message}")
    status_cell.font = _status_font(result.status)
    status_cell.fill = _status_fill(result.status)
    status_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 30

    table_row = row + 2
    if result.schedule.empty and result.breakdown.empty:
        ws.cell(row=table_row, column=1, value="No data available.")
        return

    next_row = table_row
    if not result.schedule.empty:
        next_row = _write_dataframe(ws, result.schedule, start_row=table_row)
        diff_row = next_row - 2
        if str(ws.cell(row=diff_row, column=1).value or "").startswith("DIFFERENCE"):
            for c in range(1, 5):
                ws.cell(row=diff_row, column=c).font = BOLD
                if result.status != "ok":
                    ws.cell(row=diff_row, column=c).fill = PatternFill("solid", fgColor=AMBER)

    if not result.breakdown.empty:
        header_row = next_row
        ws.cell(row=header_row, column=1, value=f"BREAKDOWN OF CLOSING BALANCE ({result.breakdown_label})").font = SCHEDULE_FONT
        breakdown_start = header_row + 1
        end_row = _write_dataframe(ws, result.breakdown, start_row=breakdown_start)
        # bold + flag the three summary rows appended at the bottom (total / TB balance / unexplained diff)
        for offset in range(1, 4):
            r = end_row - 1 - offset
            for c in range(1, 3):
                ws.cell(row=r, column=c).font = BOLD
        diff_row = end_row - 2
        if result.status != "ok" and str(ws.cell(row=diff_row, column=1).value or "").startswith("UNEXPLAINED"):
            for c in range(1, 3):
                ws.cell(row=diff_row, column=c).fill = PatternFill("solid", fgColor=AMBER)

    ws.freeze_panes = f"A{table_row + 1}"


def build_control_account_sheet_formulas(wb: Workbook, client_name: str, current_label: str, ref: str, result: ControlAccountResult, refs: DataRefs, header_cells: dict | None = None):
    """Same schedule as build_control_account_sheet, but every Dr/Cr cell is
    a live formula against the DATA sheets. A hidden helper column (F)
    holds each row's raw signed balance so the Debit/Credit split (and the
    DIFFERENCE row's own arithmetic) can reference it without repeating the
    same SUMPRODUCT twice per row."""
    sheet_name = f"{ref} {result.account_name}"[:31]
    ws = wb.create_sheet(sheet_name)
    suffix = "" if "control account" in result.account_name.lower() else " CONTROL ACCOUNT"
    title = f"{result.account_name.upper()}{suffix} (nominal {result.account_code})"
    row = _write_title(ws, client_name, current_label, title, ref, header_cells=header_cells)

    status_cell = ws.cell(row=row, column=1, value=f"Status: {result.status.upper()} - {result.message}")
    status_cell.font = _status_font(result.status)
    status_cell.fill = _status_fill(result.status)
    status_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 30

    code = result.account_code
    has_movement = refs.nominal_current is not None and len(result.schedule) >= 4

    table_row = row + 2
    headers = ["Item", "Reference", "Debit £", "Credit £"]
    for j, h in enumerate(headers):
        ws.cell(row=table_row, column=1 + j, value=h)
    _style_header_row(ws, table_row, len(headers))

    def dr_cr_from_helper(r: int):
        ws.cell(row=r, column=3, value=f"=IF(F{r}>=0,F{r},0)").number_format = CURRENCY_FMT
        ws.cell(row=r, column=4, value=f"=IF(F{r}<0,-F{r},0)").number_format = CURRENCY_FMT

    r = table_row + 1
    bfwd_row = r
    ws.cell(row=r, column=1, value="BALANCE B/FWD").border = BORDER
    ws.cell(row=r, column=2, value="Comparative year TB").border = BORDER
    bfwd_formula = sumifs_exact(refs.tb_comparative.col_range("balance"), (refs.tb_comparative.col_range("account_code"), quote(code))) if refs.tb_comparative else "=0"
    ws.cell(row=r, column=6, value=bfwd_formula)
    dr_cr_from_helper(r)
    r += 1

    move_row = None
    if has_movement:
        move_row = r
        ws.cell(row=r, column=1, value="MOVEMENTS DURING YEAR").border = BORDER
        ws.cell(row=r, column=2, value="Nominal activity detail").border = BORDER
        ws.cell(row=r, column=3, value=sumifs_exact(refs.nominal_current.col_range("debit"), (refs.nominal_current.col_range("account_code"), quote(code)))).number_format = CURRENCY_FMT
        ws.cell(row=r, column=4, value=sumifs_exact(refs.nominal_current.col_range("credit"), (refs.nominal_current.col_range("account_code"), quote(code)))).number_format = CURRENCY_FMT
    else:
        ws.cell(row=r, column=1, value="MOVEMENTS DURING YEAR").border = BORDER
        ws.cell(row=r, column=2, value="No nominal activity detail available").border = BORDER
    r += 1

    cfwd_row = r
    ws.cell(row=r, column=1, value="BALANCE C/FWD (per TB)").border = BORDER
    ws.cell(row=r, column=2, value="Current year TB").border = BORDER
    cfwd_formula = sumifs_exact(refs.tb_current.col_range("balance"), (refs.tb_current.col_range("account_code"), quote(code))) if refs.tb_current else "=0"
    ws.cell(row=r, column=6, value=cfwd_formula)
    dr_cr_from_helper(r)
    r += 1

    if move_row is not None:
        diff_row = r
        ws.cell(row=r, column=1, value="DIFFERENCE (computed c/fwd vs per TB)").font = BOLD
        ws.cell(row=r, column=2, value="").border = BORDER
        ws.cell(row=r, column=6, value=f"=F{cfwd_row}-(F{bfwd_row}+C{move_row}-D{move_row})")
        dr_cr_from_helper(r)
        for c in (1, 2, 3, 4):
            ws.cell(row=r, column=c).font = BOLD
            if result.status != "ok":
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=AMBER)
        ws.cell(row=r, column=1).border = BORDER
        r += 1

    ws.column_dimensions["F"].hidden = True
    next_row = r + 1

    if not result.breakdown.empty:
        aged_refs = refs.aged_debtors if "debtor" in result.breakdown_label.lower() else refs.aged_creditors
        party_col = "customer" if "debtor" in result.breakdown_label.lower() else "supplier"
        if aged_refs is not None:
            ws.cell(row=next_row, column=1, value=f"BREAKDOWN OF CLOSING BALANCE ({result.breakdown_label})").font = SCHEDULE_FONT
            b_row = next_row + 1
            for j, h in enumerate(["Party", "Amount £"]):
                ws.cell(row=b_row, column=1 + j, value=h)
            _style_header_row(ws, b_row, 2)

            n_parties = aged_refs.last_row - aged_refs.first_row + 1
            first_party_row = b_row + 1
            for i in range(n_parties):
                src_row = aged_refs.first_row + i
                dest_row = first_party_row + i
                ws.cell(row=dest_row, column=1, value="=" + cell_ref(aged_refs.sheet_name, f"{aged_refs.columns[party_col]}{src_row}")).border = BORDER
                amt_cell = ws.cell(row=dest_row, column=2, value="=" + cell_ref(aged_refs.sheet_name, f"{aged_refs.columns['total']}{src_row}"))
                amt_cell.number_format = CURRENCY_FMT
                amt_cell.border = BORDER

            total_row = first_party_row + n_parties
            ws.cell(row=total_row, column=1, value="TOTAL PER BREAKDOWN").font = BOLD
            total_cell = ws.cell(row=total_row, column=2, value=f"=SUM(B{first_party_row}:B{total_row - 1})")
            total_cell.number_format = CURRENCY_FMT
            total_cell.font = BOLD

            tb_row = total_row + 1
            ws.cell(row=tb_row, column=1, value="BALANCE PER TB").font = BOLD
            tb_formula = f"=ABS({sumifs_exact(refs.tb_current.col_range('balance'), (refs.tb_current.col_range('account_code'), quote(code)))[1:]})"
            tb_cell = ws.cell(row=tb_row, column=2, value=tb_formula)
            tb_cell.number_format = CURRENCY_FMT
            tb_cell.font = BOLD

            diff_bd_row = tb_row + 1
            ws.cell(row=diff_bd_row, column=1, value="UNEXPLAINED DIFFERENCE (for preparer to review)").font = BOLD
            diff_bd_cell = ws.cell(row=diff_bd_row, column=2, value=f"=B{total_row}-B{tb_row}")
            diff_bd_cell.number_format = CURRENCY_FMT
            diff_bd_cell.font = BOLD
            if result.status != "ok":
                for c in (1, 2):
                    ws.cell(row=diff_bd_row, column=c).fill = PatternFill("solid", fgColor=AMBER)

    ws.freeze_panes = f"A{table_row + 1}"


def build_pl_statement_sheet_formulas(wb: Workbook, client_name: str, period_label: str, ref: str, pl_result: StatementResult, refs: DataRefs, header_cells: dict | None = None) -> tuple[str, str | None]:
    """Same summary lines as build_statement_sheet's P&L, but every summary
    figure is a live SUMPRODUCT-by-category formula against DATA_PL, and the
    detail table is a row-for-row cell reference into DATA_PL rather than a
    duplicated literal table. Returns (sheet_name, net-profit cell ref) so
    the Balance Sheet formula sheet can bridge the current year's profit in
    with a real cross-sheet reference instead of a re-computed literal."""
    sheet_name = f"{ref} Profit and Loss"[:31]
    ws = wb.create_sheet(sheet_name)
    row = _write_title(ws, client_name, period_label, "PROFIT AND LOSS ACCOUNT", ref, header_cells=header_cells)

    status_cell = ws.cell(row=row, column=1, value=f"Status: {pl_result.status.upper()} - {pl_result.message}")
    status_cell.font = _status_font(pl_result.status)
    status_cell.fill = _status_fill(pl_result.status)
    status_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 30

    if refs.pl_current is None or refs.pl_current.is_empty():
        ws.cell(row=row + 2, column=1, value="No P&L data uploaded.")
        return sheet_name, None

    table_row = row + 2
    for j, h in enumerate(["Line", "Amount"]):
        ws.cell(row=table_row, column=1 + j, value=h)
    _style_header_row(ws, table_row, 2)

    amount_range = refs.pl_current.col_range("amount")
    category_range = refs.pl_current.col_range("category")

    r = table_row + 1
    turnover_row = r
    ws.cell(row=r, column=1, value="Turnover").border = BORDER
    ws.cell(row=r, column=2, value=sum_of_values(amount_range, category_range, [quote(v) for v in sorted(_PL_TURNOVER)])).number_format = CURRENCY_FMT
    r += 1

    direct_row = r
    ws.cell(row=r, column=1, value="Direct costs").border = BORDER
    ws.cell(row=r, column=2, value=sum_of_values(amount_range, category_range, [quote(v) for v in sorted(_PL_DIRECT_COSTS)])).number_format = CURRENCY_FMT
    r += 1

    gross_row = r
    ws.cell(row=r, column=1, value="GROSS PROFIT").font = BOLD
    gp_cell = ws.cell(row=r, column=2, value=f"=B{turnover_row}+B{direct_row}")
    gp_cell.number_format = CURRENCY_FMT
    gp_cell.font = BOLD
    r += 1

    overheads_row = r
    ws.cell(row=r, column=1, value="Overheads & other expenses").border = BORDER
    oh_cell = ws.cell(row=r, column=2, value=f"=SUM({amount_range})-B{turnover_row}-B{direct_row}")
    oh_cell.number_format = CURRENCY_FMT
    r += 1

    net_row = r
    ws.cell(row=r, column=1, value="NET PROFIT / (LOSS) FOR THE YEAR").font = BOLD
    net_cell = ws.cell(row=r, column=2, value=f"=B{gross_row}+B{overheads_row}")
    net_cell.number_format = CURRENCY_FMT
    net_cell.font = BOLD
    r += 2

    ws.cell(row=r, column=1, value="DETAIL BY ACCOUNT").font = SCHEDULE_FONT
    r += 1
    detail_headers = ["Account Code", "Account Name", "Category", "Amount"]
    for j, h in enumerate(detail_headers):
        ws.cell(row=r, column=1 + j, value=h)
    _style_header_row(ws, r, len(detail_headers))
    first_detail = r + 1
    n = refs.pl_current.last_row - refs.pl_current.first_row + 1
    for i in range(n):
        src = refs.pl_current.first_row + i
        dest = first_detail + i
        for col_idx, field in enumerate(["account_code", "account_name", "category"]):
            ws.cell(row=dest, column=1 + col_idx, value="=" + cell_ref(refs.pl_current.sheet_name, f"{refs.pl_current.columns[field]}{src}")).border = BORDER
        amt_cell = ws.cell(row=dest, column=4, value="=" + cell_ref(refs.pl_current.sheet_name, f"{refs.pl_current.columns['amount']}{src}"))
        amt_cell.number_format = CURRENCY_FMT
        amt_cell.border = BORDER

    ws.freeze_panes = f"A{table_row + 1}"
    return sheet_name, f"B{net_row}"


def build_bs_statement_sheet_formulas(
    wb: Workbook, client_name: str, period_label: str, ref: str, bs_result: StatementResult, refs: DataRefs,
    pl_sheet_name: str | None, pl_net_profit_cell: str | None, header_cells: dict | None = None,
) -> str:
    """Same lines as build_statement_sheet's Balance Sheet, including the
    explicit CHECK row, but computed via live formulas against DATA_BS - and
    the current year's profit is bridged in via a real cross-sheet reference
    to the P&L formula sheet's own NET PROFIT cell, not a re-typed number.
    Hidden column F holds each row's raw (unflipped) signed value, since the
    net-assets/total-equity/CHECK arithmetic needs the raw figures while the
    visible Amount column shows liabilities/equity the conventional way
    round (credits displayed positive)."""
    sheet_name = f"{ref} Balance Sheet"[:31]
    ws = wb.create_sheet(sheet_name)
    row = _write_title(ws, client_name, period_label, "BALANCE SHEET", ref, header_cells=header_cells)

    status_cell = ws.cell(row=row, column=1, value=f"Status: {bs_result.status.upper()} - {bs_result.message}")
    status_cell.font = _status_font(bs_result.status)
    status_cell.fill = _status_fill(bs_result.status)
    status_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 30

    if refs.bs_current is None or refs.bs_current.is_empty():
        ws.cell(row=row + 2, column=1, value="No Balance Sheet data uploaded.")
        return sheet_name

    table_row = row + 2
    for j, h in enumerate(["Line", "Amount"]):
        ws.cell(row=table_row, column=1 + j, value=h)
    _style_header_row(ws, table_row, 2)
    ws.column_dimensions["F"].hidden = True

    amount_range = refs.bs_current.col_range("amount")
    category_range = refs.bs_current.col_range("category")

    def raw_sum(values: set[str]) -> str:
        return sum_of_values(amount_range, category_range, [quote(v) for v in sorted(values)])

    r = table_row + 1
    fixed_row = r
    ws.cell(row=r, column=1, value="Fixed assets").border = BORDER
    ws.cell(row=r, column=6, value=raw_sum(_BS_FIXED_ASSETS))
    ws.cell(row=r, column=2, value=f"=F{r}").number_format = CURRENCY_FMT
    r += 1

    current_assets_row = r
    ws.cell(row=r, column=1, value="Current assets").border = BORDER
    ws.cell(row=r, column=6, value=raw_sum(_BS_CURRENT_ASSETS))
    ws.cell(row=r, column=2, value=f"=F{r}").number_format = CURRENCY_FMT
    r += 1

    total_assets_row = r
    ws.cell(row=r, column=1, value="TOTAL ASSETS").font = BOLD
    ta_cell = ws.cell(row=r, column=2, value=f"=B{fixed_row}+B{current_assets_row}")
    ta_cell.number_format = CURRENCY_FMT
    ta_cell.font = BOLD
    r += 1

    current_liab_row = r
    ws.cell(row=r, column=1, value="Current liabilities").border = BORDER
    ws.cell(row=r, column=6, value=raw_sum(_BS_CURRENT_LIABILITIES))
    ws.cell(row=r, column=2, value=f"=-F{r}").number_format = CURRENCY_FMT
    r += 1

    long_term_row = r
    ws.cell(row=r, column=1, value="Long-term liabilities").border = BORDER
    ws.cell(row=r, column=6, value=raw_sum(_BS_LONG_TERM_LIABILITIES))
    ws.cell(row=r, column=2, value=f"=-F{r}").number_format = CURRENCY_FMT
    r += 1

    total_liab_row = r
    ws.cell(row=r, column=1, value="TOTAL LIABILITIES").font = BOLD
    tl_cell = ws.cell(row=r, column=2, value=f"=-(F{current_liab_row}+F{long_term_row})")
    tl_cell.number_format = CURRENCY_FMT
    tl_cell.font = BOLD
    r += 1

    net_assets_row = r
    ws.cell(row=r, column=1, value="NET ASSETS").font = BOLD
    na_cell = ws.cell(row=r, column=2, value=f"=B{total_assets_row}+F{current_liab_row}+F{long_term_row}")
    na_cell.number_format = CURRENCY_FMT
    na_cell.font = BOLD
    r += 1

    equity_bfwd_row = r
    ws.cell(row=r, column=1, value="Equity brought forward").border = BORDER
    ws.cell(row=r, column=6, value=raw_sum(_BS_EQUITY))
    ws.cell(row=r, column=2, value=f"=-F{r}").number_format = CURRENCY_FMT
    r += 1

    profit_row = r
    ws.cell(row=r, column=1, value="Profit/(loss) for the year (per P&L, not yet closed in the TB)").border = BORDER
    net_profit_ref = ("=" + cell_ref(pl_sheet_name, pl_net_profit_cell)) if (pl_sheet_name and pl_net_profit_cell) else "=0"
    ws.cell(row=r, column=6, value=f"=-({net_profit_ref[1:]})")  # raw equity_current_year: a profit increases (credit) equity, negative in this convention
    profit_cell = ws.cell(row=r, column=2, value=net_profit_ref)
    profit_cell.number_format = CURRENCY_FMT
    r += 1

    total_equity_row = r
    ws.cell(row=r, column=1, value="TOTAL EQUITY").font = BOLD
    ws.cell(row=r, column=6, value=f"=F{equity_bfwd_row}+F{profit_row}")
    te_cell = ws.cell(row=r, column=2, value=f"=-F{r}")
    te_cell.number_format = CURRENCY_FMT
    te_cell.font = BOLD
    r += 1

    check_row = r
    ws.cell(row=r, column=1, value="CHECK: Net Assets - Total Equity (should be £0)").font = BOLD
    check_cell = ws.cell(row=r, column=2, value=f"=B{net_assets_row}+F{total_equity_row}")
    check_cell.number_format = CURRENCY_FMT
    check_cell.font = BOLD
    if bs_result.status != "ok":
        for c in (1, 2):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=AMBER)

    r += 2
    ws.cell(row=r, column=1, value="DETAIL BY ACCOUNT").font = SCHEDULE_FONT
    r += 1
    detail_headers = ["Account Code", "Account Name", "Category", "Amount"]
    for j, h in enumerate(detail_headers):
        ws.cell(row=r, column=1 + j, value=h)
    _style_header_row(ws, r, len(detail_headers))
    first_detail = r + 1
    n = refs.bs_current.last_row - refs.bs_current.first_row + 1
    for i in range(n):
        src = refs.bs_current.first_row + i
        dest = first_detail + i
        for col_idx, field in enumerate(["account_code", "account_name", "category"]):
            ws.cell(row=dest, column=1 + col_idx, value="=" + cell_ref(refs.bs_current.sheet_name, f"{refs.bs_current.columns[field]}{src}")).border = BORDER
        amt_cell = ws.cell(row=dest, column=4, value="=" + cell_ref(refs.bs_current.sheet_name, f"{refs.bs_current.columns['amount']}{src}"))
        amt_cell.number_format = CURRENCY_FMT
        amt_cell.border = BORDER

    ws.freeze_panes = f"A{table_row + 1}"
    return sheet_name


def build_fixed_asset_category_sheet_formulas(
    wb: Workbook, client_name: str, period_label: str, ref: str, result: FixedAssetResult, refs: DataRefs,
    grouped_codes: dict[str, dict[str, list[str]]], header_cells: dict | None = None,
):
    """Formula-linked version of the category-level fixed asset rollforward:
    same cost/depreciation/NBV columns as the Python-computed sheet, but
    every b/fwd, c/fwd, additions and depreciation figure is a live
    SUMPRODUCT formula against DATA_TB_Current/DATA_TB_Comparative/
    DATA_Nominal, OR'd across every account code Python grouped into that
    category (grouped_codes - see fixed_assets.group_fixed_asset_codes).
    Python still decides the category grouping itself (a text-parsing
    problem, not something an Excel formula should attempt), but every
    number in the rollforward recalculates from the raw data."""
    sheet_name = f"{ref} Fixed Asset Register (Category)"[:31]
    ws = wb.create_sheet(sheet_name)
    row = _write_title(ws, client_name, period_label, "FIXED ASSET REGISTER (CATEGORY SUMMARY)", ref, header_cells=header_cells)

    status_cell = ws.cell(row=row, column=1, value=f"Status: {result.status.upper()} - {result.message}")
    status_cell.font = _status_font(result.status)
    status_cell.fill = _status_fill(result.status)
    status_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 30

    if result.detail.empty or refs.tb_current is None:
        ws.cell(row=row + 2, column=1, value="No Fixed Asset accounts found in the trial balance.")
        return sheet_name

    table_row = row + 2
    headers = [
        "Category", "Cost b/fwd", "Additions", "Disposals (cost)", "Cost c/fwd (per TB)", "Cost diff",
        "Acc. depreciation b/fwd", "Depreciation charge", "Depreciation on disposals", "Acc. depreciation c/fwd (per TB)",
        "Depreciation diff", "NBV b/fwd", "NBV c/fwd",
        "System est. method", "System est. rate %", "System est. depreciation", "Booked vs system est. (diff)",
    ]
    for j, h in enumerate(headers):
        ws.cell(row=table_row, column=1 + j, value=h)
    _style_header_row(ws, table_row, len(headers))

    tb_c_bal, tb_c_code = refs.tb_current.col_range("balance"), refs.tb_current.col_range("account_code")
    tb_p_bal, tb_p_code = (refs.tb_comparative.col_range("balance"), refs.tb_comparative.col_range("account_code")) if refs.tb_comparative else (None, None)
    nom_debit, nom_credit, nom_code = (
        (refs.nominal_current.col_range("debit"), refs.nominal_current.col_range("credit"), refs.nominal_current.col_range("account_code"))
        if refs.nominal_current else (None, None, None)
    )

    def or_sum(sum_range, criteria_range, codes: list[str]) -> str:
        if sum_range is None or not codes:
            return "=0"
        return sum_of_values(sum_range, criteria_range, [quote(c) for c in codes])

    r = table_row + 1
    for category in result.detail["Category"]:
        codes = grouped_codes.get(category, {"cost": [], "depreciation": []})
        cost_codes, dep_codes = codes["cost"], codes["depreciation"]
        row_data = result.detail[result.detail["Category"] == category].iloc[0]

        ws.cell(row=r, column=1, value=category).border = BORDER

        b_fwd_cell = ws.cell(row=r, column=2, value=or_sum(tb_p_bal, tb_p_code, cost_codes))
        b_fwd_cell.number_format = CURRENCY_FMT
        add_cell = ws.cell(row=r, column=3, value=or_sum(nom_debit, nom_code, cost_codes))
        add_cell.number_format = CURRENCY_FMT
        disp_cell = ws.cell(row=r, column=4, value=or_sum(nom_credit, nom_code, cost_codes))
        disp_cell.number_format = CURRENCY_FMT
        c_fwd_cell = ws.cell(row=r, column=5, value=or_sum(tb_c_bal, tb_c_code, cost_codes))
        c_fwd_cell.number_format = CURRENCY_FMT
        diff_cell = ws.cell(row=r, column=6, value=f"=(B{r}+C{r}-D{r})-E{r}")
        diff_cell.number_format = CURRENCY_FMT

        dep_b_fwd_raw = or_sum(tb_p_bal, tb_p_code, dep_codes)
        dep_b_fwd_cell = ws.cell(row=r, column=7, value=f"=-({dep_b_fwd_raw[1:]})")
        dep_b_fwd_cell.number_format = CURRENCY_FMT
        dep_charge_cell = ws.cell(row=r, column=8, value=or_sum(nom_credit, nom_code, dep_codes))
        dep_charge_cell.number_format = CURRENCY_FMT
        dep_disp_cell = ws.cell(row=r, column=9, value=or_sum(nom_debit, nom_code, dep_codes))
        dep_disp_cell.number_format = CURRENCY_FMT
        dep_c_fwd_raw = or_sum(tb_c_bal, tb_c_code, dep_codes)
        dep_c_fwd_cell = ws.cell(row=r, column=10, value=f"=-({dep_c_fwd_raw[1:]})")
        dep_c_fwd_cell.number_format = CURRENCY_FMT
        dep_diff_cell = ws.cell(row=r, column=11, value=f"=(G{r}+H{r}-I{r})-J{r}")
        dep_diff_cell.number_format = CURRENCY_FMT

        nbv_b_fwd_cell = ws.cell(row=r, column=12, value=f"=B{r}-G{r}")
        nbv_b_fwd_cell.number_format = CURRENCY_FMT
        nbv_c_fwd_cell = ws.cell(row=r, column=13, value=f"=E{r}-J{r}")
        nbv_c_fwd_cell.number_format = CURRENCY_FMT

        # System-estimate columns are Python-side heuristic values (a
        # default rate inferred from the category name - see
        # fixed_assets.DEFAULT_CATEGORY_DEPRECIATION), not raw-data
        # recalculations, so unlike every other column here they're
        # written as plain values rather than SUMPRODUCT formulas.
        ws.cell(row=r, column=14, value=row_data["System est. method"]).border = BORDER
        rate_cell = ws.cell(row=r, column=15, value=row_data["System est. rate %"])
        rate_cell.border = BORDER
        est_dep_cell = ws.cell(row=r, column=16, value=row_data["System est. depreciation"])
        est_dep_cell.number_format = CURRENCY_FMT
        est_diff_cell = ws.cell(row=r, column=17, value=row_data["Booked vs system est. (diff)"])
        est_diff_cell.number_format = CURRENCY_FMT

        flagged = abs(row_data["Cost diff"]) > MATERIALITY_AMOUNT or abs(row_data["Depreciation diff"]) > MATERIALITY_AMOUNT
        if flagged:
            for c in (6, 11):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=AMBER)
        r += 1

    next_row = table_row + len(result.detail) + 2
    if result.extra_detail is not None and not result.extra_detail.empty:
        ws.cell(row=next_row, column=1, value=result.extra_detail_label.upper()).font = SCHEDULE_FONT
        _write_dataframe(ws, result.extra_detail, start_row=next_row + 1)

    ws.freeze_panes = f"A{table_row + 1}"
    return sheet_name


def build_matrix_sheet(wb: Workbook, client_name: str, current_label: str, ref: str, result: MatrixResult, header_cells: dict | None = None):
    sheet_name = f"{ref} {result.account_name} Analysis"[:31]
    ws = wb.create_sheet(sheet_name)
    title = f"NOMINAL ACTIVITY ANALYSIS - {result.account_name} ({result.account_code})"
    row = _write_title(ws, client_name, current_label, title, ref, header_cells=header_cells)

    status_cell = ws.cell(row=row, column=1, value=f"Status: {result.status.upper()} - {result.message}")
    status_cell.font = _status_font(result.status)
    status_cell.fill = _status_fill(result.status)
    status_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 30

    table_row = row + 2
    if result.matrix is None or result.matrix.empty:
        ws.cell(row=table_row, column=1, value="No data available.")
        return
    df = result.matrix.rename(columns={"date": "Date", "reference": "Reference", "description": "Description", "contact": "Contact"})
    _write_dataframe(ws, df, start_row=table_row)
    ws.freeze_panes = f"A{table_row + 1}"


def build_matrix_sheet_formulas(
    wb: Workbook, client_name: str, current_label: str, ref: str, result: MatrixResult, refs: DataRefs,
    nominal_current: pd.DataFrame, header_cells: dict | None = None,
):
    """Formula-linked version of the nominal activity matrix: same
    (date/reference/description/contact) x contra-account pivot, but every
    cell is a live SUMPRODUCT re-summing the DATA_Nominal 'net' column for
    exactly the row_ids Python bucketed into that cell - see
    nominal_matrix.build_matrix_row_groups(). Python still decides the
    bucketing itself (which contra account each transaction belongs to, and
    which accounts make the top-N-by-value cut vs fold into OTHER - a
    text/ranking decision, not something an Excel formula should attempt),
    but the amount in every cell recalculates from the raw data, and the
    row/column totals and the DIFF self-check are formulas too."""
    sheet_name = f"{ref} {result.account_name} Analysis"[:31]
    ws = wb.create_sheet(sheet_name)
    title = f"NOMINAL ACTIVITY ANALYSIS - {result.account_name} ({result.account_code})"
    row = _write_title(ws, client_name, current_label, title, ref, header_cells=header_cells)

    status_cell = ws.cell(row=row, column=1, value=f"Status: {result.status.upper()} - {result.message}")
    status_cell.font = _status_font(result.status)
    status_cell.fill = _status_fill(result.status)
    status_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 30

    table_row = row + 2
    if refs.nominal_current is None:
        ws.cell(row=table_row, column=1, value="No data available.")
        return sheet_name

    pivot_rows, col_names = build_matrix_row_groups(result.account_code, with_row_ids(nominal_current))
    if not pivot_rows:
        ws.cell(row=table_row, column=1, value="No data available.")
        return sheet_name

    headers = ["Date", "Reference", "Description", "Contact", *col_names, "TOTAL", "DIFF"]
    for j, h in enumerate(headers):
        ws.cell(row=table_row, column=1 + j, value=h)
    _style_header_row(ws, table_row, len(headers))

    net_range = refs.nominal_current.col_range("net")
    row_id_range = refs.nominal_current.col_range("row_id")
    first_col = 5  # after Date/Reference/Description/Contact
    total_col = first_col + len(col_names)
    diff_col = total_col + 1

    r = table_row + 1
    for pivot_row in pivot_rows:
        ws.cell(row=r, column=1, value=pivot_row["date"]).border = BORDER
        ws.cell(row=r, column=2, value=pivot_row["reference"]).border = BORDER
        ws.cell(row=r, column=3, value=pivot_row["description"]).border = BORDER
        ws.cell(row=r, column=4, value=pivot_row["contact"]).border = BORDER

        all_row_ids: list = []
        for j, col in enumerate(col_names):
            row_ids = pivot_row["row_ids_by_column"].get(col, [])
            all_row_ids.extend(row_ids)
            formula = sum_of_values(net_range, row_id_range, [literal(rid) for rid in row_ids]) if row_ids else "=0"
            cell = ws.cell(row=r, column=first_col + j, value=formula)
            cell.number_format = CURRENCY_FMT
            cell.border = BORDER

        total_col_letters = get_column_letter(first_col), get_column_letter(total_col - 1)
        total_cell = ws.cell(row=r, column=total_col, value=f"=SUM({total_col_letters[0]}{r}:{total_col_letters[1]}{r})")
        total_cell.number_format = CURRENCY_FMT
        total_cell.font = BOLD

        actual_formula = sum_of_values(net_range, row_id_range, [literal(rid) for rid in all_row_ids]) if all_row_ids else "=0"
        diff_cell = ws.cell(row=r, column=diff_col, value=f"={get_column_letter(total_col)}{r}-({actual_formula[1:]})")
        diff_cell.number_format = CURRENCY_FMT
        r += 1

    ws.freeze_panes = f"A{table_row + 1}"
    return sheet_name


def build_asset_register_sheet(wb: Workbook, client_name: str, current_label: str, ref: str, result: AssetRegisterResult, header_cells: dict | None = None):
    ws = wb.create_sheet(f"{ref} Fixed Asset Register"[:31])
    row = _write_title(ws, client_name, current_label, "FIXED ASSET REGISTER (ASSET DETAIL)", ref, header_cells=header_cells)

    status_cell = ws.cell(row=row, column=1, value=f"Status: {result.status.upper()} - {result.message}")
    status_cell.font = _status_font(result.status)
    status_cell.fill = _status_fill(result.status)
    status_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 30

    r = row + 2
    if not result.summary.empty:
        r = _write_dataframe(ws, result.summary, start_row=r)

    if not result.asset_schedule.empty:
        ws.cell(row=r, column=1, value="ASSET SCHEDULE").font = SCHEDULE_FONT
        r = _write_dataframe(ws, result.asset_schedule, start_row=r + 1)

    if not result.new_additions.empty:
        ws.cell(row=r, column=1, value="NEW ADDITIONS FOUND IN NOMINAL ACTIVITY - NOT YET IN REGISTER").font = SCHEDULE_FONT
        r = _write_dataframe(ws, result.new_additions, start_row=r + 1)

    if not result.possible_disposals.empty:
        ws.cell(row=r, column=1, value="POSSIBLE DISPOSALS (CREDIT MOVEMENTS ON FIXED ASSET COST CODES)").font = SCHEDULE_FONT
        r = _write_dataframe(ws, result.possible_disposals, start_row=r + 1)

    if result.summary.empty and result.asset_schedule.empty:
        ws.cell(row=r, column=1, value="No data available.")
    ws.freeze_panes = f"A{row + 1}"


def build_closing_register_sheet(wb: Workbook, client_name: str, current_label: str, ref: str, closing_register: pd.DataFrame, header_cells: dict | None = None):
    """Same presentation/column layout as the prior-year register upload,
    so this sheet's contents can be taken straight into next year's job as
    its opening register - no reformatting needed."""
    ws = wb.create_sheet(f"{ref} FA Register (closing)"[:31])
    row = _write_title(ws, client_name, current_label, "FIXED ASSET REGISTER - CLOSING POSITION (FOR NEXT YEAR'S OPENING UPLOAD)", ref, header_cells=header_cells)
    if closing_register.empty:
        ws.cell(row=row, column=1, value="No data available.")
        return
    _write_dataframe(ws, closing_register, start_row=row)
    ws.freeze_panes = f"A{row + 1}"


def build_corporation_tax_sheet(wb: Workbook, client_name: str, current_label: str, ref: str, ct: CTComputation, header_cells: dict | None = None):
    ws = wb.create_sheet(f"{ref} Corporation Tax"[:31])
    row = _write_title(ws, client_name, current_label, "CORPORATION TAX COMPUTATION", ref, header_cells=header_cells)

    status_cell = ws.cell(row=row, column=1, value=f"Status: {ct.status.upper()} - {ct.message}")
    status_cell.font = _status_font(ct.status)
    status_cell.fill = _status_fill(ct.status)
    status_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 30

    r = row + 2
    rates = ct.rates_used
    lines = [
        ("Profit per accounts", ct.accounting_profit, False),
        ("Add: disallowable expenses (preparer input)", ct.disallowable_additions, False),
        ("Less: capital allowances (preparer input)", -ct.capital_allowances, False),
        ("TAXABLE TOTAL PROFITS", ct.taxable_profit, True),
        ("Augmented profits (for rate banding)", ct.augmented_profits, False),
        ("Associated companies", ct.associated_companies, False),
        ("Accounting period (days)", ct.period_days, False),
        ("Small profits rate threshold (scaled)", ct.lower_limit, False),
        ("Main rate threshold (scaled)", ct.upper_limit, False),
        ("Rate band applied", ct.band, False),
        ("Marginal relief", ct.marginal_relief, False),
        ("CORPORATION TAX CHARGE (computed)", ct.tax_charge, True),
        ("Tax charge per accounts (TB)", ct.booked_tax_charge if ct.booked_tax_charge is not None else "not found in TB", False),
        ("Variance", ct.variance if ct.variance is not None else "", False),
    ]
    for label, value, bold in lines:
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=value)
        c1.border = BORDER
        c2.border = BORDER
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            c2.number_format = CURRENCY_FMT
        if bold:
            c1.font = BOLD
            c2.font = BOLD
        r += 1

    r += 1
    ws.cell(row=r, column=1, value=f"Rates used: small profits {rates.small_profits_rate:.0%}, main rate {rates.main_rate:.0%}, "
                                    f"thresholds £{rates.lower_limit:,.0f}/£{rates.upper_limit:,.0f}, marginal relief fraction {rates.marginal_relief_fraction:.4f}").font = SUBTITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value=f"Rates last verified {rates.as_at} against {rates.source} - confirm current before relying on this for a filing.").font = SUBTITLE_FONT

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20


PERCENT_FMT = "0.00%"


def build_corporation_tax_sheet_formulas(
    wb: Workbook, client_name: str, current_label: str, ref: str, ct: CTComputation,
    pl_sheet_name: str | None = None, pl_net_profit_cell: str | None = None, header_cells: dict | None = None,
) -> str:
    """Same proforma as build_corporation_tax_sheet, but every computed line
    (taxable profit, the scaled thresholds, the rate band, marginal relief,
    the tax charge, the variance) is a live formula, and the rates that
    drive them (small profits rate, main rate, thresholds, marginal relief
    fraction) are written as their own cells lower on the sheet and
    referenced by formula rather than baked into the computation as
    literals - so re-pointing this sheet at next year's rates is a matter
    of editing those cells, not regenerating the workbook.

    "Profit per accounts" is a live cross-sheet reference to the P&L
    formula sheet's own NET PROFIT cell when one is supplied, continuing
    the same cross-sheet linking used for the Balance Sheet.

    Simplification: "Augmented profits" always formula-links to the taxable
    profit cell on this sheet (the common case - no associated companies'
    profits to add in). Overriding augmented profits separately from
    taxable profit is a rarely-used HMRC provision not yet surfaced in the
    app's inputs (see README known limitations), so it isn't represented
    as an independent formula-editable cell here."""
    sheet_name = f"{ref} Corporation Tax"[:31]
    ws = wb.create_sheet(sheet_name)
    row = _write_title(ws, client_name, current_label, "CORPORATION TAX COMPUTATION", ref, header_cells=header_cells)

    status_cell = ws.cell(row=row, column=1, value=f"Status: {ct.status.upper()} - {ct.message}")
    status_cell.font = _status_font(ct.status)
    status_cell.fill = _status_fill(ct.status)
    status_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 30

    rates = ct.rates_used
    r = row + 2

    profit_row = r
    ws.cell(row=r, column=1, value="Profit per accounts").border = BORDER
    if pl_sheet_name and pl_net_profit_cell:
        profit_cell = ws.cell(row=r, column=2, value="=" + cell_ref(pl_sheet_name, pl_net_profit_cell))
    else:
        profit_cell = ws.cell(row=r, column=2, value=ct.accounting_profit)
    profit_cell.number_format = CURRENCY_FMT
    profit_cell.border = BORDER
    r += 1

    disallow_row = r
    ws.cell(row=r, column=1, value="Add: disallowable expenses (preparer input)").border = BORDER
    disallow_cell = ws.cell(row=r, column=2, value=ct.disallowable_additions)
    disallow_cell.number_format = CURRENCY_FMT
    disallow_cell.border = BORDER
    r += 1

    capex_row = r
    ws.cell(row=r, column=1, value="Less: capital allowances (preparer input)").border = BORDER
    ws.cell(row=r, column=6, value=ct.capital_allowances)  # hidden raw (positive) figure
    capex_cell = ws.cell(row=r, column=2, value=f"=-F{r}")
    capex_cell.number_format = CURRENCY_FMT
    capex_cell.border = BORDER
    r += 1

    taxable_row = r
    ws.cell(row=r, column=1, value="TAXABLE TOTAL PROFITS").font = BOLD
    taxable_cell = ws.cell(row=r, column=2, value=f"=B{profit_row}+B{disallow_row}+B{capex_row}")
    taxable_cell.number_format = CURRENCY_FMT
    taxable_cell.font = BOLD
    r += 1

    augmented_row = r
    ws.cell(row=r, column=1, value="Augmented profits (for rate banding)").border = BORDER
    augmented_cell = ws.cell(row=r, column=2, value=f"=B{taxable_row}")
    augmented_cell.number_format = CURRENCY_FMT
    augmented_cell.border = BORDER
    r += 1

    assoc_row = r
    ws.cell(row=r, column=1, value="Associated companies").border = BORDER
    ws.cell(row=r, column=2, value=ct.associated_companies).border = BORDER
    r += 1

    period_row = r
    ws.cell(row=r, column=1, value="Accounting period (days)").border = BORDER
    ws.cell(row=r, column=2, value=ct.period_days).border = BORDER
    r += 1

    lower_row = r
    ws.cell(row=r, column=1, value="Small profits rate threshold (scaled)").border = BORDER
    r += 1

    upper_row = r
    ws.cell(row=r, column=1, value="Main rate threshold (scaled)").border = BORDER
    r += 1

    band_row = r
    ws.cell(row=r, column=1, value="Rate band applied").border = BORDER
    r += 1

    relief_row = r
    ws.cell(row=r, column=1, value="Marginal relief").border = BORDER
    r += 1

    charge_row = r
    ws.cell(row=r, column=1, value="CORPORATION TAX CHARGE (computed)").font = BOLD
    r += 1

    booked_row = r
    ws.cell(row=r, column=1, value="Tax charge per accounts (TB)").border = BORDER
    if ct.booked_tax_charge is not None:
        booked_cell = ws.cell(row=r, column=2, value=ct.booked_tax_charge)
        booked_cell.number_format = CURRENCY_FMT
    else:
        ws.cell(row=r, column=2, value="not found in TB")
    ws.cell(row=r, column=2).border = BORDER
    r += 1

    variance_row = r
    ws.cell(row=r, column=1, value="Variance").border = BORDER
    r += 2

    ws.cell(row=r, column=1, value="RATES USED (edit here to re-point this sheet at a future year's rates)").font = SCHEDULE_FONT
    r += 1
    small_rate_row = r
    ws.cell(row=r, column=1, value="Small profits rate").border = BORDER
    small_rate_cell = ws.cell(row=r, column=2, value=rates.small_profits_rate)
    small_rate_cell.number_format = PERCENT_FMT
    small_rate_cell.border = BORDER
    r += 1
    main_rate_row = r
    ws.cell(row=r, column=1, value="Main rate").border = BORDER
    main_rate_cell = ws.cell(row=r, column=2, value=rates.main_rate)
    main_rate_cell.number_format = PERCENT_FMT
    main_rate_cell.border = BORDER
    r += 1
    lower_const_row = r
    ws.cell(row=r, column=1, value="Small profits rate threshold (unscaled)").border = BORDER
    lower_const_cell = ws.cell(row=r, column=2, value=rates.lower_limit)
    lower_const_cell.number_format = CURRENCY_FMT
    lower_const_cell.border = BORDER
    r += 1
    upper_const_row = r
    ws.cell(row=r, column=1, value="Main rate threshold (unscaled)").border = BORDER
    upper_const_cell = ws.cell(row=r, column=2, value=rates.upper_limit)
    upper_const_cell.number_format = CURRENCY_FMT
    upper_const_cell.border = BORDER
    r += 1
    fraction_row = r
    ws.cell(row=r, column=1, value="Marginal relief fraction").border = BORDER
    fraction_cell = ws.cell(row=r, column=2, value=rates.marginal_relief_fraction)
    fraction_cell.border = BORDER
    r += 1
    ws.cell(row=r, column=1, value=f"Rates last verified {rates.as_at} against {rates.source} - confirm current before relying on this for a filing.").font = SUBTITLE_FONT

    scale = f"(1+B{assoc_row})"
    period_fraction = f"MIN(1,MAX(0,B{period_row}/365))"
    lower_cell = ws.cell(row=lower_row, column=2, value=f"=ROUND(B{lower_const_row}/{scale}*{period_fraction},2)")
    lower_cell.number_format = CURRENCY_FMT
    lower_cell.border = BORDER
    upper_cell = ws.cell(row=upper_row, column=2, value=f"=ROUND(B{upper_const_row}/{scale}*{period_fraction},2)")
    upper_cell.number_format = CURRENCY_FMT
    upper_cell.border = BORDER

    taxable, augmented, lower, upper = f"B{taxable_row}", f"B{augmented_row}", f"B{lower_row}", f"B{upper_row}"
    small_rate, main_rate, fraction = f"B{small_rate_row}", f"B{main_rate_row}", f"B{fraction_row}"

    band_cell = ws.cell(row=band_row, column=2, value=(
        f'=IF({taxable}<=0,"no tax due (loss/nil profit)",'
        f'IF({augmented}<={lower},"small profits rate",'
        f'IF({augmented}>={upper},"main rate","marginal relief")))'
    ))
    band_cell.border = BORDER

    relief_cell = ws.cell(row=relief_row, column=2, value=(
        f'=IF(AND({taxable}>0,{augmented}>{lower},{augmented}<{upper}),'
        f'({upper}-{augmented})*({taxable}/{augmented})*{fraction},0)'
    ))
    relief_cell.number_format = CURRENCY_FMT
    relief_cell.border = BORDER

    charge_cell = ws.cell(row=charge_row, column=2, value=(
        f'=ROUND(IF({taxable}<=0,0,'
        f'IF({augmented}<={lower},{taxable}*{small_rate},'
        f'IF({augmented}>={upper},{taxable}*{main_rate},'
        f'{taxable}*{main_rate}-B{relief_row}))),2)'
    ))
    charge_cell.number_format = CURRENCY_FMT
    charge_cell.font = BOLD

    if ct.booked_tax_charge is not None:
        variance_cell = ws.cell(row=variance_row, column=2, value=f"=B{charge_row}-B{booked_row}")
        variance_cell.number_format = CURRENCY_FMT
        variance_cell.border = BORDER
        if ct.status != "ok":
            for c in (1, 2):
                ws.cell(row=variance_row, column=c).fill = PatternFill("solid", fgColor=AMBER)
    else:
        ws.cell(row=variance_row, column=2, value="").border = BORDER

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["F"].hidden = True
    return sheet_name


def build_points_forward_sheet(wb: Workbook, client_name: str, period_label: str, ref: str, header_cells: dict | None = None):
    ws = wb.create_sheet(f"{ref} Points Forward"[:31])
    row = _write_title(ws, client_name, period_label, "POINTS FORWARD / OPEN QUERIES", ref, header_cells=header_cells)
    headers = ["#", "Area", "Query", "Raised by", "Date raised", "Response / Resolution", "Status"]
    for j, h in enumerate(headers):
        ws.cell(row=row, column=1 + j, value=h)
    _style_header_row(ws, row, len(headers))
    for r in range(row + 1, row + 21):
        ws.cell(row=r, column=1, value=r - row)
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
    widths = [4, 18, 45, 14, 14, 45, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    ws.freeze_panes = f"A{row + 1}"


# The confirmation-only half of a real manual-job review checklist: items
# that need a preparer to check something outside the data this system
# ingests (an agreement, a bank statement, an HMRC login) rather than
# something a data-driven check can flag on its own - see
# compliance_checks.py for the automatable subset (DLA monthly threshold +
# S455, dividends vs reserves, petty cash running balance, loan/BBL/HP
# presence), which is deliberately NOT repeated here to avoid the same
# point appearing twice in different places in the pack.
COMPLIANCE_CHECKLIST_ITEMS = [
    ("Fixed Assets", "Depreciation method and rate applied consistently with last year for brought-forward assets."),
    ("Fixed Assets", "Current year additions have a depreciation method/rate assigned (see the Fixed Asset Register schedule for additions found in nominal activity but not yet in the register)."),
    ("Fixed Assets", "Any disposals in the year identified and the resulting profit/loss on disposal calculated."),
    ("Bank", "Statement received and matches the closing balance for every bank account held during the year, including any account opened or closed mid-year."),
    ("Bank", "Any unrecorded bank receipts or payments (per statement but not per the books) identified and posted."),
    ("Stock", "Closing stock figure confirmed - by count, or by an estimation basis (e.g. GP ratio) agreed and documented."),
    ("Trade Debtors", "Sales invoices for the year agree to sales per the VAT working; any missing sales journal identified."),
    ("Trade Debtors", "Unmatched sales (invoiced but not receipted) agreed as the closing debtors figure; opening debtors traced to receipt."),
    ("Trade Debtors", "For Xero jobs: outstanding debtor receipts after the year end checked for correct allocation to P&L vs balance sheet, with the receipt date noted on the debtors schedule."),
    ("Prepayments", "Prepayment basis confirmed (per invoice or per last year), and any accountancy fee prepayment correctly reversed/re-booked for the current year."),
    ("Other Debtors", "Opening balance (if any) traced and cleared; s455 note prepared if any balance relates to a director."),
    ("Petty Cash", "Closing petty cash balance agreed to the count/last year's basis; any deficit covered through the DLA rather than left as a negative cash balance."),
    ("Trade Creditors", "Purchases per the VAT working agree to purchase invoices; unmatched purchases agreed as the closing creditors figure."),
    ("Trade Creditors", "Any expense assumed paid in cash, or paid through the DLA, checked against last year's treatment."),
    ("Accruals", "Opening accruals reversed and the current year's accountancy/other accrual booked and reconciled to the source ledger."),
    ("VAT", "VAT scheme confirmed (standard/cash/flat rate; registered or deregistered in year) and consistent with the return filed."),
    ("VAT", "Input VAT correctly claimed/reversed on expenses where relevant (e.g. entertainment, personal use)."),
    ("PAYE / Wages", "Payroll figures received for the year and agreed to the PAYE control/wages control account; employment allowance claimed correctly if more than one director."),
    ("PAYE / Wages", "Any subcontractor payments (CIS) identified, suffered CIS reconciled to HMRC."),
    ("Pension", "Pension contributions for the year received and the closing liability agreed to the pension provider/payment made after the year end."),
    ("Loans / HP / Credit Card", "Agreement or statement received for every loan/HP/credit card facility; closing balance split between amounts due within one year and after one year."),
    ("Loans / HP / Credit Card", "Interest for the year booked correctly (per statement, or per calculator where no statement is available)."),
    ("Directors' Loan Account", "DLA transaction-by-transaction detail agrees to the movements schedule; any capital introduced or drawings for the year identified."),
    ("Directors' Loan Account", "Use-of-home-as-office and any other expenses due to be booked through the DLA (per last year's treatment) have been included."),
    ("Dividends", "Dividend minutes/vouchers in place; dividend correctly apportioned to the tax year and rounded appropriately; split according to shareholdings."),
    ("Government Grants", "Any government grant (e.g. SEISS for a sole trader) shown correctly - in the capital account for a sole trader, separately (and in the CT600) for a limited company."),
    ("Corporation Tax", "CT liability agreed to the HMRC online account; loss brought forward/carried back correctly offset or claimed; refund due tracked to the correct bank account."),
    ("Corporation Tax", "CT600 box 775 (expenditure on machinery/plant not a long-life asset or integral feature, excluding anything in box 760) reviewed and completed where relevant."),
]


def build_compliance_checklist_sheet(wb: Workbook, client_name: str, period_label: str, ref: str, header_cells: dict | None = None):
    """A static, editable pro-forma of the confirmation-only checklist
    items - not computed from data (there's nothing in a TB or nominal
    activity export that answers "was the HP agreement received?"), so
    this is a structured place for the preparer to work through and tick
    off, the same role the real firm's own manual-job checklist plays."""
    ws = wb.create_sheet(f"{ref} Compliance Checklist"[:31])
    row = _write_title(ws, client_name, period_label, "COMPLIANCE CHECKLIST (MANUAL REVIEW POINTS)", ref, header_cells=header_cells)
    ws.cell(row=row, column=1, value=(
        "Confirmation-only items that need a document, statement, or external check rather than something "
        "the uploaded data can answer on its own. Data-driven checks for DLA/S455, dividends vs reserves, "
        "petty cash, and loan/BBL/HP presence are on their own schedules - see the Index."
    )).font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 30

    table_row = row + 2
    headers = ["Area", "Checklist item", "Status (Yes / No / N/A)", "Notes"]
    for j, h in enumerate(headers):
        ws.cell(row=table_row, column=1 + j, value=h)
    _style_header_row(ws, table_row, len(headers))

    r = table_row + 1
    for area, item in COMPLIANCE_CHECKLIST_ITEMS:
        ws.cell(row=r, column=1, value=area).border = BORDER
        ws.cell(row=r, column=2, value=item).border = BORDER
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=3).border = BORDER
        ws.cell(row=r, column=4).border = BORDER
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
        r += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 75
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 45
    ws.freeze_panes = f"A{table_row + 1}"


def _generate_schedules(
    wb: Workbook,
    ref: _RefCounter,
    index_ws: Worksheet,
    client_name: str,
    current_label: str,
    comparative_label: str,
    data: dict,
    results: list[ReconResult],
    control_account_results: list[ControlAccountResult],
    matrix_results: list[MatrixResult],
    ct_computation: CTComputation | None,
    fixed_asset_result: FixedAssetResult | None,
    asset_register_result: AssetRegisterResult | None,
    enabled=lambda key: True,
    place=lambda key, call: call(),
    header_cells: dict | None = None,
) -> Workbook:
    """The shared core both build_workbook (generic layout, a fresh
    workbook) and build_workbook_into_template (a practice's real
    uploaded template as the base) drive - every schedule this system
    knows how to build, in one place, so the two paths can't quietly drift
    apart on what gets generated.

    `enabled(key)` gates whether a schedule is generated at all (always
    True for the generic path; template config-driven for the template
    path). `place(key, call)` invokes a builder call and hands back its
    return value unchanged - the template path additionally repositions
    whichever new sheet the call created, per that schedule's
    insert_after_sheet config; the generic path just calls straight
    through, since a freshly generated workbook has no fixed layout to
    respect."""
    entries = []

    tb_on = enabled("tb_lead_schedule")
    tb_ref = ref.next() if tb_on else None
    if tb_on:
        entries.append({"ref": tb_ref, "title": "TB Lead Schedule", "status": "ok" if data.get("tb_current") is not None and not data["tb_current"].empty else "n/a", "message": "Current vs comparative, variance flagged"})

    pl_statement = build_pl_statement(data.get("pl_current"))
    bs_statement = build_bs_statement(data.get("bs_current"), pl_statement.net_profit)

    pl_on, bs_on = enabled("profit_and_loss"), enabled("balance_sheet")
    pl_ref = ref.next() if pl_on else None
    bs_ref = ref.next() if bs_on else None
    if pl_on:
        entries.append({"ref": pl_ref, "title": "Profit & Loss", "status": pl_statement.status, "message": pl_statement.message if pl_statement.status != "ok" else ""})
    if bs_on:
        entries.append({"ref": bs_ref, "title": "Balance Sheet", "status": bs_statement.status, "message": bs_statement.message if bs_statement.status != "n/a" else ""})

    ct_ref = None
    if ct_computation is not None and enabled("corporation_tax"):
        ct_ref = ref.next()
        entries.append({"ref": ct_ref, "title": "Corporation Tax Computation", "status": ct_computation.status, "message": ct_computation.message})

    fa_ref = None
    if fixed_asset_result is not None and fixed_asset_result.status != "n/a" and enabled("fixed_asset_category"):
        fa_ref = ref.next()
        entries.append({"ref": fa_ref, "title": fixed_asset_result.name, "status": fixed_asset_result.status, "message": fixed_asset_result.message})

    ar_ref = None
    cr_ref = None
    if asset_register_result is not None and asset_register_result.status != "n/a" and enabled("fixed_asset_register"):
        ar_ref = ref.next()
        entries.append({"ref": ar_ref, "title": "Fixed asset register (asset detail)", "status": asset_register_result.status, "message": asset_register_result.message})
        if not asset_register_result.closing_register.empty:
            cr_ref = ref.next()
            entries.append({"ref": cr_ref, "title": "Fixed asset register (closing - for next year)", "status": "n/a", "message": "Same layout as the prior-year upload - carry this straight into next year's job."})

    # (config key, sheet title) per recon/anomaly result name - None title
    # means "feeds another schedule, no separate tab of its own" (unrelated
    # to enabled()); a name absent from this map falls back to using its
    # own name as both the config key and (truncated) sheet title, so a
    # future new check works out of the box without a config change
    RESULT_SCHEDULE_INFO = {
        "TB self-balance check": ("tb_balance_check", "TB Balance Check"),
        "Current vs comparative variance analysis": (None, None),
        "Debtors control account reconciliation": ("debtors_recon", "Debtors Recon"),
        "Creditors control account reconciliation": ("creditors_recon", "Creditors Recon"),
        "Bank reconciliation": ("bank_recon", "Bank Recon"),
        "VAT return cross-check": ("vat_recon", "VAT Recon"),
        "Nominal activity review (reallocation candidates)": ("nominal_review", "Nominal Review"),
        "Contact coding consistency (possible miscoding)": ("contact_coding_consistency", "Contact Coding Check"),
        "Duplicate transaction check": ("duplicate_check", "Duplicate Check"),
        "Unusual posting date check (weekend manual journals)": ("unusual_posting_dates", "Weekend Journals"),
        "Directors' loan account review (S455 / £10,000 monthly threshold)": ("dla_review", "DLA Review"),
        "Dividend vs distributable reserves review": ("dividend_reserves_review", "Dividend Review"),
        "Petty cash running balance review": ("petty_cash_review", "Petty Cash Review"),
        "Loan facility review (BBL / Hire Purchase / Bank Loan)": ("loan_facility_review", "Loan Facility Review"),
    }
    recon_refs = {}
    recon_sheet_names = {}
    for res in results:
        config_key, sheet_title = RESULT_SCHEDULE_INFO.get(res.name, (res.name, res.name[:31]))
        if sheet_title is None or not enabled(config_key):
            continue
        recon_refs[res.name] = ref.next()
        recon_sheet_names[res.name] = sheet_title
        entries.append({"ref": recon_refs[res.name], "title": res.name, "status": res.status, "message": res.message})

    ca_on = enabled("control_account_rollforward")
    ca_refs = {}
    if ca_on:
        for r in control_account_results:
            ca_refs[r.account_code] = ref.next()
            entries.append({"ref": ca_refs[r.account_code], "title": f"{r.account_name} control account", "status": r.status, "message": r.message})

    mx_on = enabled("nominal_matrix")
    mx_refs = {}
    if mx_on:
        for r in matrix_results:
            mx_refs[r.account_code] = ref.next()
            entries.append({"ref": mx_refs[r.account_code], "title": f"{r.account_name} nominal analysis", "status": r.status, "message": r.message})

    cl_on = enabled("compliance_checklist")
    cl_ref = ref.next() if cl_on else None
    if cl_on:
        entries.append({"ref": cl_ref, "title": "Compliance Checklist (manual review points)", "status": "n/a", "message": "To be completed by preparer - see the data-driven checks above for the automatable subset"})

    pf_on = enabled("points_forward")
    pf_ref = ref.next() if pf_on else None
    if pf_on:
        entries.append({"ref": pf_ref, "title": "Points Forward / Open Queries", "status": "n/a", "message": "To be completed by preparer/reviewer"})

    build_index_sheet(index_ws, client_name, current_label, comparative_label, entries)

    # raw-data sheets every formula-linked schedule below references, so the
    # workbook recalculates like a manually-built working paper rather than
    # holding Python-computed literals - see data_sheets.py / xlformulas.py
    refs = write_data_sheets(wb, data)

    variance_result = next((r for r in results if r.name == "Current vs comparative variance analysis"), None)
    if tb_on:
        if refs.tb_current is not None:
            place("tb_lead_schedule", lambda: build_tb_lead_schedule_formulas(wb, client_name, current_label, tb_ref, variance_result.detail if variance_result else pd.DataFrame(), refs, header_cells=header_cells))
        else:
            place("tb_lead_schedule", lambda: build_tb_lead_schedule(wb, client_name, current_label, tb_ref, variance_result.detail if variance_result else pd.DataFrame(), header_cells=header_cells))

    pl_sheet_name, pl_net_profit_cell = (None, None)
    if pl_on:
        pl_sheet_name, pl_net_profit_cell = place("profit_and_loss", lambda: build_pl_statement_sheet_formulas(wb, client_name, current_label, pl_ref, pl_statement, refs, header_cells=header_cells))
    if bs_on:
        place("balance_sheet", lambda: build_bs_statement_sheet_formulas(wb, client_name, current_label, bs_ref, bs_statement, refs, pl_sheet_name, pl_net_profit_cell, header_cells=header_cells))

    if ct_ref is not None:
        place("corporation_tax", lambda: build_corporation_tax_sheet_formulas(wb, client_name, current_label, ct_ref, ct_computation, pl_sheet_name, pl_net_profit_cell, header_cells=header_cells))

    if fa_ref is not None:
        grouped_codes = group_fixed_asset_codes(data.get("tb_current"))
        if grouped_codes and refs.tb_current is not None:
            place("fixed_asset_category", lambda: build_fixed_asset_category_sheet_formulas(wb, client_name, current_label, fa_ref, fixed_asset_result, refs, grouped_codes, header_cells=header_cells))
        else:
            place("fixed_asset_category", lambda: build_recon_sheet(wb, client_name, current_label, fa_ref, f"{fa_ref} Fixed Assets", fixed_asset_result, header_cells=header_cells))

    if ar_ref is not None:
        place("fixed_asset_register", lambda: build_asset_register_sheet(wb, client_name, current_label, ar_ref, asset_register_result, header_cells=header_cells))

    if cr_ref is not None:
        place("fixed_asset_register", lambda: build_closing_register_sheet(wb, client_name, current_label, cr_ref, asset_register_result.closing_register, header_cells=header_cells))

    for res in results:
        if res.name not in recon_refs:
            continue
        config_key, _ = RESULT_SCHEDULE_INFO.get(res.name, (res.name, res.name[:31]))
        r = recon_refs[res.name]
        sheet_title = recon_sheet_names[res.name]
        place(config_key, lambda r=r, sheet_title=sheet_title, res=res: build_recon_sheet(wb, client_name, current_label, r, f"{r} {sheet_title}", res, header_cells=header_cells))

    if ca_on:
        for r in control_account_results:
            if refs.tb_current is not None:
                place("control_account_rollforward", lambda r=r: build_control_account_sheet_formulas(wb, client_name, current_label, ca_refs[r.account_code], r, refs, header_cells=header_cells))
            else:
                place("control_account_rollforward", lambda r=r: build_control_account_sheet(wb, client_name, current_label, ca_refs[r.account_code], r, header_cells=header_cells))

    if mx_on:
        for r in matrix_results:
            if refs.nominal_current is not None:
                place("nominal_matrix", lambda r=r: build_matrix_sheet_formulas(wb, client_name, current_label, mx_refs[r.account_code], r, refs, data.get("nominal_current"), header_cells=header_cells))
            else:
                place("nominal_matrix", lambda r=r: build_matrix_sheet(wb, client_name, current_label, mx_refs[r.account_code], r, header_cells=header_cells))

    if cl_on:
        place("compliance_checklist", lambda: build_compliance_checklist_sheet(wb, client_name, current_label, cl_ref, header_cells=header_cells))

    if pf_on:
        place("points_forward", lambda: build_points_forward_sheet(wb, client_name, current_label, pf_ref, header_cells=header_cells))

    return wb


def build_workbook(
    client_name: str,
    current_label: str,
    comparative_label: str,
    data: dict,
    results: list[ReconResult],
    control_account_results: list[ControlAccountResult] | None = None,
    matrix_results: list[MatrixResult] | None = None,
    ct_computation: CTComputation | None = None,
    fixed_asset_result: FixedAssetResult | None = None,
    asset_register_result: AssetRegisterResult | None = None,
    header_cells: dict | None = None,
) -> Workbook:
    """Builds the working paper pack in this system's own generic layout -
    a fresh workbook, sequential numbering from 1, everything enabled, the
    house-style A1/A2/A3 header convention unless header_cells overrides it.
    See build_workbook_into_template for generating into a practice's real
    uploaded template instead."""
    wb = Workbook()
    wb.active.title = "Index"
    return _generate_schedules(
        wb, _RefCounter(), wb.active,
        client_name, current_label, comparative_label, data, results,
        control_account_results or [], matrix_results or [],
        ct_computation, fixed_asset_result, asset_register_result,
        header_cells=header_cells,
    )


def build_workbook_into_template(
    template_source: str | Path | bytes,
    template_config: dict,
    client_name: str,
    current_label: str,
    comparative_label: str,
    data: dict,
    results: list[ReconResult],
    control_account_results: list[ControlAccountResult] | None = None,
    matrix_results: list[MatrixResult] | None = None,
    ct_computation: CTComputation | None = None,
    fixed_asset_result: FixedAssetResult | None = None,
    asset_register_result: AssetRegisterResult | None = None,
) -> Workbook:
    """Same schedules as build_workbook, generated into a copy of a
    practice's real uploaded template file instead of a fresh workbook -
    which schedules get generated, where each one is inserted relative to
    the template's own sheets, where numbering starts, and which cells the
    CLIENT NAME/PERIOD/SCHEDULE TITLE header block lands in on every
    generated sheet, are all driven by the template's customisation config
    (storage.DEFAULT_TEMPLATE_CONFIG documents every key) - real templates
    studied against this system use at least three different header-cell
    conventions (a simple A1/A2/A3 block; a "Client"/"Year End"/"Subject"
    labelled row layout with a separate Ref cell; a "Name of company:"/
    "Start period:"/"End period:" key-value block), so this can't be
    hardcoded the way the generic layout's A1/A2/A3 default is.

    The template's own sheets are never modified or overwritten - every
    generated schedule is a brand new sheet (openpyxl auto-suffixes a name
    that collides with one already in the template, e.g. "Index" ->
    "Index1", rather than erroring or overwriting), positioned via
    insert_after_sheet and otherwise left wherever it's created.

    Not yet wired to the config: materiality (recon.MATERIALITY_AMOUNT and
    equivalents in the other computation modules are still fixed
    module-level constants) - see README known limitations.
    """
    # template_source is bytes once templates live in Postgres rather than
    # on disk (a local path wouldn't survive between serverless
    # invocations) - openpyxl reads a file-like object just as happily as
    # a path, so wrap bytes in BytesIO rather than requiring a real file
    source = io.BytesIO(template_source) if isinstance(template_source, (bytes, bytearray)) else template_source
    wb = openpyxl.load_workbook(source)
    schedules_cfg = template_config.get("schedules", {})
    numbering_cfg = template_config.get("numbering", {})
    header_cells = template_config.get("header_cells")

    def enabled(key: str) -> bool:
        return schedules_cfg.get(key, {}).get("enabled", True)

    def reposition(sheet_name: str, key: str):
        target = schedules_cfg.get(key, {}).get("insert_after_sheet")
        if not target or target not in wb.sheetnames:
            return
        target_idx = wb.sheetnames.index(target)
        current_idx = wb.sheetnames.index(sheet_name)
        wb.move_sheet(sheet_name, offset=(target_idx + 1) - current_idx)

    def place(key: str, call):
        before = list(wb.sheetnames)
        result = call()
        new_names = [n for n in wb.sheetnames if n not in before]
        if new_names:
            reposition(new_names[-1], key)
        return result

    index_ws = wb.create_sheet("Index")
    index_target = schedules_cfg.get("index", {}).get("insert_after_sheet")
    if index_target and index_target in wb.sheetnames:
        reposition(index_ws.title, "index")
    else:
        # no configured position: the index goes first, ahead of every one
        # of the template's own sheets (never destructive - this only
        # shifts the template's existing sheets later, it can't drop them)
        current_idx = wb.sheetnames.index(index_ws.title)
        wb.move_sheet(index_ws.title, offset=-current_idx)

    ref = _RefCounter(start=int(numbering_cfg.get("start_at", 1)))
    return _generate_schedules(
        wb, ref, index_ws,
        client_name, current_label, comparative_label, data, results,
        control_account_results or [], matrix_results or [],
        ct_computation, fixed_asset_result, asset_register_result,
        enabled=enabled, place=place, header_cells=header_cells,
    )
