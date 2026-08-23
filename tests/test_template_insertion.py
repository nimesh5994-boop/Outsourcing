"""Tests for generating into a copy of a practice's real uploaded template
file, instead of the system's own generic layout - verifies the template's
own sheets are never touched, schedules land where the config says, the
enabled/disabled toggle is honoured, and numbering respects start_at."""
import copy

import openpyxl
import pytest

from app import control_accounts, corporation_tax, fixed_assets, nominal_matrix, recon
from app.excel_builder import _write_title, build_workbook_into_template
from app.storage import DEFAULT_TEMPLATE_CONFIG


def _make_template(tmp_path, sheet_names):
    wb = openpyxl.Workbook()
    wb.active.title = sheet_names[0]
    wb.active["A1"] = f"original content - {sheet_names[0]}"
    for name in sheet_names[1:]:
        ws = wb.create_sheet(name)
        ws["A1"] = f"original content - {name}"
    path = tmp_path / "practice_template.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def generation_inputs(canonical_data):
    results = recon.run_all_recons(canonical_data)
    ca_results = control_accounts.build_all_rollforwards(
        canonical_data["tb_current"], canonical_data["tb_comparative"], canonical_data["nominal_current"],
        canonical_data["aged_debtors"], canonical_data["aged_creditors"],
    )
    mx_results = nominal_matrix.build_all_matrices(canonical_data["tb_current"], canonical_data["nominal_current"])
    ct_computation = corporation_tax.compute(accounting_profit=float(canonical_data["pl_current"]["amount"].sum()))
    fixed_asset_result = fixed_assets.category_level_rollforward(
        canonical_data["tb_current"], canonical_data["tb_comparative"], canonical_data["nominal_current"]
    )
    return {
        "data": canonical_data, "results": results, "control_account_results": ca_results,
        "matrix_results": mx_results, "ct_computation": ct_computation, "fixed_asset_result": fixed_asset_result,
    }


def test_template_sheets_are_preserved_untouched(tmp_path, generation_inputs):
    template_path = _make_template(tmp_path, ["Cover Page", "Firm Notes"])
    config = copy.deepcopy(DEFAULT_TEMPLATE_CONFIG)

    wb = build_workbook_into_template(
        template_path, config, "Brightwell Landscaping Supplies Limited",
        "Year ended 31 December 2025", "Year ended 31 December 2024", **generation_inputs,
    )

    assert "Cover Page" in wb.sheetnames
    assert "Firm Notes" in wb.sheetnames
    assert wb["Cover Page"]["A1"].value == "original content - Cover Page"
    assert wb["Firm Notes"]["A1"].value == "original content - Firm Notes"
    # generated content added on top, not just the two original sheets
    assert len(wb.sheetnames) > 2


def test_index_sheet_does_not_overwrite_templates_own_first_sheet(tmp_path, generation_inputs):
    template_path = _make_template(tmp_path, ["Cover Page"])
    config = copy.deepcopy(DEFAULT_TEMPLATE_CONFIG)

    wb = build_workbook_into_template(
        template_path, config, "Brightwell Landscaping Supplies Limited",
        "Year ended 31 December 2025", "Year ended 31 December 2024", **generation_inputs,
    )

    assert wb["Cover Page"]["A1"].value == "original content - Cover Page"
    assert "Index" in wb.sheetnames
    # default (no insert_after_sheet configured): index goes first
    assert wb.sheetnames.index("Index") < wb.sheetnames.index("Cover Page")


def test_disabled_schedule_is_skipped(tmp_path, generation_inputs):
    template_path = _make_template(tmp_path, ["Cover Page"])
    config = copy.deepcopy(DEFAULT_TEMPLATE_CONFIG)
    config["schedules"]["corporation_tax"]["enabled"] = False

    wb = build_workbook_into_template(
        template_path, config, "Brightwell Landscaping Supplies Limited",
        "Year ended 31 December 2025", "Year ended 31 December 2024", **generation_inputs,
    )

    assert not any("Corporation Tax" in name for name in wb.sheetnames)
    assert any("TB Lead Schedule" in name for name in wb.sheetnames)  # sibling schedules unaffected


def test_insert_after_sheet_positions_the_schedule(tmp_path, generation_inputs):
    template_path = _make_template(tmp_path, ["Cover Page", "Anchor Sheet", "Back Matter"])
    config = copy.deepcopy(DEFAULT_TEMPLATE_CONFIG)
    config["schedules"]["tb_lead_schedule"]["insert_after_sheet"] = "Anchor Sheet"

    wb = build_workbook_into_template(
        template_path, config, "Brightwell Landscaping Supplies Limited",
        "Year ended 31 December 2025", "Year ended 31 December 2024", **generation_inputs,
    )

    tb_sheet = next(n for n in wb.sheetnames if "TB Lead Schedule" in n)
    anchor_idx = wb.sheetnames.index("Anchor Sheet")
    back_matter_idx = wb.sheetnames.index("Back Matter")
    tb_idx = wb.sheetnames.index(tb_sheet)
    assert anchor_idx < tb_idx < back_matter_idx


def test_numbering_start_at_is_respected(tmp_path, generation_inputs):
    template_path = _make_template(tmp_path, ["Cover Page"])
    config = copy.deepcopy(DEFAULT_TEMPLATE_CONFIG)
    config["numbering"]["start_at"] = 21

    wb = build_workbook_into_template(
        template_path, config, "Brightwell Landscaping Supplies Limited",
        "Year ended 31 December 2025", "Year ended 31 December 2024", **generation_inputs,
    )

    assert any(n.startswith("21 ") for n in wb.sheetnames)
    assert not any(n.startswith("1 ") for n in wb.sheetnames)


def test_sheet_name_collision_with_template_is_handled_not_fatal(tmp_path, generation_inputs):
    # the template already has its own sheet literally named "Index" -
    # openpyxl auto-suffixes ours rather than erroring or overwriting
    template_path = _make_template(tmp_path, ["Index", "Cover Page"])
    config = copy.deepcopy(DEFAULT_TEMPLATE_CONFIG)

    wb = build_workbook_into_template(
        template_path, config, "Brightwell Landscaping Supplies Limited",
        "Year ended 31 December 2025", "Year ended 31 December 2024", **generation_inputs,
    )

    assert wb["Index"]["A1"].value == "original content - Index"
    assert any(n != "Index" and "Index" in n for n in wb.sheetnames)


def test_formula_linked_schedules_still_reference_data_sheets_in_template_mode(tmp_path, generation_inputs):
    # the formula-linked engine must work the same way when the base
    # workbook is a loaded template, not just a fresh Workbook()
    template_path = _make_template(tmp_path, ["Cover Page"])
    config = copy.deepcopy(DEFAULT_TEMPLATE_CONFIG)

    wb = build_workbook_into_template(
        template_path, config, "Brightwell Landscaping Supplies Limited",
        "Year ended 31 December 2025", "Year ended 31 December 2024", **generation_inputs,
    )

    assert "DATA_TB_Current" in wb.sheetnames
    tb_sheet_name = next(n for n in wb.sheetnames if "TB Lead Schedule" in n)
    ws = wb[tb_sheet_name]
    formula_cells = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith("=")]
    assert formula_cells, "expected live formulas on the TB Lead Schedule sheet even in template mode"


def test_write_title_default_uses_a1_a2_a3():
    ws = openpyxl.Workbook().active
    next_row = _write_title(ws, "Acme Ltd", "Year ended 31 December 2025", "TB LEAD SCHEDULE", ref="1")
    assert ws["A1"].value == "ACME LTD"
    assert ws["A2"].value == "YEAR ENDED 31 DECEMBER 2025"
    assert ws["A3"].value == "1  TB LEAD SCHEDULE"
    assert next_row == 5  # unchanged from the pre-header_cells default


def test_write_title_honours_custom_header_cells_and_computes_content_row():
    # mirrors the "Client"/"Year End"/"Subject" row-1-3 convention found in
    # the real Xero Ltd / Sage-QBO-FreeAgent / Manual Job templates - values
    # in column B rather than column A
    ws = openpyxl.Workbook().active
    header_cells = {"client_name_cell": "B1", "period_cell": "B2", "schedule_title_cell": "B3"}
    next_row = _write_title(ws, "Acme Ltd", "Year ended 31 December 2025", "TB LEAD SCHEDULE", ref="1", header_cells=header_cells)
    assert ws["B1"].value == "ACME LTD"
    assert ws["B2"].value == "YEAR ENDED 31 DECEMBER 2025"
    assert ws["B3"].value == "1  TB LEAD SCHEDULE"
    assert ws["A1"].value is None  # nothing leaks into the default cells
    assert next_row == 5


def test_write_title_content_row_follows_the_lowest_configured_header_cell():
    # a convention where the header block spans further down than rows 1-3
    # (e.g. the "Name of company:"/"Start period:"/"End period:" block seen
    # in a newer real template) needs schedule content pushed down further
    ws = openpyxl.Workbook().active
    header_cells = {"client_name_cell": "B1", "period_cell": "B2", "schedule_title_cell": "A5"}
    next_row = _write_title(ws, "Acme Ltd", "Year ended 31 December 2025", "TB LEAD SCHEDULE", header_cells=header_cells)
    assert next_row == 7  # max row (5) + 2


def test_template_config_header_cells_applies_to_generated_schedules(tmp_path, generation_inputs):
    template_path = _make_template(tmp_path, ["Cover Page"])
    config = copy.deepcopy(DEFAULT_TEMPLATE_CONFIG)
    config["header_cells"] = {"client_name_cell": "B1", "period_cell": "B2", "schedule_title_cell": "B3"}

    wb = build_workbook_into_template(
        template_path, config, "Brightwell Landscaping Supplies Limited",
        "Year ended 31 December 2025", "Year ended 31 December 2024", **generation_inputs,
    )

    tb_sheet_name = next(n for n in wb.sheetnames if "TB Lead Schedule" in n)
    ws = wb[tb_sheet_name]
    assert ws["B1"].value == "BRIGHTWELL LANDSCAPING SUPPLIES LIMITED"
    assert ws["B2"].value == "YEAR ENDED 31 DECEMBER 2025"
    assert "TRIAL BALANCE LEAD SCHEDULE" in ws["B3"].value
    assert ws["A1"].value is None
