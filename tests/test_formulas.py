"""Verifies formula-linked schedules actually evaluate correctly - not just
that the string looks like a formula, but that a real formula engine
computes the same numbers the equivalent Python (pandas) computation does.

Uses the `formulas` library (pure-Python Excel formula evaluator) since
LibreOffice can't be relied on in this sandbox to recalculate and verify.
These tests are skipped if `formulas` isn't installed - it's a heavy,
dev-only dependency (requirements-dev.txt), not part of the app's runtime
requirements.
"""
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

pytest.importorskip("formulas")
import formulas  # noqa: E402

from app import control_accounts as ca  # noqa: E402
from app import corporation_tax as ct_mod  # noqa: E402
from app import financial_statements as fs  # noqa: E402
from app import fixed_assets as fa  # noqa: E402
from app import nominal_matrix as nm  # noqa: E402
from app import recon, xero_reports as xr  # noqa: E402
from app.data_sheets import write_data_sheets  # noqa: E402
from app.excel_builder import (  # noqa: E402
    build_asset_register_sheet_formulas,
    build_bs_statement_sheet_formulas,
    build_control_account_sheet_formulas,
    build_corporation_tax_sheet_formulas,
    build_fixed_asset_category_sheet_formulas,
    build_matrix_sheet_formulas,
    build_pl_statement_sheet_formulas,
    build_tb_lead_schedule_formulas,
)
from app.parsers import FileDataSource  # noqa: E402

SAMPLE_DIR = Path(__file__).resolve().parent.parent / "sample_data"


def _evaluate(path) -> dict:
    xl = formulas.ExcelModel().loads(str(path)).finish()
    return xl.calculate()


def _cell(sol: dict, workbook_filename: str, sheet: str, coord: str):
    key = f"'[{workbook_filename}]{sheet.upper()}'!{coord}"
    return sol[key].value[0][0]


def _cell_or_none(sol: dict, workbook_filename: str, sheet: str, coord: str):
    key = f"'[{workbook_filename}]{sheet.upper()}'!{coord}"
    if key not in sol:
        return None
    return sol[key].value[0][0]


@pytest.fixture
def canonical_tb():
    tb_current, tb_comparative = xr.parse_trial_balance(
        FileDataSource(SAMPLE_DIR / "trial_balance_current_xero.xlsx")
    )
    return tb_current, tb_comparative


def test_tb_lead_schedule_formulas_match_python_ground_truth(tmp_path, canonical_tb):
    tb_current, tb_comparative = canonical_tb
    variance = recon.variance_analysis(tb_current, tb_comparative)

    wb = Workbook()
    refs = write_data_sheets(wb, {"tb_current": tb_current, "tb_comparative": tb_comparative})
    build_tb_lead_schedule_formulas(wb, "Brightwell Landscaping Supplies Limited", "Year ended 31 December 2025", "1", variance.detail, refs)
    wb.remove(wb["Sheet"])

    out = tmp_path / "tb_formula.xlsx"
    wb.save(out)

    sol = _evaluate(out)
    sheet_name = "1 TB Lead Schedule"

    errors = [k for k, v in sol.items() if f"[{out.name}]{sheet_name.upper()}" in k and "VALUE!" in str(v.value)]
    assert not errors, f"formula errors: {errors}"

    for i, row in enumerate(variance.detail.itertuples()):
        r = 6 + i
        assert _cell(sol, out.name, sheet_name, f"C{r}") == pytest.approx(row.current_year, abs=0.01)
        assert _cell(sol, out.name, sheet_name, f"D{r}") == pytest.approx(row.comparative_year, abs=0.01)
        assert _cell(sol, out.name, sheet_name, f"E{r}") == pytest.approx(row.variance_amount, abs=0.01)
        assert bool(_cell(sol, out.name, sheet_name, f"G{r}")) == bool(row.flag)


def test_tb_lead_schedule_formulas_respect_a_configured_materiality(tmp_path, canonical_tb):
    # a tighter materiality than the £500/10% default should flag rows the
    # default wouldn't - and the live Excel formula (not just the Python
    # side) needs to bake in that same tighter value, since a template's
    # configured materiality is meant to reach the actual workbook
    tb_current, tb_comparative = canonical_tb
    tight_materiality, tight_pct = 50.0, 0.01
    variance = recon.variance_analysis(tb_current, tb_comparative, materiality=tight_materiality, variance_pct_threshold=tight_pct)
    assert variance.detail["flag"].any(), "fixture doesn't exercise a flagged row at the tightened threshold"

    wb = Workbook()
    refs = write_data_sheets(wb, {"tb_current": tb_current, "tb_comparative": tb_comparative})
    build_tb_lead_schedule_formulas(
        wb, "Brightwell Landscaping Supplies Limited", "Year ended 31 December 2025", "1", variance.detail, refs,
        materiality=tight_materiality, variance_pct_threshold=tight_pct,
    )
    wb.remove(wb["Sheet"])

    out = tmp_path / "tb_formula_tight_materiality.xlsx"
    wb.save(out)

    sol = _evaluate(out)
    sheet_name = "1 TB Lead Schedule"
    for i, row in enumerate(variance.detail.itertuples()):
        r = 6 + i
        assert bool(_cell(sol, out.name, sheet_name, f"G{r}")) == bool(row.flag)


def test_control_account_formulas_match_python_ground_truth(tmp_path, canonical_data):
    results = ca.build_all_rollforwards(
        canonical_data["tb_current"], canonical_data["tb_comparative"], canonical_data["nominal_current"],
        canonical_data["aged_debtors"], canonical_data["aged_creditors"],
    )
    trade_creditors = next(r for r in results if r.account_code == "8010")

    wb = Workbook()
    refs = write_data_sheets(wb, canonical_data)
    build_control_account_sheet_formulas(wb, "Brightwell Landscaping Supplies Limited", "Year ended 31 December 2025", "1", trade_creditors, refs)
    wb.remove(wb["Sheet"])

    out = tmp_path / "ca_formula.xlsx"
    wb.save(out)

    sol = _evaluate(out)
    sheet_name = f"1 {trade_creditors.account_name}"[:31]

    errors = [k for k, v in sol.items() if f"[{out.name}]{sheet_name.upper()}" in k and ("VALUE!" in str(v.value) or "REF!" in str(v.value))]
    assert not errors, f"formula errors: {errors}"

    # BALANCE C/FWD (per TB) row: Debit/Credit split of the current TB balance
    schedule_rows = {row["Item"]: i for i, row in trade_creditors.schedule.iterrows()}
    c_fwd_row_idx = schedule_rows["BALANCE C/FWD (per TB)"]
    expected_debit = trade_creditors.schedule.iloc[c_fwd_row_idx]["Debit £"]
    expected_credit = trade_creditors.schedule.iloc[c_fwd_row_idx]["Credit £"]
    excel_row = 8 + c_fwd_row_idx  # row 5 = status, table header at row 7, first data row 8
    assert _cell(sol, out.name, sheet_name, f"C{excel_row}") == pytest.approx(expected_debit, abs=0.01)
    assert _cell(sol, out.name, sheet_name, f"D{excel_row}") == pytest.approx(expected_credit, abs=0.01)

    # the breakdown's "unexplained difference" should be ~0 on this data (confirmed against real client data separately)
    for r in range(6, 6 + len(trade_creditors.schedule) + len(trade_creditors.breakdown) + 6):
        label = _cell_or_none(sol, out.name, sheet_name, f"A{r}")
        if label and "UNEXPLAINED" in str(label):
            # this dataset's aged creditors listing doesn't fully explain the
            # TB balance (a known, deliberate gap - see
            # test_control_account_breakdown_flags_unexplained_difference in
            # test_pipeline.py), so just confirm the formula reproduces the
            # same gap the Python computation found, not that it's zero
            unexplained_row = trade_creditors.breakdown[
                trade_creditors.breakdown["Party"].astype(str).str.startswith("UNEXPLAINED")
            ].iloc[0]
            assert _cell(sol, out.name, sheet_name, f"B{r}") == pytest.approx(unexplained_row["Amount £"], abs=0.5)
            break
    else:
        pytest.fail("did not find the UNEXPLAINED DIFFERENCE row")


def test_pl_bs_formulas_match_python_ground_truth_and_tie_out(tmp_path, canonical_data):
    pl_result = fs.build_pl_statement(canonical_data["pl_current"])
    bs_result = fs.build_bs_statement(canonical_data["bs_current"], pl_result.net_profit)
    assert bs_result.status == "ok"  # sample data is built to tie to £0.00 - see test_pipeline.py

    wb = Workbook()
    refs = write_data_sheets(wb, canonical_data)
    pl_sheet_name, pl_net_profit_cell = build_pl_statement_sheet_formulas(
        wb, "Brightwell Landscaping Supplies Limited", "Year ended 31 December 2025", "1", pl_result, refs
    )
    bs_sheet_name = build_bs_statement_sheet_formulas(
        wb, "Brightwell Landscaping Supplies Limited", "Year ended 31 December 2025", "2", bs_result, refs,
        pl_sheet_name, pl_net_profit_cell,
    )
    wb.remove(wb["Sheet"])

    out = tmp_path / "pl_bs_formula.xlsx"
    wb.save(out)

    sol = _evaluate(out)

    for sheet in (pl_sheet_name, bs_sheet_name):
        errors = [k for k, v in sol.items() if f"[{out.name}]{sheet.upper()}" in k and ("VALUE!" in str(v.value) or "REF!" in str(v.value))]
        assert not errors, f"formula errors on {sheet}: {errors}"

    pl_line_to_row = {line: 8 + i for i, line in enumerate(pl_result.statement["Line"])}
    for line, amount in zip(pl_result.statement["Line"], pl_result.statement["Amount"]):
        r = pl_line_to_row[line]
        assert _cell(sol, out.name, pl_sheet_name, f"B{r}") == pytest.approx(amount, abs=0.01)

    bs_line_to_row = {line: 8 + i for i, line in enumerate(bs_result.statement["Line"])}
    for line, amount in zip(bs_result.statement["Line"], bs_result.statement["Amount"]):
        r = bs_line_to_row[line]
        assert _cell(sol, out.name, bs_sheet_name, f"B{r}") == pytest.approx(amount, abs=0.01)

    check_row = bs_line_to_row["CHECK: Net Assets - Total Equity (should be £0)"]
    assert _cell(sol, out.name, bs_sheet_name, f"B{check_row}") == pytest.approx(0.0, abs=0.01)


def test_bs_statement_formulas_render_unrecognised_category_note(tmp_path):
    # a Suspense-type account outside the five recognised B/S categories -
    # excluded from every SUMPRODUCT total, so the CHECK row shows a real
    # gap - and the note/table naming it should be written into the sheet
    # without disturbing the DETAIL BY ACCOUNT formulas below it
    bs = pd.DataFrame([
        {"account_code": "1", "account_name": "Bank", "category": "Current Asset", "amount": 5000.0},
        {"account_code": "2", "account_name": "Share Capital", "category": "Equity", "amount": -100.0},
        {"account_code": "3", "account_name": "Retained Earnings", "category": "Equity", "amount": -9900.0},
        {"account_code": "4", "account_name": "Suspense Account", "category": "Unclassified", "amount": 5000.0},
    ])
    bs_result = fs.build_bs_statement(bs, net_profit=0.0)
    assert bs_result.status == "review"
    assert not bs_result.unrecognized_detail.empty

    wb = Workbook()
    refs = write_data_sheets(wb, {"bs_current": bs})
    bs_sheet_name = build_bs_statement_sheet_formulas(
        wb, "Test Client Ltd", "Year ended 31 December 2025", "1", bs_result, refs, None, None,
    )
    wb.remove(wb["Sheet"])

    out = tmp_path / "bs_unrecognised_category.xlsx"
    wb.save(out)

    sol = _evaluate(out)
    errors = [k for k, v in sol.items() if f"[{out.name}]{bs_sheet_name.upper()}" in k and ("VALUE!" in str(v.value) or "REF!" in str(v.value))]
    assert not errors, f"formula errors on {bs_sheet_name}: {errors}"

    bs_line_to_row = {line: 8 + i for i, line in enumerate(bs_result.statement["Line"])}
    check_row = bs_line_to_row["CHECK: Net Assets - Total Equity (should be £0)"]
    assert _cell(sol, out.name, bs_sheet_name, f"B{check_row}") == pytest.approx(-5000.0, abs=0.01)

    ws = load_workbook(out)[bs_sheet_name]
    found_note = any(cell.value and "UNRECOGNISED ACCOUNT TYPE" in str(cell.value) for row in ws.iter_rows() for cell in row)
    found_account_name = any(cell.value == "Suspense Account" for row in ws.iter_rows() for cell in row)
    found_amount = any(cell.value == 5000.0 for row in ws.iter_rows() for cell in row)
    assert found_note, "unrecognised-account note not written to the sheet"
    assert found_account_name, "Suspense Account not named in the unrecognised-account detail table"
    assert found_amount, "the unrecognised account's £5,000 amount not written to the detail table"

    # DETAIL BY ACCOUNT still lists every account, including the
    # unrecognised one - the note is additive, not a replacement
    detail_labels = [cell.value for row in ws.iter_rows() for cell in row if cell.value == "DETAIL BY ACCOUNT"]
    assert detail_labels, "DETAIL BY ACCOUNT header missing after the unrecognised-account note block"


def _fa_tb_row(code, name, account_type, debit, credit):
    return {"account_code": code, "account_name": name, "account_type": account_type,
            "debit": debit, "credit": credit, "balance": debit - credit}


def test_fixed_asset_category_formulas_match_python_ground_truth(tmp_path):
    # mirrors test_pipeline.py's unpunctuated-name regression fixture, with
    # nominal movements added so additions/depreciation-charge formulas get
    # exercised (and deliberately tie to zero, so the category ties out)
    tb_current = pd.DataFrame([
        _fa_tb_row("6350", "IT EQUIPMENT COST BROUGHT FORWARD", "Fixed Asset", 1612.19, 0),
        _fa_tb_row("6351", "IT EQUIPMENT COST OF ADDITIONS", "Fixed Asset", 1514.15, 0),
        _fa_tb_row("6360", "IT EQUIPMENT ACCUMULATED DEPRECIATION BROUGHT FORWARD", "Fixed Asset", 0, 189.57),
    ])
    tb_comparative = pd.DataFrame([
        _fa_tb_row("6350", "IT EQUIPMENT COST BROUGHT FORWARD", "Fixed Asset", 1612.19, 0),
        _fa_tb_row("6360", "IT EQUIPMENT ACCUMULATED DEPRECIATION BROUGHT FORWARD", "Fixed Asset", 0, 89.57),
    ])
    nominal_current = pd.DataFrame([
        {"account_code": "6351", "debit": 1514.15, "credit": 0.0},
        {"account_code": "6360", "debit": 0.0, "credit": 100.0},
    ])

    result = fa.category_level_rollforward(tb_current, tb_comparative, nominal_current)
    assert result.status == "ok"
    grouped_codes = fa.group_fixed_asset_codes(tb_current)

    wb = Workbook()
    refs = write_data_sheets(wb, {"tb_current": tb_current, "tb_comparative": tb_comparative, "nominal_current": nominal_current})
    sheet_name = build_fixed_asset_category_sheet_formulas(
        wb, "Brightwell Landscaping Supplies Limited", "Year ended 31 December 2025", "1", result, refs, grouped_codes
    )
    wb.remove(wb["Sheet"])

    out = tmp_path / "fa_formula.xlsx"
    wb.save(out)

    sol = _evaluate(out)

    errors = [k for k, v in sol.items() if f"[{out.name}]{sheet_name.upper()}" in k and ("VALUE!" in str(v.value) or "REF!" in str(v.value))]
    assert not errors, f"formula errors: {errors}"

    col_letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    columns = [
        "Category", "Cost b/fwd", "Additions", "Disposals (cost)", "Cost c/fwd (per TB)", "Cost diff",
        "Acc. depreciation b/fwd", "Depreciation charge", "Depreciation on disposals", "Acc. depreciation c/fwd (per TB)",
        "Depreciation diff", "NBV b/fwd", "NBV c/fwd",
    ]
    row0 = result.detail.iloc[0]
    r = 8  # row5=status, table header row7, first data row8
    for col, letter in zip(columns, col_letters):
        if col == "Category":
            assert _cell(sol, out.name, sheet_name, f"{letter}{r}") == row0[col]
        else:
            assert _cell(sol, out.name, sheet_name, f"{letter}{r}") == pytest.approx(row0[col], abs=0.01)


def test_asset_register_formulas_match_python_ground_truth(tmp_path):
    # 3 still-held assets (reducing balance, straight line, and a 0%-rate
    # freehold-style asset that must not error), 1 already-disposed asset
    # that must be excluded, a genuine addition, and a TB deliberately
    # written with three different casings of "Fixed Asset" (Fixed Asset/
    # fixed asset/FIXED ASSET) to prove the formula's account_type match
    # is case-insensitive the same way Python's own .str.lower() compare is.
    prior_register = pd.DataFrame([
        {"asset_id": "FA-001", "description": "Ford Transit Van", "category": "Motor Vehicles",
         "date_acquired": pd.Timestamp("2022-03-15"), "cost": 18000.0, "depreciation_method": "Reducing Balance",
         "depreciation_rate": 25.0, "accumulated_depreciation_b_fwd": 10195.31, "disposed": "No"},
        {"asset_id": "FA-002", "description": "Dell Precision Workstation", "category": "Computer Equipment",
         "date_acquired": pd.Timestamp("2021-09-01"), "cost": 2200.0, "depreciation_method": "Straight Line",
         "depreciation_rate": 33.3, "accumulated_depreciation_b_fwd": 1900.0, "disposed": "No"},
        {"asset_id": "FA-003", "description": "Freehold display cabinet", "category": "Fixtures and Fittings",
         "date_acquired": pd.Timestamp("2020-06-01"), "cost": 2000.0, "depreciation_method": "Straight Line",
         "depreciation_rate": 0.0, "accumulated_depreciation_b_fwd": 500.0, "disposed": "No"},
        {"asset_id": "FA-004", "description": "Old delivery van (sold)", "category": "Motor Vehicles",
         "date_acquired": pd.Timestamp("2019-01-01"), "cost": 9000.0, "depreciation_method": "Reducing Balance",
         "depreciation_rate": 25.0, "accumulated_depreciation_b_fwd": 9000.0, "disposed": "Yes"},
    ])
    tb_current = pd.DataFrame([
        _fa_tb_row("3000", "MOTOR VEHICLES COST", "Fixed Asset", 18000, 0),
        _fa_tb_row("3050", "MOTOR VEHICLES DEPRECIATION", "Fixed Asset", 0, 13695.31),
        _fa_tb_row("6350", "COMPUTER EQUIPMENT - COST", "fixed asset", 2200, 0),
        _fa_tb_row("6360", "COMPUTER EQUIPMENT - DEPRECIATION", "fixed asset", 0, 2266.6),
        _fa_tb_row("4100", "FIXTURES AND FITTINGS COST", "FIXED ASSET", 2000, 0),
        _fa_tb_row("4150", "FIXTURES AND FITTINGS DEPRECIATION", "FIXED ASSET", 0, 500),
        _fa_tb_row("9000", "SALES", "Sales", 0, 50000),
    ])
    nominal_current = pd.DataFrame([
        {"date": pd.Timestamp("2025-07-01"), "account_code": "3000", "account_name": "MOTOR VEHICLES COST",
         "reference": "INV-501", "description": "New delivery van", "contact": "Van Dealer Ltd", "debit": 0.0, "credit": 0.0},
    ])

    result = fa.asset_level_rollforward(prior_register, nominal_current, tb_current, period_days=365)
    assert result.status == "review"  # deliberate variance, asserted below
    assert list(result.asset_schedule["Asset ID"]) == ["FA-001", "FA-002", "FA-003"]  # FA-004 excluded

    wb = Workbook()
    refs = write_data_sheets(wb, {"tb_current": tb_current, "fixed_asset_register": prior_register, "nominal_current": nominal_current})
    build_asset_register_sheet_formulas(wb, "Test Client", "Year ended 31 December 2025", "5", result, refs, materiality=500.0)
    wb.remove(wb["Sheet"])
    sheet_name = "5 Fixed Asset Register"

    out = tmp_path / "far_asset_formula.xlsx"
    wb.save(out)
    sol = _evaluate(out)

    errors = [k for k, v in sol.items() if f"[{out.name}]{sheet_name.upper()}" in k and ("VALUE!" in str(v.value) or "REF!" in str(v.value) or "#NAME?" in str(v.value))]
    assert not errors, f"formula errors: {errors}"

    # per-asset rollforward columns: E..L, rows 10/11/12 for FA-001/002/003
    columns = ["Cost", "Acc. Dep. b/fwd", "Depreciation Charge", "Acc. Dep. c/fwd", "NBV b/fwd", "NBV c/fwd"]
    col_letters = ["E", "H", "I", "J", "K", "L"]
    for i, row_num in enumerate([10, 11, 12]):
        row_data = result.asset_schedule.iloc[i]
        for col, letter in zip(columns, col_letters):
            assert _cell(sol, out.name, sheet_name, f"{letter}{row_num}") == pytest.approx(row_data[col], abs=0.01), f"{col} for row {row_num}"

    # the 0%-rate asset (FA-003, row 12) must charge exactly zero, not error
    assert _cell(sol, out.name, sheet_name, "I12") == pytest.approx(0.0, abs=0.001)

    # summary block: totals + the live TB tie-out, case-insensitive account_type match
    summary_row = 17  # label row15, header row16, values row17 (no new additions in this fixture - debit is 0)
    summary = result.summary.iloc[0]
    assert _cell(sol, out.name, sheet_name, f"A{summary_row}") == pytest.approx(summary["Total cost (register + unrecorded additions)"], abs=0.01)
    assert _cell(sol, out.name, sheet_name, f"B{summary_row}") == pytest.approx(summary["Total accumulated depreciation"], abs=0.01)
    assert _cell(sol, out.name, sheet_name, f"C{summary_row}") == pytest.approx(summary["Total NBV (register)"], abs=0.01)
    assert _cell(sol, out.name, sheet_name, f"D{summary_row}") == pytest.approx(summary["TB fixed asset net balance"], abs=0.01)
    assert _cell(sol, out.name, sheet_name, f"E{summary_row}") == pytest.approx(summary["Variance"], abs=0.01)


def test_asset_register_formulas_fall_back_to_plain_values_without_register_data_sheet(tmp_path):
    # no fixed_asset_register data sheet written (e.g. the register upload
    # is missing) - build_asset_register_sheet_formulas must not be called
    # with refs.fixed_asset_register as None in real use (excel_builder
    # gates on it), but the function itself still degrades safely rather
    # than raising if it ever is.
    prior_register = pd.DataFrame([
        {"asset_id": "FA-001", "description": "Ford Transit Van", "category": "Motor Vehicles",
         "date_acquired": pd.Timestamp("2022-03-15"), "cost": 18000.0, "depreciation_method": "Reducing Balance",
         "depreciation_rate": 25.0, "accumulated_depreciation_b_fwd": 10195.31, "disposed": "No"},
    ])
    result = fa.asset_level_rollforward(prior_register, None, None, period_days=365)
    wb = Workbook()
    refs = write_data_sheets(wb, {})
    build_asset_register_sheet_formulas(wb, "Test Client", "2025", "5", result, refs)
    ws = wb["5 Fixed Asset Register"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "FA-001" in values


@pytest.mark.parametrize("accounting_profit,booked_tax_charge", [
    (30_000, 5_600.00),   # small profits rate band
    (150_000, 36_000.00),  # marginal relief band (the known 24% reference point)
    (300_000, 74_990.00),  # main rate band
])
def test_corporation_tax_formulas_match_python_ground_truth(tmp_path, accounting_profit, booked_tax_charge):
    ct = ct_mod.compute(accounting_profit=accounting_profit, booked_tax_charge=booked_tax_charge)

    wb = Workbook()
    sheet_name = build_corporation_tax_sheet_formulas(wb, "Brightwell Landscaping Supplies Limited", "Year ended 31 December 2025", "1", ct)
    wb.remove(wb["Sheet"])

    out = tmp_path / f"ct_formula_{accounting_profit}.xlsx"
    wb.save(out)

    sol = _evaluate(out)

    errors = [k for k, v in sol.items() if f"[{out.name}]{sheet_name.upper()}" in k and ("VALUE!" in str(v.value) or "REF!" in str(v.value))]
    assert not errors, f"formula errors: {errors}"

    rows = {
        "profit": 7, "disallow": 8, "capex": 9, "taxable": 10, "augmented": 11,
        "assoc": 12, "period": 13, "lower": 14, "upper": 15, "band": 16,
        "relief": 17, "charge": 18, "booked": 19, "variance": 20,
    }
    assert _cell(sol, out.name, sheet_name, f"B{rows['taxable']}") == pytest.approx(ct.taxable_profit, abs=0.01)
    assert _cell(sol, out.name, sheet_name, f"B{rows['lower']}") == pytest.approx(ct.lower_limit, abs=0.01)
    assert _cell(sol, out.name, sheet_name, f"B{rows['upper']}") == pytest.approx(ct.upper_limit, abs=0.01)
    assert str(_cell(sol, out.name, sheet_name, f"B{rows['band']}")) == ct.band
    assert _cell(sol, out.name, sheet_name, f"B{rows['relief']}") == pytest.approx(ct.marginal_relief, abs=0.01)
    assert _cell(sol, out.name, sheet_name, f"B{rows['charge']}") == pytest.approx(ct.tax_charge, abs=0.01)
    assert _cell(sol, out.name, sheet_name, f"B{rows['variance']}") == pytest.approx(ct.variance, abs=0.01)


def test_nominal_matrix_formulas_match_python_ground_truth(tmp_path, canonical_data):
    tb_current, nominal_current = canonical_data["tb_current"], canonical_data["nominal_current"]
    result = nm.build_matrix("8010", "TRADE CREDITORS", nominal_current)
    assert not result.matrix.empty

    wb = Workbook()
    refs = write_data_sheets(wb, canonical_data)
    sheet_name = build_matrix_sheet_formulas(
        wb, "Brightwell Landscaping Supplies Limited", "Year ended 31 December 2025", "1", result, refs, nominal_current
    )
    wb.remove(wb["Sheet"])

    out = tmp_path / "matrix_formula.xlsx"
    wb.save(out)

    sol = _evaluate(out)

    errors = [k for k, v in sol.items() if f"[{out.name}]{sheet_name.upper()}" in k and ("VALUE!" in str(v.value) or "REF!" in str(v.value))]
    assert not errors, f"formula errors: {errors}"

    col_names = [c for c in result.matrix.columns if c not in ("date", "reference", "description", "contact", "TOTAL", "DIFF")]
    n_cols = len(col_names)
    first_col, total_col, diff_col = 5, 5 + n_cols, 5 + n_cols + 1
    from openpyxl.utils import get_column_letter as _gcl

    matrix_sorted = result.matrix.sort_values(["date", "reference", "description", "contact"]).reset_index(drop=True)
    for i, py_row in matrix_sorted.iterrows():
        r = 8 + i  # row5=status, table header row7, first data row8
        for j, col in enumerate(col_names):
            letter = _gcl(first_col + j)
            assert _cell(sol, out.name, sheet_name, f"{letter}{r}") == pytest.approx(py_row[col], abs=0.01)
        assert _cell(sol, out.name, sheet_name, f"{_gcl(total_col)}{r}") == pytest.approx(py_row["TOTAL"], abs=0.01)
        assert _cell(sol, out.name, sheet_name, f"{_gcl(diff_col)}{r}") == pytest.approx(py_row["DIFF"], abs=0.01)
