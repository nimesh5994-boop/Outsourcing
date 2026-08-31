"""Integration tests for the Postgres-backed storage layer and the FastAPI
HTTP routes that depend on it - the part of the storage.py rewrite (JSON
files on disk -> Postgres entities/files/mapping_profiles tables) that the
rest of the suite never exercises, since every other test drives the
computation modules directly with in-memory DataFrames/fixtures.

Runs against a real, throwaway Postgres schema (not the app's production
Neon database) so the whole path - entity CRUD, file BYTEA storage,
template normalisation, and the practice -> template -> client -> job ->
upload -> generate -> download HTTP flow - is proven against a real
database engine, not a mock. Skipped automatically when no test database
is reachable (e.g. a machine without Postgres installed); set
TEST_DATABASE_URL to point at one explicitly.
"""
import io
import os
import uuid
from pathlib import Path

import psycopg
import pytest
from fastapi.testclient import TestClient

SAMPLE_DIR = Path(__file__).resolve().parent.parent / "sample_data"

TEST_DATABASE_URL = os.environ.get(
    "TEST_DATABASE_URL", "postgresql://wpa_test:wpa_test@127.0.0.1:5432/wpa_test"
)


def _database_available() -> bool:
    try:
        conn = psycopg.connect(TEST_DATABASE_URL, connect_timeout=3)
        conn.close()
        return True
    except Exception:
        return False


pytestmark = pytest.mark.skipif(
    not _database_available(),
    reason=(
        f"no reachable Postgres test database at {TEST_DATABASE_URL!r} - "
        "set TEST_DATABASE_URL, or run one locally, to exercise storage.py"
    ),
)


@pytest.fixture
def http_client(monkeypatch):
    """A TestClient wired to a throwaway schema in the test database, so
    repeated/parallel test runs never collide and nothing here can ever
    touch the real app's data."""
    schema = f"test_{uuid.uuid4().hex[:12]}"
    admin_conn = psycopg.connect(TEST_DATABASE_URL, autocommit=True)
    admin_conn.execute(f'CREATE SCHEMA "{schema}"')
    admin_conn.close()

    separator = "&" if "?" in TEST_DATABASE_URL else "?"
    scoped_dsn = f"{TEST_DATABASE_URL}{separator}options=-csearch_path%3D{schema}"
    monkeypatch.setenv("DATABASE_URL", scoped_dsn)
    monkeypatch.setenv("SECRET_KEY", "test-secret-key-not-for-production")

    from app import storage
    storage._conn = None  # force reconnect against the schema-scoped DSN

    from app.main import app
    with TestClient(app) as c:
        yield c

    storage._conn = None
    admin_conn = psycopg.connect(TEST_DATABASE_URL, autocommit=True)
    admin_conn.execute(f'DROP SCHEMA "{schema}" CASCADE')
    admin_conn.close()


def _signup(c: TestClient, practice_name="Acme & Co", admin_name="Alex Partner",
            admin_email="alex@acme.test", admin_password="hunter2hunter") -> str:
    """Creates a practice + its first (partner) user and logs the client in
    via the session cookie, exactly as a real signup would. Returns the new
    practice_id."""
    resp = c.post(
        "/practices",
        data={
            "practice_name": practice_name, "admin_name": admin_name,
            "admin_email": admin_email, "admin_password": admin_password,
        },
        follow_redirects=False,
    )
    assert resp.status_code == 303
    return resp.headers["location"].rsplit("/", 1)[-1]


def _make_template_bytes() -> bytes:
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cover Page"
    ws["A1"] = "FIRM TEMPLATE COVER"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_storage_entity_and_file_roundtrip(http_client):
    """Direct storage.py calls: entities and file BYTEA both survive a
    reconnect-free roundtrip through the real database."""
    from app import storage

    practice = storage.create_practice("Test Practice")
    assert practice["id"]
    assert storage.get_practice(practice["id"])["name"] == "Test Practice"

    file_id = storage.save_file("upload", None, "sample.txt", b"hello world", content_type="text/plain")
    assert storage.load_file(file_id) == b"hello world"

    storage.save_mapping_profile("client_x", "trial_balance", "xero", {"Account": "account_name"})
    assert storage.load_mapping_profile("client_x", "trial_balance", "xero") == {"Account": "account_name"}
    assert storage.load_mapping_profile("client_x", "trial_balance", "sage") is None


def test_full_http_flow_practice_to_download(http_client):
    """signup (practice + partner login) -> template upload (normalised) ->
    client -> job -> report upload -> generate -> download, all over real
    HTTP requests against a real Postgres-backed app instance, as the
    logged-in partner."""
    c = http_client
    practice_id = _signup(c)

    resp = c.post(
        f"/practices/{practice_id}/templates",
        data={"name": "Standard Template"},
        files={"file": ("template.xlsx", _make_template_bytes(),
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 303

    from app import storage
    templates_list = storage.list_templates(practice_id)
    assert len(templates_list) == 1
    template_id = templates_list[0]["id"]
    assert templates_list[0]["normalisation"]["normalised"] is True

    resp = c.post(
        f"/practices/{practice_id}/clients",
        data={"name": "Sample Client Ltd", "template_id": template_id},
        follow_redirects=False,
    )
    assert resp.status_code == 303
    client_id = resp.headers["location"].rsplit("/", 1)[-1]

    resp = c.post(
        f"/clients/{client_id}/jobs",
        data={"current_period_start": "2024-01-01", "current_period_end": "2024-12-31"},
        follow_redirects=False,
    )
    assert resp.status_code == 303
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    tb_path = SAMPLE_DIR / "trial_balance_current_xero.xlsx"
    with open(tb_path, "rb") as fh:
        resp = c.post(
            f"/jobs/{job_id}/uploads",
            files={"files": (tb_path.name, fh,
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            follow_redirects=False,
        )
    assert resp.status_code == 303  # Xero-native TB auto-confirms straight through, no mapping-confirm redirect

    job = storage.get_job(job_id)
    assert len(job["uploads"]) == 1
    upload = next(iter(job["uploads"].values()))
    assert storage.load_file(upload["file_id"]) is not None

    resp = c.post(f"/jobs/{job_id}/generate", follow_redirects=False)
    assert resp.status_code == 303

    resp = c.get(f"/jobs/{job_id}/download")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert len(resp.content) > 1000

    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Cover Page" in wb.sheetnames  # the template's own sheet survived untouched
    assert wb["Cover Page"]["A1"].value == "FIRM TEMPLATE COVER"


def test_unauthenticated_request_redirects_to_login(http_client):
    c = http_client
    resp = c.get("/practices/practice_doesnotmatter", follow_redirects=False)
    assert resp.status_code == 303
    assert resp.headers["location"].startswith("/login")


def test_wrong_password_rejected(http_client):
    c = http_client
    _signup(c, admin_email="owner@acme.test", admin_password="correct-horse-battery")
    resp = c.post("/login", data={"email": "owner@acme.test", "password": "wrong"}, follow_redirects=False)
    assert resp.status_code == 400
    assert "Invalid email or password" in resp.text


def test_preparer_scoped_to_granted_clients_only(http_client):
    """The core RBAC guarantee: a preparer can see/act on a client they've
    been granted access to, and gets a 403 - not the client's data - for
    one they haven't, even though both clients are in the same practice."""
    c = http_client
    practice_id = _signup(c, admin_email="partner@acme.test")

    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Client A"}, follow_redirects=False)
    client_a_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Client B"}, follow_redirects=False)
    client_b_id = resp.headers["location"].rsplit("/", 1)[-1]

    resp = c.post(f"/practices/{practice_id}/users", data={
        "name": "Prep One", "email": "prep@acme.test", "password": "prepper-pass", "role": "preparer",
    }, follow_redirects=False)
    assert resp.status_code == 303

    from app import storage
    prep_user = storage.get_user_by_email("prep@acme.test")
    resp = c.post(
        f"/practices/{practice_id}/users/{prep_user['id']}/client-access",
        data={"client_ids": [client_a_id]}, follow_redirects=False,
    )
    assert resp.status_code == 303

    # log out the partner, log in as the preparer
    c.post("/logout")
    resp = c.post("/login", data={"email": "prep@acme.test", "password": "prepper-pass"}, follow_redirects=False)
    assert resp.status_code == 303

    assert c.get(f"/clients/{client_a_id}").status_code == 200
    resp = c.get(f"/clients/{client_b_id}", follow_redirects=False)
    assert resp.status_code == 403

    # preparer's client list only shows the granted client
    resp = c.get(f"/practices/{practice_id}/clients")
    assert "Client A" in resp.text
    assert "Client B" not in resp.text

    # preparer role restrictions: no creating clients, no managing users/templates
    assert c.post(f"/practices/{practice_id}/clients", data={"name": "Client C"}, follow_redirects=False).status_code == 403
    assert c.get(f"/practices/{practice_id}/users", follow_redirects=False).status_code == 403
    assert c.post(f"/practices/{practice_id}/templates", data={"name": "X"}, follow_redirects=False).status_code == 403


def test_manager_sees_all_clients_but_cannot_manage_users(http_client):
    c = http_client
    practice_id = _signup(c, admin_email="partner2@acme.test")
    c.post(f"/practices/{practice_id}/clients", data={"name": "Only Client"}, follow_redirects=False)

    c.post(f"/practices/{practice_id}/users", data={
        "name": "Mgr One", "email": "mgr@acme.test", "password": "manager-pass", "role": "manager",
    }, follow_redirects=False)

    c.post("/logout")
    c.post("/login", data={"email": "mgr@acme.test", "password": "manager-pass"}, follow_redirects=False)

    resp = c.get(f"/practices/{practice_id}/clients")
    assert "Only Client" in resp.text  # managers see every client, no grant needed

    assert c.get(f"/practices/{practice_id}/users", follow_redirects=False).status_code == 403


def test_cross_practice_access_denied(http_client):
    """A user from one practice can't reach another practice's pages, even
    though both exist in the same database."""
    c = http_client
    practice_a_id = _signup(c, admin_email="a@firm-a.test")

    from app.main import app as fastapi_app
    other = TestClient(fastapi_app)
    resp = other.post("/practices", data={
        "practice_name": "Firm B", "admin_name": "B Partner",
        "admin_email": "b@firm-b.test", "admin_password": "firm-b-password",
    }, follow_redirects=False)
    assert resp.status_code == 303

    resp = other.get(f"/practices/{practice_a_id}", follow_redirects=False)
    assert resp.status_code == 404


def _make_pdf_trial_balance() -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

    rows = [
        ["Account Code", "Account Name", "Debit", "Credit"],
        ["1000", "Bank Current Account", "12000.00", ""],
        ["4000", "Sales", "", "75000.00"],
    ]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    table = Table(rows)
    table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    doc.build([table])
    return buf.getvalue()


def test_bulk_upload_mixed_files_auto_detect_and_confirm_chain(http_client):
    """The real workflow this feature exists for: drop in several files of
    different kinds and formats at once (a genuine Xero export, a generic
    CSV, and a PDF) with no report-type/platform/period picked up front.
    The Xero file should auto-confirm silently (real parser validation, not
    a guess); the other two should get queued for a quick confirm each,
    chained one after another; and the PDF (uploaded alongside a current-
    period TB) should be guessed as the comparative TB via the "second
    upload of this type" heuristic, since it has no period text of its own."""
    pytest.importorskip("reportlab", reason="reportlab is a dev-only dependency for building test PDFs")
    c = http_client
    practice_id = _signup(c, admin_email="bulk@acme.test")

    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Bulk Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
        "comparative_period_start": "2024-01-01", "comparative_period_end": "2024-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    tb_path = SAMPLE_DIR / "trial_balance_current_xero.xlsx"
    bank_path = SAMPLE_DIR / "bank_statement_current.csv"
    files_payload = [
        ("files", (tb_path.name, tb_path.read_bytes(),
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ("files", (bank_path.name, bank_path.read_bytes(), "text/csv")),
        ("files", ("prior_year_tb.pdf", _make_pdf_trial_balance(), "application/pdf")),
    ]
    resp = c.post(f"/jobs/{job_id}/uploads", files=files_payload, follow_redirects=False)
    assert resp.status_code == 303

    from app import storage
    job = storage.get_job(job_id)
    by_filename = {u["filename"]: u for u in job["uploads"].values()}
    assert by_filename[tb_path.name]["confirmed"] is True  # Xero-native: no confirm step needed
    assert by_filename[tb_path.name]["report_type"] == "trial_balance"
    assert by_filename[bank_path.name]["report_type"] == "bank_statement"
    assert by_filename["prior_year_tb.pdf"]["report_type"] == "trial_balance"

    # walk the confirm chain the redirect points at, for every unconfirmed upload
    location = resp.headers["location"]
    hops = 0
    while location and "/mapping" in location and hops < 5:
        hops += 1
        assert c.get(location).status_code == 200
        resp = c.post(location, data={"action": "confirm"}, follow_redirects=False)
        assert resp.status_code == 303
        location = resp.headers["location"]
    assert hops == 2  # bank statement + PDF both needed a confirm; the Xero file didn't

    job = storage.get_job(job_id)
    assert all(u["confirmed"] for u in job["uploads"].values())
    by_filename = {u["filename"]: u for u in job["uploads"].values()}
    assert by_filename["prior_year_tb.pdf"]["period"] == "comparative"

    resp = c.post(f"/jobs/{job_id}/generate", follow_redirects=False)
    assert resp.status_code == 303
    resp = c.get(f"/jobs/{job_id}/download")
    assert resp.status_code == 200
    assert len(resp.content) > 1000


def _make_multisheet_vat_workbook() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Box 1", "Box 2", "Box 3", "Box 4", "Box 5", "Box 6", "Box 7", "Box 8", "Box 9"])
    summary.append([1000, 0, 1000, 200, 800, 15000, 5000, 0, 0])
    detail = wb.create_sheet("Detail")
    detail.append(["Date", "Account Code", "Description", "Debit", "Credit"])
    detail.append(["01/03/2025", "4000", "Sales invoice 101", "", "1000"])
    detail.append(["05/03/2025", "5000", "Purchase invoice 55", "500", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_multisheet_workbook_expands_into_one_upload_per_sheet(http_client):
    """A VAT return export with separate Summary/Detail tabs used to be
    silently reduced to whichever sheet pandas reads by default - every
    sheet should now become its own classified sub-upload."""
    c = http_client
    practice_id = _signup(c, admin_email="multisheet@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Multisheet Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    resp = c.post(
        f"/jobs/{job_id}/uploads",
        files={"files": ("vat_return.xlsx", _make_multisheet_vat_workbook(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 303

    from app import storage
    job = storage.get_job(job_id)
    assert len(job["uploads"]) == 2
    by_sheet = {u["sheet_name"]: u for u in job["uploads"].values()}
    assert by_sheet["Summary"]["report_type"] == "vat_return"
    assert by_sheet["Detail"]["report_type"] == "nominal_activity"
    assert by_sheet["Summary"]["display_name"] == "vat_return.xlsx (Summary)"


def test_report_notes_and_section_scoped_upload(http_client):
    """The job page shows a fixed section per report type, each with an
    editable instruction note (persisted per client, not per job) and its
    own upload form - a file dropped into a section is classified as that
    section's report type directly, no guessing needed for that part."""
    c = http_client
    practice_id = _signup(c, admin_email="sections@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Sections Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    resp = c.get(f"/jobs/{job_id}")
    assert resp.status_code == 200
    assert "VAT Return" in resp.text
    assert "Add VAT Return file(s)" in resp.text

    resp = c.post(f"/clients/{client_id}/notes/vat_return", data={
        "note": "Detail tab has the transaction-level data.", "next": f"/jobs/{job_id}",
    }, follow_redirects=False)
    assert resp.status_code == 303
    assert resp.headers["location"] == f"/jobs/{job_id}"

    from app import storage
    client = storage.get_client(client_id)
    assert client["report_notes"]["vat_return"] == "Detail tab has the transaction-level data."
    assert "Instruction note (set)" in c.get(f"/jobs/{job_id}").text

    bank_path = SAMPLE_DIR / "bank_statement_current.csv"
    resp = c.post(
        f"/jobs/{job_id}/uploads",
        data={"section_report_type": "bank_statement"},
        files={"files": (bank_path.name, bank_path.read_bytes(), "text/csv")},
        follow_redirects=False,
    )
    assert resp.status_code == 303
    job = storage.get_job(job_id)
    upload = next(iter(job["uploads"].values()))
    assert upload["report_type"] == "bank_statement"


def test_ai_reconciliation_note_reaches_the_generated_workbook(http_client, monkeypatch):
    """Full wiring, end to end: template config opts in -> generate() calls
    the (mocked) agent for a flagged check -> the note lands on that
    check's sheet in the actual downloaded workbook. The Anthropic API is
    never really called in tests - see test_reconciliation_agent.py for
    the agent's own unit tests; this just proves main.py -> excel_builder
    wiring works, using a mock at the same patch point."""
    from unittest.mock import MagicMock, patch

    c = http_client
    monkeypatch.setenv("ANTHROPIC_API_KEY", "test-key")
    practice_id = _signup(c, admin_email="ainotes@acme.test")

    resp = c.post(
        f"/practices/{practice_id}/templates",
        data={"name": "AI Notes Template"},
        files={"file": ("template.xlsx", _make_template_bytes(),
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    from app import storage
    template = storage.list_templates(practice_id)[0]
    template["config"]["ai_reconciliation_notes"]["enabled"] = True
    storage.save_template(template)

    resp = c.post(f"/practices/{practice_id}/clients", data={
        "name": "AI Notes Client", "template_id": template["id"],
    }, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    resp = c.post(f"/clients/{client_id}/notes/vat_return", data={
        "note": "Detail tab has the transaction-level data.", "next": f"/jobs/{job_id}",
    }, follow_redirects=False)
    assert resp.status_code == 303

    # a VAT return with no TB/P&L uploaded alongside it - guaranteed to
    # flag (turnover and the VAT control balance both default to 0)
    vat_csv = b"Box 1,Box 2,Box 3,Box 4,Box 5,Box 6,Box 7,Box 8,Box 9\n1000,0,1000,200,800,15000,5000,0,0\n"
    resp = c.post(
        f"/jobs/{job_id}/uploads", data={"section_report_type": "vat_return"},
        files={"files": ("vat.csv", vat_csv, "text/csv")}, follow_redirects=False,
    )
    location = resp.headers["location"]
    assert "/mapping" in location
    upload_id = location.rsplit("/", 2)[-2]
    job = storage.get_job(job_id)
    upload = job["uploads"][upload_id]
    # real form submissions resubmit the pre-filled (auto-suggested)
    # column mapping; a bare {"action": "confirm"} would silently map
    # every column to None, zeroing out the VAT figures and defeating
    # the point of this test (nothing would ever flag).
    confirm_data = {"action": "confirm"}
    confirm_data.update({f"col__{col}": field for col, field in (upload["mapping"] or {}).items() if field})
    resp = c.post(location, data=confirm_data, follow_redirects=False)
    assert resp.status_code == 303

    mock_client = MagicMock()
    block = MagicMock()
    block.type = "text"
    block.text = "The £800 Box 5 figure has no nominal ledger postings to compare against - nothing was uploaded."
    response = MagicMock()
    response.content = [block]
    mock_client.messages.create.return_value = response

    with patch("anthropic.Anthropic", return_value=mock_client):
        resp = c.post(f"/jobs/{job_id}/generate", follow_redirects=False)
    assert resp.status_code == 303
    assert mock_client.messages.create.called

    resp = c.get(f"/jobs/{job_id}/download")
    assert resp.status_code == 200

    import io as _io

    import openpyxl
    wb = openpyxl.load_workbook(_io.BytesIO(resp.content))
    vat_sheet_names = [s for s in wb.sheetnames if "VAT" in s.upper()]
    assert vat_sheet_names, wb.sheetnames
    sheet_text = "\n".join(
        str(cell.value) for row in wb[vat_sheet_names[0]].iter_rows() for cell in row if cell.value
    )
    assert "AI-ASSISTED NOTE" in sheet_text
    assert "no nominal ledger postings" in sheet_text


def test_generate_progress_stream_reports_all_ten_steps_in_order(http_client):
    """The SSE progress endpoint (added so the person generating a
    working paper can see what's happening instead of staring at a
    blocked page) must report every one of the ten real steps, in the
    right order, ending with a completion event - and must leave the job
    in exactly the same generated state as the classic POST route does."""
    c = http_client
    practice_id = _signup(c, admin_email="progress@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Progress Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    resp = c.get(f"/jobs/{job_id}/generate/progress")
    assert resp.status_code == 200
    assert "EventSource" in resp.text

    resp = c.get(f"/jobs/{job_id}/generate/stream")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/event-stream")

    import json as _json
    events = [
        _json.loads(line[len("data: "):])
        for line in resp.text.split("\n") if line.startswith("data: ")
    ]
    assert events[-1]["step"] == "complete"
    assert events[-1]["redirect"] == f"/jobs/{job_id}"
    assert "at" in events[-1]

    from app.main import GENERATE_STEPS
    step_events = [e for e in events[:-1] if e["step"] != 5]  # step 5 (AI notes) is off by default -> "skipped" only
    seen_steps = [e["step"] for e in step_events]
    assert seen_steps == sorted(seen_steps)  # strictly non-decreasing: steps arrive in order
    for n in range(1, len(GENERATE_STEPS) + 1):
        if n == 5:
            assert any(e["step"] == 5 and e["status"] == "skipped" for e in events)
            continue
        assert any(e["step"] == n and e["status"] == "running" for e in events)
        assert any(e["step"] == n and e["status"] == "done" for e in events)

    from app import storage
    job = storage.get_job(job_id)
    assert job["status"] == "generated"
    assert job["output_file_id"]


def test_generate_progress_is_persisted_and_summarized_with_timings(http_client):
    """Progress events aren't just streamed live and thrown away - each one
    is saved onto the job as it happens, so a page loaded after the run (or
    after a stream that nobody watched) can still show what happened and how
    long each step took. This also covers the classic synchronous POST route,
    which shares the same generator and must persist identically."""
    c = http_client
    practice_id = _signup(c, admin_email="persist-progress@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Persisted Progress Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    resp = c.post(f"/jobs/{job_id}/generate", follow_redirects=False)
    assert resp.status_code == 303

    from app import storage
    from app.main import GENERATE_STEPS, _summarize_progress

    job = storage.get_job(job_id)
    progress = job.get("progress")
    assert progress, "generate() must persist progress events onto the job even on the classic POST route"
    assert progress[-1] == {"step": "complete", "redirect": f"/jobs/{job_id}", "at": progress[-1]["at"]}
    for evt in progress:
        assert "at" in evt

    summary = _summarize_progress(progress)
    assert summary["finished"] is True
    assert summary["error_message"] is None
    assert summary["total_seconds"] is not None and summary["total_seconds"] >= 0

    rows_by_step = {r["step"]: r for r in summary["rows"]}
    for n in range(1, len(GENERATE_STEPS) + 1):
        row = rows_by_step[n]
        assert row["label"] == GENERATE_STEPS[n - 1]
        if n == 5:  # AI notes: off by default -> skipped, no duration
            assert row["status"] == "skipped"
            assert row["seconds"] is None
        else:
            assert row["status"] == "done"
            assert row["seconds"] is not None and row["seconds"] >= 0

    resp = c.get(f"/jobs/{job_id}")
    assert resp.status_code == 200
    assert "Last generation" in resp.text
    assert f"finished in {summary['total_seconds']:.1f}s" in resp.text


def test_run_step_with_retry_recovers_from_a_transient_failure(monkeypatch):
    """The retry helper itself (used by steps 1 and 10, the only two that
    touch Postgres): a step whose work raises once and then succeeds must
    end up "done" with the right result, having emitted exactly one
    "retrying" event in between - and never sleep for real in tests."""
    from app import main

    monkeypatch.setattr(main.time, "sleep", lambda *_: None)

    calls = {"n": 0}

    def flaky():
        calls["n"] += 1
        if calls["n"] == 1:
            raise ConnectionError("simulated transient DB blip")
        return "workbook built"

    events = []

    def event_fn(n, status, **extra):
        evt = {"step": n, "status": status, **extra}
        events.append(evt)
        return evt

    gen = main._run_step_with_retry(event_fn, 10, flaky)
    result = None
    try:
        while True:
            next(gen)
    except StopIteration as stop:
        result = stop.value

    assert result == "workbook built"
    assert calls["n"] == 2
    statuses = [e["status"] for e in events]
    assert statuses == ["running", "retrying", "done"]
    assert events[1]["attempt"] == 1
    assert events[1]["max_attempts"] == main.STEP_RETRY_ATTEMPTS
    assert "simulated transient DB blip" in events[1]["error"]


def test_run_step_with_retry_gives_up_after_max_attempts(monkeypatch):
    """A persistent (not transient) failure must still surface as a real
    error once every attempt is exhausted - retry must never silently
    swallow a genuine failure."""
    from app import main

    monkeypatch.setattr(main.time, "sleep", lambda *_: None)

    calls = {"n": 0}

    def always_fails():
        calls["n"] += 1
        raise ConnectionError("still down")

    events = []

    def event_fn(n, status, **extra):
        evt = {"step": n, "status": status, **extra}
        events.append(evt)
        return evt

    gen = main._run_step_with_retry(event_fn, 1, always_fails)
    with pytest.raises(ConnectionError, match="still down"):
        list(gen)

    assert calls["n"] == main.STEP_RETRY_ATTEMPTS
    statuses = [e["status"] for e in events]
    assert statuses == ["running"] + ["retrying"] * (main.STEP_RETRY_ATTEMPTS - 1)


def test_generate_recovers_from_a_transient_storage_blip(http_client, monkeypatch):
    """End-to-end: a real generate run where step 1's storage.load_file
    call fails once (simulating a transient Postgres blip) must still
    finish successfully via one retry - and the retry must be visible in
    the persisted progress trail, not just silently absorbed."""
    c = http_client
    practice_id = _signup(c, admin_email="retry-recover@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Retry Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    tb_path = SAMPLE_DIR / "trial_balance_current_xero.xlsx"
    with open(tb_path, "rb") as fh:
        resp = c.post(
            f"/jobs/{job_id}/uploads",
            files={"files": (tb_path.name, fh,
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            follow_redirects=False,
        )
    assert resp.status_code == 303  # Xero-native TB auto-confirms

    from app import main, storage

    monkeypatch.setattr(main.time, "sleep", lambda *_: None)
    original_load_file = storage.load_file
    calls = {"n": 0}

    def flaky_load_file(file_id):
        calls["n"] += 1
        if calls["n"] == 1:
            raise ConnectionError("simulated transient DB blip")
        return original_load_file(file_id)

    monkeypatch.setattr(storage, "load_file", flaky_load_file)

    resp = c.post(f"/jobs/{job_id}/generate", follow_redirects=False)
    assert resp.status_code == 303
    assert calls["n"] == 2  # one failed attempt, one that succeeded

    job = storage.get_job(job_id)
    assert job["status"] == "generated"
    retry_events = [e for e in job["progress"] if e.get("status") == "retrying"]
    assert len(retry_events) == 1
    assert retry_events[0]["step"] == 1
    assert "simulated transient DB blip" in retry_events[0]["error"]

    summary = main._summarize_progress(job["progress"])
    step1_row = next(r for r in summary["rows"] if r["step"] == 1)
    assert step1_row["status"] == "done"
    assert step1_row["retries"] == 1


def test_summarize_progress_reports_failure(http_client):
    """If a run blew up partway through, the summary must say so plainly
    rather than silently reporting 'finished' - this is what the job detail
    page's failure banner is driven from."""
    from app.main import _summarize_progress

    progress = [
        {"step": 1, "total": 10, "label": "Load data", "status": "running", "at": "2026-01-01T00:00:00+00:00"},
        {"step": 1, "total": 10, "label": "Load data", "status": "done", "at": "2026-01-01T00:00:01+00:00"},
        {"step": 2, "total": 10, "label": "Reconcile", "status": "running", "at": "2026-01-01T00:00:01+00:00"},
        {"step": "error", "message": "boom", "at": "2026-01-01T00:00:02+00:00"},
    ]
    summary = _summarize_progress(progress)
    assert summary["finished"] is False
    assert summary["error_message"] == "boom"
    assert summary["total_seconds"] is None
    row2 = next(r for r in summary["rows"] if r["step"] == 2)
    assert row2["status"] == "running"
    assert row2["seconds"] is None


def test_summarize_progress_empty_is_none(http_client):
    from app.main import _summarize_progress
    assert _summarize_progress(None) is None
    assert _summarize_progress([]) is None


def test_generate_stream_unauthenticated_redirects_to_login(http_client):
    c = http_client
    c.cookies.clear()
    resp = c.get("/jobs/job_doesnotmatter/generate/stream", follow_redirects=False)
    assert resp.status_code == 303
    assert resp.headers["location"].startswith("/login")


def test_upload_route_rejects_confirm_with_no_report_type(http_client):
    """Guard against silently confirming an upload the system couldn't
    classify and the user didn't pick a type for either - that would mark
    it 'confirmed' with an empty mapping, contributing nothing to the
    generated workbook with no visible sign anything was wrong."""
    c = http_client
    practice_id = _signup(c, admin_email="guard@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Guard Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    resp = c.post(
        f"/jobs/{job_id}/uploads",
        files={"files": ("mystery.csv", b"Foo,Bar\n1,2\n", "text/csv")},
        follow_redirects=False,
    )
    location = resp.headers["location"]
    assert "/mapping" in location

    from app import storage
    job = storage.get_job(job_id)
    upload = next(iter(job["uploads"].values()))
    assert upload["report_type"] == ""

    resp = c.post(location, data={"action": "confirm"}, follow_redirects=False)
    assert resp.status_code == 303
    assert resp.headers["location"] == location  # bounced back, not confirmed

    job = storage.get_job(job_id)
    upload = next(iter(job["uploads"].values()))
    assert upload["confirmed"] is False


def _vat_recon_workbook(rows: list[list]) -> bytes:
    """A simple single-sheet Excel file with headers that match the VAT
    reconciliation alias dictionaries (mapping.ALIASES["vat_gl"] etc.)
    closely enough for auto-suggestion to map every column on its own."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Reference", "Contact", "Description", "Net Amount", "VAT Amount"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _confirm_all_pending_uploads(c: TestClient, job_id: str) -> None:
    """Walks every unconfirmed upload's mapping-confirm page and resubmits
    its auto-suggested mapping (real column headers here match the alias
    dictionaries, so nothing needs a manual override) - same pattern as
    the other tests that need a confirmed upload before generating."""
    from app import storage as storage_module

    while True:
        job = storage_module.get_job(job_id)
        pending = next((u for u in job["uploads"].values() if not u["confirmed"]), None)
        if not pending:
            return
        confirm_data = {"action": "confirm"}
        confirm_data.update({f"col__{col}": field for col, field in (pending["mapping"] or {}).items() if field})
        resp = c.post(f"/jobs/{job_id}/uploads/{pending['id']}/mapping", data=confirm_data, follow_redirects=False)
        assert resp.status_code == 303


def test_vat_reconciliation_end_to_end_through_http_and_into_the_workbook(http_client):
    """Full VAT Reconciliation workspace flow over real HTTP requests: a
    General Ledger upload, two Filed VAT Return detail uploads (Box 1 and
    Box 4, each simulating the "combine multiple files" case with two
    files apiece), settings, the standalone Run button, and finally that
    the same results land in the generated workbook (see main.py's step
    2, which folds vat_reconciliation.reconcile() into the main results
    list) - proving this section works both on its own and as part of
    the full pipeline."""
    c = http_client
    practice_id = _signup(c, admin_email="vat-recon@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "VAT Recon Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    gl_bytes = _vat_recon_workbook([
        ["2025-01-15", "INV-100", "Acme Ltd", "Sale", 1000, 200],   # matches Box 1 filed item exactly
        ["2025-01-20", "INV-101", "Beta Ltd", "Sale", 500, 100],    # matches by reference, VAT amount differs (variance)
        ["2025-02-01", "BILL-200", "Gamma Supplies", "Purchase", 800, 160],  # matches Box 4 filed item exactly
        ["2025-03-01", "MYSTERY-1", "Nobody Ltd", "Unexplained", 50, 10],    # neither box's filed return mentions this
    ])
    resp = c.post(
        f"/jobs/{job_id}/uploads", data={"section_report_type": "vat_gl"},
        files={"files": ("gl.xlsx", gl_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 303

    filed_sales_q1 = _vat_recon_workbook([["2025-01-15", "INV-100", "Acme Ltd", "", 1000, 200]])
    filed_sales_q2 = _vat_recon_workbook([["2025-01-20", "INV-101", "Beta Ltd", "", 500, 105]])  # £5 variance vs GL
    resp = c.post(
        f"/jobs/{job_id}/uploads", data={"section_report_type": "vat_filed_sales"},
        files=[
            ("files", ("filed_sales_q1.xlsx", filed_sales_q1, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("files", ("filed_sales_q2.xlsx", filed_sales_q2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
        follow_redirects=False,
    )
    assert resp.status_code == 303

    filed_purchases = _vat_recon_workbook([["2025-02-01", "BILL-200", "Gamma Supplies", "", 800, 160]])
    resp = c.post(
        f"/jobs/{job_id}/uploads", data={"section_report_type": "vat_filed_purchases"},
        files={"files": ("filed_purchases.xlsx", filed_purchases, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 303

    from app import storage
    job = storage.get_job(job_id)
    assert len(job["uploads"]) == 4  # 1 GL + 2 filed sales files + 1 filed purchases file
    by_type = {}
    for u in job["uploads"].values():
        by_type.setdefault(u["report_type"], []).append(u)
    assert len(by_type["vat_gl"]) == 1
    assert len(by_type["vat_filed_sales"]) == 2
    assert len(by_type["vat_filed_purchases"]) == 1

    _confirm_all_pending_uploads(c, job_id)
    job = storage.get_job(job_id)
    assert all(u["confirmed"] for u in job["uploads"].values())

    resp = c.post(f"/jobs/{job_id}/vat-recon-settings", data={"vat_recon_basis": "accrual", "vat_recon_tolerance": "0.0"}, follow_redirects=False)
    assert resp.status_code == 303

    resp = c.post(f"/jobs/{job_id}/vat-recon/run", follow_redirects=False)
    assert resp.status_code == 303

    job = storage.get_job(job_id)
    results = {r["name"]: r for r in job["vat_recon_results"]}
    assert set(results) == {
        "VAT Recon - Box 1 (Sales)", "VAT Recon - Box 4 (Purchases)", "VAT Recon - General Ledger Coverage",
        "VAT Recon - suggested box for unmatched General Ledger items",
    }
    assert results["VAT Recon - Box 1 (Sales)"]["status"] == "review"  # INV-101 variance
    assert results["VAT Recon - Box 4 (Purchases)"]["status"] == "ok"
    assert results["VAT Recon - General Ledger Coverage"]["status"] == "review"  # MYSTERY-1 unaccounted for
    box1_review = results["VAT Recon - Box 1 (Sales)"]["extra_detail"]
    assert any(row["Reference"] == "INV-101" and row["Exception type"] == "Matched but VAT amount differs" for row in box1_review)

    resp = c.get(f"/jobs/{job_id}")
    assert resp.status_code == 200
    assert "VAT Reconciliation" in resp.text
    assert "INV-101" in resp.text

    resp = c.post(f"/jobs/{job_id}/generate", follow_redirects=False)
    assert resp.status_code == 303

    job = storage.get_job(job_id)
    summary_by_name = {s["name"]: s for s in job["summary"]}
    assert summary_by_name["VAT Recon - Box 1 (Sales)"]["status"] == "review"
    assert summary_by_name["VAT Recon - Box 4 (Purchases)"]["status"] == "ok"

    import openpyxl as _openpyxl
    resp = c.get(f"/jobs/{job_id}/download")
    assert resp.status_code == 200
    wb = _openpyxl.load_workbook(io.BytesIO(resp.content))
    assert any("VAT Recon - Box 1" in name for name in wb.sheetnames)
    assert any("VAT Recon - Box 4" in name for name in wb.sheetnames)


def test_vat_reconciliation_cash_basis_combination_matching_through_http(http_client):
    """Cash-basis combination matching (vat_reconciliation.match_box
    passes 4 and 5 - several GL legs summing to one filed row, and the
    mirror case, several filed rows summing to one GL row) had only ever
    been exercised directly against hand-built DataFrames before, never
    through the real upload -> mapping-confirm -> run HTTP path. Also
    covers a real bug found the same way: a filed return row with no date
    column at all (this test's third invoice) leaves `date` as `pd.NaT`
    after apply_mapping - a perfectly ordinary real-world event (not
    every export tracks an invoice date the same way) - and main.py's
    _jsonable helper checked `isinstance(value, pd.Timestamp)` to convert
    dates for JSONB storage, but pd.NaT is its own singleton type, never
    a Timestamp subclass, so that bare NaT reached json.dumps() and
    crashed the whole save with an unhandled 500 instead of completing
    the run. See test_jsonable.py for the fix's direct unit tests; this
    test locks in the fix at the HTTP layer, where it actually surfaced,
    alongside confirming both combination directions work."""
    c = http_client
    practice_id = _signup(c, admin_email="vat-combo@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "VAT Combo Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    # Acme Ltd: 3 partial GL receipts (400+350+250=1000) that only tie out
    # combined -> Pass 4 (several GL legs -> one filed Box 1 row).
    # Beta Supplies: one combined GL payment (600) that is the sum of two
    # separate filed Box 4 invoices (250+350) -> Pass 5 (one GL row ->
    # several filed rows). Echo Ltd: a plain reference match (no
    # combination involved) whose filed invoice has no date at all.
    gl_bytes = _vat_recon_workbook([
        ["2025-06-05", "RCPT-A1", "Acme Ltd", "Part payment 1", 2000, 400],
        ["2025-06-12", "RCPT-A2", "Acme Ltd", "Part payment 2", 1750, 350],
        ["2025-06-19", "RCPT-A3", "Acme Ltd", "Part payment 3", 1250, 250],
        ["2025-06-22", "PAY-BETA-1", "Beta Supplies", "Combined supplier payment", 3000, 600],
        ["2025-06-25", "INV-E700", "Echo Ltd", "Sale", 300, 60],
    ])
    resp = c.post(
        f"/jobs/{job_id}/uploads", data={"section_report_type": "vat_gl"},
        files={"files": ("gl.xlsx", gl_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 303

    filed_sales = _vat_recon_workbook([["2025-06-01", "INV-A500", "Acme Ltd", "", 5000, 1000]])
    resp = c.post(
        f"/jobs/{job_id}/uploads", data={"section_report_type": "vat_filed_sales"},
        files={"files": ("filed_sales.xlsx", filed_sales, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 303

    # No Date column at all - some filed-return exports genuinely don't
    # carry one at invoice-line grain - so `date` comes back as pd.NaT
    # for this row once apply_mapping defaults and parses it.
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Reference", "Customer", "Description", "Net Amount", "VAT Amount"])
    ws.append(["INV-E700", "Echo Ltd", "", 300, 60])
    buf = io.BytesIO()
    wb.save(buf)
    filed_sales_no_date = buf.getvalue()
    resp = c.post(
        f"/jobs/{job_id}/uploads", data={"section_report_type": "vat_filed_sales"},
        files={"files": ("filed_sales_no_date.xlsx", filed_sales_no_date, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 303

    filed_purchases = _vat_recon_workbook([
        ["2025-06-15", "PINV-1", "Beta Supplies", "", 1250, 250],
        ["2025-06-18", "PINV-2", "Beta Supplies", "", 1750, 350],
    ])
    resp = c.post(
        f"/jobs/{job_id}/uploads", data={"section_report_type": "vat_filed_purchases"},
        files={"files": ("filed_purchases.xlsx", filed_purchases, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 303

    _confirm_all_pending_uploads(c, job_id)

    resp = c.post(f"/jobs/{job_id}/vat-recon-settings", data={"vat_recon_basis": "cash", "vat_recon_tolerance": "0.0"}, follow_redirects=False)
    assert resp.status_code == 303

    resp = c.post(f"/jobs/{job_id}/vat-recon/run", follow_redirects=False)
    assert resp.status_code == 303, "this used to 500 - a blank/NaT filed-invoice date crashed save_job()"

    from app import storage
    job = storage.get_job(job_id)
    results = {r["name"]: r for r in job["vat_recon_results"]}

    box1 = results["VAT Recon - Box 1 (Sales)"]
    assert box1["status"] == "ok"
    assert len(box1["matched_detail"]) == 2
    by_ref = {r["Filed reference"]: r for r in box1["matched_detail"]}
    assert "combined GL postings" in by_ref["INV-A500"]["Match basis"]
    assert by_ref["INV-A500"]["GL VAT"] == 1000.0
    assert by_ref["INV-E700"]["Match basis"] == "reference"
    assert by_ref["INV-E700"]["Filed date"] is None, "the missing date should serialize as None, not crash"

    box4 = results["VAT Recon - Box 4 (Purchases)"]
    assert box4["status"] == "ok"
    assert len(box4["matched_detail"]) == 1
    assert "combined filed items" in box4["matched_detail"][0]["Match basis"]
    assert box4["matched_detail"][0]["Filed reference"] == "PINV-1; PINV-2"


def _make_nominal_matrix_gl_workbook() -> bytes:
    """A Xero-native 'Account Transactions' export with two sections:
    'Repairs and Maintenance' (the nominal matrix's subject, 14 rows
    exercising the OTHER-bucket top-10 cap, a multi-code-split 'and N
    more' related-account entry, an unallocated entry, and a contact with
    a dominant allocation history for the suggestion engine) and 'Motor
    Expenses' (one clean row, a contrast that should come back "ok")."""
    from datetime import datetime as _dt

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Account Transactions"
    ws.append(["Account Transactions"])
    ws.append(["Matrix Test Co"])
    ws.append(["For the period 1 January 2025 to 31 December 2025"])
    ws.append(["Date", "Source", "Contact", "Contact Group", "Description", "Invoice Number", "Reference",
                "Debit", "Credit", "Running Balance", "Gross", "Net", "VAT", "VAT Rate", "VAT Rate Name",
                "Account Code", "Account Type", "Related account"])

    def row(date, contact, ref, debit, code, related):
        ws.append([date, "Spend Money", contact, None, contact, None, ref, debit, 0, None, None, None, None,
                    None, "No VAT", code, "Overhead", related])

    ws.append(["Repairs and Maintenance"])
    row(_dt(2025, 2, 1), "Fix It Ltd", "R-1", 600, "6200", "1200 - Bank Current Account")
    row(_dt(2025, 3, 1), "Fix It Ltd", "R-2", 620, "6200", "1200 - Bank Current Account")
    row(_dt(2025, 4, 1), "Fix It Ltd", "R-3", 550, "6200", "")  # unallocated
    row(_dt(2025, 4, 15), "Mixed Vendor", "R-4", 700, "6200",
        "1200 - Bank Current Account, 1250 - Petty Cash and 2 more")  # multi-code split
    suppliers = [
        ("Alpha Supplies", 5000), ("Bravo Supplies", 4500), ("Charlie Supplies", 4000), ("Delta Supplies", 3500),
        ("Echo Supplies", 3000), ("Foxtrot Supplies", 2500), ("Golf Supplies", 2000), ("Hotel Supplies", 1500),
        ("India Supplies", 1000),
    ]
    for i, (name, amount) in enumerate(suppliers):
        row(_dt(2025, 5, 1 + i), name, f"S-{i}", amount, "6200", f"30{i:02d} - {name} Contra")
    row(_dt(2025, 6, 1), "Juliet Supplies", "S-J", 800, "6200", "3010 - Juliet Supplies Contra")
    ws.append(["Total Repairs and Maintenance", None, None, None, None, None, None, 0, 0, 0, 0, 0, 0, None, None, None, None, None])

    ws.append(["Motor Expenses"])
    row(_dt(2025, 3, 10), "Motor Supplies Ltd", "M-1", 300, "6300", "1200 - Bank Current Account")
    ws.append(["Total Motor Expenses", None, None, None, None, None, None, 0, 0, 0, 0, 0, 0, None, None, None, None, None])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_nominal_matrix_and_accruals_prepayments_through_full_generate(http_client):
    """Neither Nominal Analysis Matrix nor Accruals & Prepayments has a
    standalone run-on-its-own section (unlike VAT/PAYE/Control Accounts/
    Debtors & Creditors/FAR/Bank) - both only run inside the full Generate
    pipeline, so this drives that pipeline end to end rather than a
    dedicated "Run" route. Nominal Matrix needs Xero-native nominal
    activity specifically (contra_code/contra_name/contra_needs_review -
    see app/nominal_matrix.py's own docstring on why generic-mapped
    uploads can't produce this schedule); Accruals & Prepayments works
    off ordinary generic-mapped TB/nominal activity."""
    c = http_client
    practice_id = _signup(c, admin_email="matrix-accruals@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Matrix Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
        "comparative_period_start": "2024-01-01", "comparative_period_end": "2024-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    resp = c.post(
        f"/jobs/{job_id}/uploads", data={"section_report_type": "nominal_activity"},
        files={"files": ("gl.xlsx", _make_nominal_matrix_gl_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 303

    from openpyxl import Workbook

    def _make_tb(rows):
        wb = Workbook()
        ws = wb.active
        ws.append(["Account Code", "Account Name", "Account Type", "Debit", "Credit"])
        for r in rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    tb_current = _make_tb([
        ["6200", "Repairs and Maintenance", "Overhead", 30270, 0],
        ["6300", "Motor Expenses", "Overhead", 300, 0],
        ["1400", "Prepaid Insurance", "Prepayment", 2000, 0],
        ["1410", "Prepaid Rent", "Prepayment", 1200, 0],
        ["2400", "Accrued Expenses", "Current Liability", 0, 800],
        ["2410", "Accruals - Other", "Current Liability", 0, 1200],
        ["2420", "Accrued Expenses - VAT", "Current Liability", 0, 300],
    ])
    tb_comparative = _make_tb([
        ["1400", "Prepaid Insurance", "Prepayment", 1500, 0],
        ["1410", "Prepaid Rent", "Prepayment", 2000, 0],
        ["2400", "Accrued Expenses", "Current Liability", 0, 600],
        ["2410", "Accruals - Other", "Current Liability", 0, 1000],
    ])
    # Explicit period on confirm - a bare TB has no date column for
    # guess_period to key off, and its "second upload of this type is
    # probably comparative" fallback only kicks in once the FIRST is
    # already confirmed, which it isn't yet when both are uploaded
    # back to back before either is confirmed.
    _upload_and_confirm_with_period(c, job_id, "trial_balance", "tb.xlsx", tb_current, "current")
    _upload_and_confirm_with_period(c, job_id, "trial_balance", "tb_comparative.xlsx", tb_comparative, "comparative")
    _confirm_all_pending_uploads(c, job_id)

    resp = c.post(f"/jobs/{job_id}/generate", follow_redirects=False)
    assert resp.status_code == 303

    from app import storage
    job = storage.get_job(job_id)
    summary_by_name = {row["name"]: row for row in job["summary"]}

    ap = summary_by_name["Accruals & Prepayments schedule"]
    assert ap["status"] == "review"
    assert "Accrued Expenses - VAT" not in ap["message"]  # excluded (accrued expenses that's actually a VAT account)

    sugg = summary_by_name["Nominal activity - suggested allocations for unallocated transactions"]
    assert sugg["status"] == "review"

    resp = c.get(f"/jobs/{job_id}/download")
    assert resp.status_code == 200
    import openpyxl as _openpyxl
    wb_out = _openpyxl.load_workbook(io.BytesIO(resp.content))
    matrix_sheet_name = next(n for n in wb_out.sheetnames if "Repairs and Maintenance" in n)
    ws_matrix = wb_out[matrix_sheet_name]
    rows_out = list(ws_matrix.iter_rows(values_only=True))

    status_row_text = rows_out[4][0]
    assert "REVIEW" in status_row_text
    assert "multi-code split" in status_row_text
    assert "550" in status_row_text
    assert "folded into" in status_row_text

    header_row = next(r for r in rows_out if r and "TOTAL" in [str(c) for c in r])
    headers = [str(c) for c in header_row if c is not None]
    assert "1200 - Bank Current Account" in headers, "Bank Current Account should survive as its own top-10 column"
    assert "OTHER (see nominal activity detail)" in headers, "expected an OTHER column given >10 distinct contra labels"

    motor_sheet_name = next(n for n in wb_out.sheetnames if "Motor Expenses" in n)
    motor_rows = list(wb_out[motor_sheet_name].iter_rows(values_only=True))
    assert "OK" in motor_rows[4][0]


def test_nominal_matrix_and_accruals_prepayments_standalone_run_through_http(http_client):
    """Nominal Analysis Matrix and Accruals & Prepayments as their own
    standalone sections, same treatment as Control Accounts/Debtors &
    Creditors/Fixed Asset Register/Bank Reconciliation above - each with
    its own independent "Run" button, reusing the job's existing Trial
    Balance/Nominal Activity uploads. MatrixResult (nominal_matrix.
    build_all_matrices) has no `name` field, unlike ReconResult, so
    _matrix_result_to_dict synthesises one from account_code/account_name
    - same shape decision as _control_account_result_to_dict."""
    c = http_client
    practice_id = _signup(c, admin_email="matrix-standalone@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Matrix Standalone Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
        "comparative_period_start": "2024-01-01", "comparative_period_end": "2024-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    resp = c.post(
        f"/jobs/{job_id}/uploads", data={"section_report_type": "nominal_activity"},
        files={"files": ("gl.xlsx", _make_nominal_matrix_gl_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 303

    from openpyxl import Workbook

    def _make_tb(rows):
        wb = Workbook()
        ws = wb.active
        ws.append(["Account Code", "Account Name", "Account Type", "Debit", "Credit"])
        for r in rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    tb_current = _make_tb([
        ["6200", "Repairs and Maintenance", "Overhead", 30270, 0],
        ["6300", "Motor Expenses", "Overhead", 300, 0],
        ["1400", "Prepaid Insurance", "Prepayment", 2000, 0],
        ["1410", "Prepaid Rent", "Prepayment", 1200, 0],
        ["2400", "Accrued Expenses", "Current Liability", 0, 800],
        ["2410", "Accruals - Other", "Current Liability", 0, 1200],
        ["2420", "Accrued Expenses - VAT", "Current Liability", 0, 300],
    ])
    tb_comparative = _make_tb([
        ["1400", "Prepaid Insurance", "Prepayment", 1500, 0],
        ["1410", "Prepaid Rent", "Prepayment", 2000, 0],
        ["2400", "Accrued Expenses", "Current Liability", 0, 600],
        ["2410", "Accruals - Other", "Current Liability", 0, 1000],
    ])
    _upload_and_confirm_with_period(c, job_id, "trial_balance", "tb.xlsx", tb_current, "current")
    _upload_and_confirm_with_period(c, job_id, "trial_balance", "tb_comparative.xlsx", tb_comparative, "comparative")
    _confirm_all_pending_uploads(c, job_id)

    resp = c.post(f"/jobs/{job_id}/nominal-matrix/run", follow_redirects=False)
    assert resp.status_code == 303
    assert resp.headers["location"] == f"/jobs/{job_id}#nominal-matrix"

    resp = c.post(f"/jobs/{job_id}/accruals-prepayments/run", follow_redirects=False)
    assert resp.status_code == 303
    assert resp.headers["location"] == f"/jobs/{job_id}#accruals-prepayments"

    from app import storage
    job = storage.get_job(job_id)

    mx_results = {r["name"]: r for r in job["nominal_matrix_results"]}
    repairs_name = next(n for n in mx_results if "Repairs and Maintenance" in n)
    repairs = mx_results[repairs_name]
    assert repairs["status"] == "review"
    assert "multi-code split" in repairs["message"]
    assert repairs["extra_detail"], "expected the OTHER-bucket breakdown"
    # Regression check for a real bug found via this exact standalone
    # section (screenshot of the rendered table): Postgres jsonb does NOT
    # preserve object key order on a save/reload round trip - it
    # re-orders keys by length, then lexicographically - so several
    # similar-length contra-account column names ("3000 - Alpha Supplies
    # Contra", "3004 - Echo Supplies Contra", ...) came back visibly
    # scrambled unless detail_columns is stored and used explicitly. This
    # asserts the real Postgres-stored order, not just an in-memory dict.
    assert repairs["detail_columns"][:4] == ["date", "reference", "description", "contact"]
    assert repairs["detail_columns"][-2:] == ["TOTAL", "DIFF"]
    assert repairs["detail_columns"].index("1200 - Bank Current Account") < repairs["detail_columns"].index("3000 - Alpha Supplies Contra")
    motor_name = next(n for n in mx_results if "Motor Expenses" in n)
    assert mx_results[motor_name]["status"] == "ok"
    suggestion_name = next(n for n in mx_results if "suggested allocations" in n.lower())
    assert mx_results[suggestion_name]["status"] == "review"
    assert mx_results[suggestion_name]["detail"][0]["Contact"] == "Fix It Ltd"

    ap_results = job["accruals_prepayments_results"]
    assert len(ap_results) == 1
    ap = ap_results[0]
    assert ap["status"] == "review"
    assert "Accrued Expenses - VAT" not in ap["message"]
    codes = {row["Nominal code"] for row in ap["detail"]}
    assert codes == {"1400", "1410", "2400", "2410"}, "the VAT-related accrual account should never appear"

    resp = c.get(f"/jobs/{job_id}")
    assert resp.status_code == 200
    assert "Nominal Analysis Matrix" in resp.text
    assert "Accruals &amp; Prepayments" in resp.text or "Accruals & Prepayments" in resp.text
    assert "Repairs and Maintenance" in resp.text


def _make_payroll_gl_workbook() -> bytes:
    """A minimal but genuinely Xero-native 'Account Transactions' export
    (title rows + one section per account + a Total row per section,
    exactly the shape app.xero_reports.parse_account_transactions
    expects) - one real wages payment, one HMRC payment, one pension
    payment, posted the way a real client export this feature was
    designed against actually does it: net-pay-only postings coded
    straight against a running account, contact carrying the payee
    name, no reference numbers."""
    from datetime import datetime as _dt

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Account Transactions"
    ws.append(["Account Transactions"])
    ws.append(["Test Co Ltd"])
    ws.append(["For the period 1 April 2025 to 31 March 2026"])
    ws.append(["Date", "Source", "Contact", "Contact Group", "Description", "Invoice Number", "Reference",
                "Debit", "Credit", "Running Balance", "Gross", "Net", "VAT", "VAT Rate", "VAT Rate Name",
                "Account Code", "Account Type", "Related account"])
    ws.append(["Wages Payable - Payroll"])
    ws.append([_dt(2025, 5, 2), "Spend Money", "Jamie Doe", None, "Jamie Doe", None, None,
                1084.84, 0, None, None, None, None, None, "No VAT", "814", "Liability", "BANK"])
    ws.append(["Total Wages Payable - Payroll", None, None, None, None, None, None, 0, 0, 0, 0, 0, 0, None, None, None, None, None])
    ws.append(["PAYE Payable"])
    ws.append([_dt(2025, 5, 20), "Spend Money", "HMRC", None, "HMRC", None, None,
                449.52, 0, None, None, None, None, None, "No VAT", "825", "Liability", "BANK"])
    ws.append(["Total PAYE Payable", None, None, None, None, None, None, 0, 0, 0, 0, 0, 0, None, None, None, None, None])
    ws.append(["Pensions Costs"])
    ws.append([_dt(2025, 5, 21), "Spend Money", "Nest", None, "Nest", None, None,
                145.60, 0, None, None, None, None, None, "No VAT", "482", "Expense", "BANK"])
    ws.append(["Total Pensions Costs", None, None, None, None, None, None, 0, 0, 0, 0, 0, 0, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_brightpay_csv(kind: str) -> bytes:
    if kind == "payroll_summary":
        text = (
            "Name,Surname,Gross pay,Taxable gross,Tax,NIC-able gross,Employee NICs,"
            "Student + Postgrad Loan deduction,Net pay,Take-home pay,Employer NICs,Employer pension,Cost to employer\n"
            "Month 1 (Ending 30 April 2025)\n"
            "Jamie,Doe,1200.00,1200.00,50.00,1200.00,25.00,0.00,1084.84,1084.84,40.16,0.00,1240.16\n"
            "TOTAL,,1200.00,1200.00,50.00,1200.00,25.00,0.00,1084.84,1084.84,40.16,0.00,1240.16\n"
        )
    elif kind == "p32":
        text = (
            "Tax period ending,Gross tax,Tax refund received,CIS deductions suffered,Student loan,Postgraduate loan,"
            "Net tax,Gross NICs,SMP recovered,NIC compensation on SMP,SPP recovered,NIC compensation on SPP,"
            "ShPP recovered,NIC compensation on ShPP,SAP recovered,NIC compensation on SAP,SPBP recovered,"
            "NIC compensation on SPBP,SNCP recovered,NIC compensation on SNCP,Employment allowance claim,"
            "Apprenticeship Levy,Total deductions from NICs,Net NICs,Amount due\n"
            "Tax Months 1 to 1 (Summary)\n"
            "05/05/2025,50.00,0.00,0.00,0.00,0.00,50.00,75.00,0.00,0.00,0.00,0.00,0.00,0.00,0.00,0.00,0.00,0.00,0.00,0.00,25.00,0.00,25.00,50.00,449.52\n"
        )
    else:
        text = (
            "Name,Surname,Employee pensionable gross,Employee pension,Employee AVCs,Employer pensionable gross,"
            "Employer pension,Employer AVCs,Employee + employer pension\n"
            "Month 1 (Ending 30 April 2025)\n"
            "Jamie,Doe,1200.00,105.60,0.00,1200.00,40.00,0.00,145.60\n"
            "TOTAL,,1200.00,105.60,0.00,1200.00,40.00,0.00,145.60\n"
        )
    return text.encode()


def test_paye_reconciliation_end_to_end_through_http_and_into_the_workbook(http_client):
    """Full PAYE Reconciliation workspace flow over real HTTP requests: a
    Xero-native General Ledger upload (the job's regular Nominal Activity
    section - no separate GL upload for this workspace), three BrightPay
    uploads (auto-detected from their own column layout, no mapping
    step), settings, the standalone Run button, and finally that the
    same three results land in the generated workbook's own tabs."""
    c = http_client
    practice_id = _signup(c, admin_email="paye-recon@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "PAYE Recon Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-04-01", "current_period_end": "2026-03-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    gl_bytes = _make_payroll_gl_workbook()
    resp = c.post(
        f"/jobs/{job_id}/uploads", data={"section_report_type": "nominal_activity"},
        files={"files": ("gl.xlsx", gl_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 303  # Xero-native Account Transactions auto-confirms straight through

    for kind, filename in [("payroll_summary", "payroll.csv"), ("p32", "p32.csv"), ("pensions", "pensions.csv")]:
        resp = c.post(
            f"/jobs/{job_id}/uploads",
            files={"files": (filename, _make_brightpay_csv(kind), "text/csv")},
            follow_redirects=False,
        )
        assert resp.status_code == 303  # BrightPay native detection auto-confirms too, no type_hint needed

    from app import storage
    job = storage.get_job(job_id)
    assert len(job["uploads"]) == 4
    by_type = {}
    for u in job["uploads"].values():
        by_type.setdefault(u["report_type"], []).append(u)
    assert len(by_type["nominal_activity"]) == 1
    assert len(by_type["paye_summary"]) == 1
    assert len(by_type["paye_p32"]) == 1
    assert len(by_type["paye_pensions"]) == 1
    assert all(u["confirmed"] for u in job["uploads"].values())

    resp = c.post(f"/jobs/{job_id}/paye-recon-settings", data={"paye_recon_tolerance": "0.0", "paye_recon_date_window_days": "60"}, follow_redirects=False)
    assert resp.status_code == 303

    resp = c.post(f"/jobs/{job_id}/paye-recon/run", follow_redirects=False)
    assert resp.status_code == 303

    job = storage.get_job(job_id)
    results = {r["name"]: r for r in job["paye_recon_results"]}
    assert set(results) == {"PAYE Recon - Net Pay by Employee", "PAYE Recon - HMRC PAYE & NI", "PAYE Recon - Pension Contributions"}
    assert results["PAYE Recon - Net Pay by Employee"]["status"] == "ok"
    assert results["PAYE Recon - HMRC PAYE & NI"]["status"] == "ok"
    assert results["PAYE Recon - Pension Contributions"]["status"] == "ok"

    resp = c.get(f"/jobs/{job_id}")
    assert resp.status_code == 200
    assert "PAYE Reconciliation" in resp.text

    resp = c.post(f"/jobs/{job_id}/generate", follow_redirects=False)
    assert resp.status_code == 303

    job = storage.get_job(job_id)
    summary_by_name = {s["name"]: s for s in job["summary"]}
    assert summary_by_name["PAYE Recon - Net Pay by Employee"]["status"] == "ok"
    assert summary_by_name["PAYE Recon - HMRC PAYE & NI"]["status"] == "ok"
    assert summary_by_name["PAYE Recon - Pension Contributions"]["status"] == "ok"

    import openpyxl as _openpyxl
    resp = c.get(f"/jobs/{job_id}/download")
    assert resp.status_code == 200
    wb = _openpyxl.load_workbook(io.BytesIO(resp.content))
    assert any("PAYE Recon - Net Pay" in name for name in wb.sheetnames)
    assert any("PAYE Recon - HMRC" in name for name in wb.sheetnames)
    assert any("PAYE Recon - Pension" in name for name in wb.sheetnames)


def _control_accounts_tb_workbook() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Account Code", "Account Name", "Account Type", "Debit", "Credit"])
    ws.append(["1100", "DEBTORS CONTROL", "Current Asset", 5000, 0])
    ws.append(["2100", "CREDITORS CONTROL", "Current Liability", 0, 3000])
    ws.append(["1200", "BANK", "Bank", 0, 2000])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _control_accounts_nominal_workbook() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Account Code", "Account Name", "Reference", "Description", "Contact", "Source Type", "Debit", "Credit"])
    ws.append(["2025-06-01", "1100", "DEBTORS CONTROL", "INV-1", "Sale", "Acme Ltd", "Invoice", 5000, 0])
    ws.append(["2025-06-15", "2100", "CREDITORS CONTROL", "BILL-1", "Purchase", "Gamma Supplies", "Bill", 0, 3000])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _aged_debtors_workbook() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Customer", "Current", "1-30 days", "31-60 days", "61-90 days", "90 days plus", "Total"])
    ws.append(["Acme Ltd", 5000, 0, 0, 0, 0, 5000])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _aged_creditors_workbook() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Supplier", "Current", "1-30 days", "31-60 days", "61-90 days", "90 days plus", "Total"])
    ws.append(["Gamma Supplies", 3000, 0, 0, 0, 0, 3000])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_control_accounts_fixture(c: TestClient, job_id: str) -> None:
    for section_report_type, filename, content in [
        ("trial_balance", "tb.xlsx", _control_accounts_tb_workbook()),
        ("nominal_activity", "gl.xlsx", _control_accounts_nominal_workbook()),
        ("aged_debtors", "aged_debtors.xlsx", _aged_debtors_workbook()),
        ("aged_creditors", "aged_creditors.xlsx", _aged_creditors_workbook()),
    ]:
        resp = c.post(
            f"/jobs/{job_id}/uploads", data={"section_report_type": section_report_type},
            files={"files": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            follow_redirects=False,
        )
        assert resp.status_code == 303
    _confirm_all_pending_uploads(c, job_id)


def test_control_accounts_standalone_run_through_http(http_client):
    """Control Accounts as its own standalone, run-on-its-own section
    (same "test one section before the full pipeline" treatment as VAT/
    PAYE Reconciliation) - reuses the job's existing Trial Balance/
    Nominal Activity/Aged Debtors/Aged Creditors uploads rather than
    needing any of its own, runs independently via its own button, and
    the results land under job["control_accounts_results"] in the same
    generic {name/status/message/detail/extra_detail/matched_detail}
    shape every other standalone section uses."""
    c = http_client
    practice_id = _signup(c, admin_email="control-accounts@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Control Accounts Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    _upload_control_accounts_fixture(c, job_id)

    resp = c.post(f"/jobs/{job_id}/control-accounts/run", follow_redirects=False)
    assert resp.status_code == 303
    assert resp.headers["location"] == f"/jobs/{job_id}#control-accounts"

    from app import storage
    job = storage.get_job(job_id)
    results = job["control_accounts_results"]
    assert results, "expected at least one control account result"
    names = {r["name"] for r in results}
    assert any("DEBTORS CONTROL" in n for n in names)
    assert any("CREDITORS CONTROL" in n for n in names)
    assert any("wrong control account" in n.lower() for n in names)  # suggest_control_account_miscoding's check
    debtors_result = next(r for r in results if "DEBTORS CONTROL" in r["name"])
    assert debtors_result["status"] in ("ok", "review")
    assert debtors_result["detail"]  # the b/fwd + movement + c/fwd schedule

    resp = c.get(f"/jobs/{job_id}")
    assert resp.status_code == 200
    assert "Control Accounts" in resp.text
    assert "DEBTORS CONTROL" in resp.text


def _far_tb_workbook() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Account Code", "Account Name", "Account Type", "Debit", "Credit"])
    ws.append(["3000", "Motor Vehicles Cost", "Fixed Asset", 18000, 0])
    ws.append(["3050", "Motor Vehicles Depreciation", "Fixed Asset", 0, 10195.31])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _far_tb_comparative_workbook() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Account Code", "Account Name", "Account Type", "Debit", "Credit"])
    ws.append(["3000", "Motor Vehicles Cost", "Fixed Asset", 12000, 0])
    ws.append(["3050", "Motor Vehicles Depreciation", "Fixed Asset", 0, 7000])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _far_nominal_workbook() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Account Code", "Account Name", "Reference", "Description", "Contact", "Source Type", "Debit", "Credit"])
    ws.append(["2025-06-01", "3000", "Motor Vehicles Cost", "INV-9", "New van purchase", "Van Dealer Ltd", "Bill", 6000, 0])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _far_register_workbook() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Asset ID", "Description", "Category", "Date Acquired", "Cost", "Depreciation Method", "Depreciation Rate", "Accumulated Depreciation Brought Forward", "Disposed?"])
    ws.append(["FA-001", "Ford Transit van", "Motor Vehicles", "15/03/2022", 12000, "Reducing Balance", 25, 7000, "No"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_and_confirm_with_period(c: TestClient, job_id: str, section_report_type: str, filename: str, content: bytes, period: str) -> None:
    """Same upload + mapping-confirm flow as _confirm_all_pending_uploads,
    but forces `period` explicitly rather than trusting the auto-detect
    guess - needed here because guess_period's "closest period end wins"
    date heuristic (see document_detection.guess_period) can pick the
    wrong side for a fixture with only one or two GL rows nowhere near
    either period end; a real full-year GL export doesn't have this
    problem, but a minimal test fixture does."""
    resp = c.post(
        f"/jobs/{job_id}/uploads", data={"section_report_type": section_report_type},
        files={"files": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 303
    from app import storage
    job = storage.get_job(job_id)
    upload = next(u for u in job["uploads"].values() if not u["confirmed"])
    confirm_data = {"action": "confirm", "period": period}
    confirm_data.update({f"col__{col}": field for col, field in (upload["mapping"] or {}).items() if field})
    resp = c.post(f"/jobs/{job_id}/uploads/{upload['id']}/mapping", data=confirm_data, follow_redirects=False)
    assert resp.status_code == 303


def _upload_far_fixture(c: TestClient, job_id: str) -> None:
    _upload_and_confirm_with_period(c, job_id, "trial_balance", "tb.xlsx", _far_tb_workbook(), "current")
    _upload_and_confirm_with_period(c, job_id, "trial_balance", "tb_comparative.xlsx", _far_tb_comparative_workbook(), "comparative")
    _upload_and_confirm_with_period(c, job_id, "nominal_activity", "gl.xlsx", _far_nominal_workbook(), "current")
    _upload_and_confirm_with_period(c, job_id, "fixed_asset_register", "far.xlsx", _far_register_workbook(), "comparative")


def test_fixed_asset_register_standalone_run_through_http(http_client):
    """Fixed Asset Register as its own standalone section, same treatment
    as Control Accounts/Debtors & Creditors above: the category-level
    rollforward and capex-miscoding scan need only Trial Balance/Nominal
    Activity; the asset-level detail additionally uses a prior-year Fixed
    Asset Register upload and comes back as an AssetRegisterResult, which
    has no `name` of its own and four data tables rather than the generic
    three - see _asset_register_result_to_dict for how that's mapped onto
    the shared results shape."""
    c = http_client
    practice_id = _signup(c, admin_email="fixed-asset-register@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "FAR Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
        "comparative_period_start": "2024-01-01", "comparative_period_end": "2024-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    _upload_far_fixture(c, job_id)

    resp = c.post(f"/jobs/{job_id}/fixed-asset-register/run", follow_redirects=False)
    assert resp.status_code == 303
    assert resp.headers["location"] == f"/jobs/{job_id}#fixed-asset-register"

    from app import storage
    job = storage.get_job(job_id)
    results = {r["name"]: r for r in job["fixed_asset_register_results"]}
    assert "Fixed asset register (category summary)" in results
    assert "Fixed asset register (asset detail)" in results
    assert any("capital expenditure" in n.lower() for n in results)

    category_result = results["Fixed asset register (category summary)"]
    assert category_result["status"] in ("ok", "review")
    assert category_result["detail"]  # per-category cost/depreciation/NBV rollforward
    assert category_result["extra_detail"]  # additions posted during the year

    register_result = results["Fixed asset register (asset detail)"]
    assert register_result["status"] in ("ok", "review")
    assert register_result["detail"]  # the rolled-forward per-asset schedule
    assert register_result["extra_detail"]  # new additions found in nominal activity, not yet in the register
    assert register_result["extra_detail"][0]["Account"] == "Motor Vehicles Cost"

    resp = c.get(f"/jobs/{job_id}")
    assert resp.status_code == 200
    assert "Fixed Asset Register" in resp.text
    assert "Ford Transit van" in resp.text


def _bank_statement_workbook() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Account", "Statement Date", "Closing Balance"])
    ws.append(["BANK CURRENT ACCOUNT", "2025-12-31", 5000])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _bank_recon_tb_workbook() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Account Code", "Account Name", "Account Type", "Debit", "Credit"])
    ws.append(["1200", "BANK CURRENT ACCOUNT", "Bank", 5000, 0])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_bank_reconciliation_standalone_run_through_http(http_client):
    """Bank Reconciliation as its own standalone section, same treatment
    as Control Accounts/Debtors & Creditors/Fixed Asset Register above -
    the simplest of the four (a single ReconResult, no separate upload of
    its own beyond the Bank Closing Statement and Trial Balance already
    used elsewhere), the last section built for this "every check open
    and testable on its own" phase (see recon.bank_reconciliation)."""
    c = http_client
    practice_id = _signup(c, admin_email="bank-recon@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Bank Recon Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    for section_report_type, filename, content in [
        ("bank_statement", "bank.xlsx", _bank_statement_workbook()),
        ("trial_balance", "tb.xlsx", _bank_recon_tb_workbook()),
    ]:
        resp = c.post(
            f"/jobs/{job_id}/uploads", data={"section_report_type": section_report_type},
            files={"files": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            follow_redirects=False,
        )
        assert resp.status_code == 303
    _confirm_all_pending_uploads(c, job_id)

    resp = c.post(f"/jobs/{job_id}/bank-recon/run", follow_redirects=False)
    assert resp.status_code == 303
    assert resp.headers["location"] == f"/jobs/{job_id}#bank-recon"

    from app import storage
    job = storage.get_job(job_id)
    results = job["bank_recon_results"]
    assert len(results) == 1
    result = results[0]
    assert result["name"] == "Bank reconciliation"
    assert result["status"] == "ok"  # statement (5000) ties exactly to the TB bank balance (5000)
    assert result["detail"][0]["Bank account"] == "BANK CURRENT ACCOUNT"

    resp = c.get(f"/jobs/{job_id}")
    assert resp.status_code == 200
    assert "Bank Reconciliation" in resp.text
    assert "BANK CURRENT ACCOUNT" in resp.text


def test_navigation_breadcrumbs_and_section_nav(http_client):
    """The Practice -> Clients -> Client -> Job breadcrumb chain, and the
    job page's sticky jump-link section nav - added after real user
    feedback that opening a client and going into a job left no way to
    move back up or jump between the many standalone-check sections
    without scrolling. Every level of the practice hierarchy should
    carry a breadcrumb back to its parent; the job page additionally
    gets a "Back to {client}" link and a jump-nav bar across every
    standalone section plus the source-documents upload area."""
    c = http_client
    practice_id = _signup(c, admin_email="nav@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Nav Test Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    resp = c.get(f"/practices/{practice_id}")
    assert resp.status_code == 200
    assert 'class="breadcrumb"' in resp.text

    resp = c.get(f"/practices/{practice_id}/clients")
    assert resp.status_code == 200
    assert 'class="breadcrumb"' in resp.text
    assert f'href="/practices/{practice_id}"' in resp.text

    resp = c.get(f"/clients/{client_id}")
    assert resp.status_code == 200
    html = resp.text
    assert 'class="breadcrumb"' in html
    assert f'href="/practices/{practice_id}"' in html
    assert f'href="/practices/{practice_id}/clients"' in html
    assert "Nav Test Client" in html

    resp = c.get(f"/jobs/{job_id}")
    assert resp.status_code == 200
    html = resp.text
    assert 'class="breadcrumb"' in html
    assert f'href="/practices/{practice_id}"' in html
    assert f'href="/practices/{practice_id}/clients"' in html
    assert f'href="/clients/{client_id}"' in html
    assert "Back to Nav Test Client" in html
    assert 'class="section-nav"' in html
    for anchor in ("#generate", "#vat-recon", "#paye-recon", "#control-accounts",
                   "#debtors-creditors", "#fixed-asset-register", "#bank-recon", "#source-documents"):
        assert f'href="{anchor}"' in html


def test_debtors_creditors_standalone_run_through_http(http_client):
    """Debtors & Creditors as its own standalone section, same treatment
    as Control Accounts above - the aged listing vs Trial Balance control
    account check, with the full customer/supplier-wise listing attached
    (see recon.debtors_creditors_control_recon), independent of Control
    Accounts, VAT/PAYE Reconciliation, and the full Generate pipeline."""
    c = http_client
    practice_id = _signup(c, admin_email="debtors-creditors@acme.test")
    resp = c.post(f"/practices/{practice_id}/clients", data={"name": "Debtors Creditors Client"}, follow_redirects=False)
    client_id = resp.headers["location"].rsplit("/", 1)[-1]
    resp = c.post(f"/clients/{client_id}/jobs", data={
        "current_period_start": "2025-01-01", "current_period_end": "2025-12-31",
    }, follow_redirects=False)
    job_id = resp.headers["location"].rsplit("/", 1)[-1]

    _upload_control_accounts_fixture(c, job_id)

    resp = c.post(f"/jobs/{job_id}/debtors-creditors/run", follow_redirects=False)
    assert resp.status_code == 303
    assert resp.headers["location"] == f"/jobs/{job_id}#debtors-creditors"

    from app import storage
    job = storage.get_job(job_id)
    results = {r["name"]: r for r in job["debtors_creditors_results"]}
    assert set(results) == {"Debtors control account reconciliation", "Creditors control account reconciliation"}
    # aged debtors total (5000) ties exactly to the TB Debtors Control debit balance (5000)
    assert results["Debtors control account reconciliation"]["status"] == "ok"
    assert results["Creditors control account reconciliation"]["status"] == "ok"
    debtors_listing = results["Debtors control account reconciliation"]["extra_detail"]
    assert debtors_listing and debtors_listing[0]["Customer"] == "Acme Ltd"

    resp = c.get(f"/jobs/{job_id}")
    assert resp.status_code == 200
    assert "Debtors &amp; Creditors" in resp.text or "Debtors & Creditors" in resp.text
    assert "Acme Ltd" in resp.text
