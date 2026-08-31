"""End-to-end test: parse the sample data (mirroring real Xero export
structures), run every reconciliation, and build the final workbook."""
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from app import anomaly_detection, compliance_checks, control_accounts, corporation_tax, excel_builder, financial_statements, fixed_assets, mapping, nominal_matrix, parsers, recon, xero_reports
from app.excel_builder import build_workbook

SAMPLE_DIR = Path(__file__).resolve().parent.parent / "sample_data"


def test_trial_balance_self_balances(canonical_data):
    assert abs(canonical_data["tb_current"]["balance"].sum()) < 0.01
    assert abs(canonical_data["tb_comparative"]["balance"].sum()) < 0.01


def test_aged_report_sums_from_detail_rows_not_formula_subtotals(canonical_data):
    # the sample files use un-evaluated '=E9'-style formulas in the subtotal
    # rows (matching Xero's real export quirk) - a naive parser would read
    # those as zero
    moss = canonical_data["aged_debtors"].set_index("customer").loc["MOSS & TURF SUPPLIES"]
    assert moss["total"] == pytest.approx(2370.00)


def test_vat_column_mapping_via_alias_suggestion(canonical_data):
    v = canonical_data["vat_return"].iloc[0]
    assert v["box5"] == pytest.approx(8380.00)
    assert v["box6"] == pytest.approx(78400.00)


def test_recon_engine_runs_all_checks(canonical_data):
    results = recon.run_all_recons(canonical_data)
    names = {r.name for r in results}
    assert "TB self-balance check" in names
    assert all(r.status in ("ok", "review", "error", "n/a") for r in results)
    tb_check = next(r for r in results if r.name == "TB self-balance check")
    assert tb_check.status == "ok"


def test_debtors_creditors_control_recon_lists_every_customer_from_the_aged_report():
    # the aggregate total-vs-TB check alone isn't enough - a preparer
    # needs the client-wise list exactly as submitted in the aged report,
    # not just a single reconciling total
    aged = pd.DataFrame([
        {"customer": "Acme Ltd", "current": 500.0, "bucket_1": 100.0, "bucket_2": 0.0, "bucket_3": 0.0, "bucket_4": 0.0, "older": 0.0, "total": 600.0},
        {"customer": "Widgets Co", "current": 200.0, "bucket_1": 0.0, "bucket_2": 0.0, "bucket_3": 0.0, "bucket_4": 0.0, "older": 0.0, "total": 200.0},
    ])
    tb = pd.DataFrame([{"account_code": "1100", "account_name": "TRADE DEBTORS CONTROL", "account_type": "Current Asset", "balance": 800.0}])
    result = recon.debtors_creditors_control_recon(aged, tb, ["debtors control", "trade debtors"], "customer", "Debtors control recon")
    assert result.status == "ok"
    assert list(result.extra_detail["Customer"]) == ["Acme Ltd", "Widgets Co"]
    assert list(result.extra_detail["Total"]) == [600.0, 200.0]
    assert result.extra_detail.iloc[0]["1 month"] == 100.0
    assert "Customer-wise listing" in result.extra_detail_label


def test_control_account_rollforwards_compute_correctly(canonical_data):
    # the hand-built sample nominal activity is an illustrative subset, not a
    # complete GL, so accounts won't all tie out perfectly - what matters is
    # that b/fwd + movements - c/fwd is computed correctly either way
    results = control_accounts.build_all_rollforwards(
        canonical_data["tb_current"], canonical_data["tb_comparative"], canonical_data["nominal_current"],
        canonical_data["aged_debtors"], canonical_data["aged_creditors"],
    )
    assert len(results) > 0
    assert all(r.status in ("ok", "review") for r in results)
    for r in results:
        b_fwd = r.schedule.iloc[0]["Debit £"] - r.schedule.iloc[0]["Credit £"]
        move_debit, move_credit = r.schedule.iloc[1]["Debit £"], r.schedule.iloc[1]["Credit £"]
        c_fwd = r.schedule.iloc[2]["Debit £"] - r.schedule.iloc[2]["Credit £"]
        computed = b_fwd + move_debit - move_credit
        rollforward_ties = abs(computed - c_fwd) <= control_accounts.MATERIALITY_AMOUNT
        if not rollforward_ties:
            assert r.status == "review"


def test_control_account_breakdown_flags_unexplained_difference(canonical_data):
    # accounts receivable's balance (96,420) is far bigger than the tiny
    # sample aged debtors listing (15,170) - the breakdown should surface
    # that gap as a single explicit number rather than silently ignore it
    results = control_accounts.build_all_rollforwards(
        canonical_data["tb_current"], canonical_data["tb_comparative"], canonical_data["nominal_current"],
        canonical_data["aged_debtors"], canonical_data["aged_creditors"],
    )
    receivable = next(r for r in results if r.account_code == "610A")
    assert not receivable.breakdown.empty
    unexplained_row = receivable.breakdown[receivable.breakdown["Party"].str.startswith("UNEXPLAINED")].iloc[0]
    assert unexplained_row["Amount £"] == pytest.approx(15170.00 - 96420.00)
    assert receivable.status == "review"


def test_nominal_matrix_flags_multi_code_splits(canonical_data):
    results = nominal_matrix.build_all_matrices(
        canonical_data["tb_current"], canonical_data["nominal_current"],
        account_codes=["8010"],
    )
    trade_creditors = results[0]
    assert trade_creditors.status == "review"
    assert "multi-code split" in trade_creditors.message


def test_corporation_tax_marginal_relief_matches_known_reference_point():
    # £150,000 at FY2023+ rates is the standard textbook reference: exactly
    # 24% effective rate, right in the middle of the marginal relief band
    result = corporation_tax.compute(accounting_profit=150_000)
    assert result.band == "marginal relief"
    assert result.tax_charge == pytest.approx(36_000.00)
    assert result.rate_applied == pytest.approx(0.24)


def test_corporation_tax_flags_variance_against_booked_charge(canonical_data):
    accounting_profit = float(canonical_data["pl_current"]["amount"].sum())
    result = corporation_tax.compute(accounting_profit=accounting_profit, booked_tax_charge=0.01)
    assert result.status == "review"
    assert result.variance == pytest.approx(result.tax_charge - 0.01)


def test_balance_sheet_check_ties_once_current_year_profit_bridged_in(canonical_data):
    pl_result = financial_statements.build_pl_statement(canonical_data["pl_current"])
    bs_result = financial_statements.build_bs_statement(canonical_data["bs_current"], pl_result.net_profit)

    net_assets = bs_result.statement.set_index("Line").loc["NET ASSETS", "Amount"]
    total_equity = bs_result.statement.set_index("Line").loc["TOTAL EQUITY", "Amount"]
    check = bs_result.statement.set_index("Line").loc["CHECK: Net Assets - Total Equity (should be £0)", "Amount"]

    assert check == pytest.approx(net_assets - total_equity, abs=0.01)
    assert bs_result.status == "ok"
    assert abs(check) <= financial_statements.MATERIALITY_AMOUNT


def test_balance_sheet_check_flags_a_genuine_gap():
    bs = pd.DataFrame([
        {"account_code": "1", "account_name": "Bank", "category": "Current Asset", "amount": 10000.0},
        {"account_code": "2", "account_name": "Share Capital", "category": "Equity", "amount": -100.0},
        # no retained earnings account at all - deliberately unbalanced
    ])
    result = financial_statements.build_bs_statement(bs, net_profit=0.0)
    assert result.status == "review"
    assert "does not balance" in result.message


def test_balance_sheet_check_names_the_account_with_an_unrecognised_category():
    # an account whose Account Type doesn't map to any of the five
    # recognised B/S categories vanishes from every total (assets,
    # liabilities, AND equity alike) - the check should name it, not just
    # report an unexplained gap
    bs = pd.DataFrame([
        # deliberately balances to zero WITHOUT the Suspense account below -
        # so its exclusion is the entire, exact cause of the gap
        {"account_code": "1", "account_name": "Bank", "category": "Current Asset", "amount": 5000.0},
        {"account_code": "2", "account_name": "Share Capital", "category": "Equity", "amount": -100.0},
        {"account_code": "3", "account_name": "Retained Earnings", "category": "Equity", "amount": -9900.0},
        # a genuinely mistyped Account Type in the source data - not one of
        # Fixed Asset/Current Asset/Bank/Current Liability/Liability/Equity
        {"account_code": "4", "account_name": "Suspense Account", "category": "Unclassified", "amount": 5000.0},
    ])
    result = financial_statements.build_bs_statement(bs, net_profit=0.0)
    assert result.status == "review"
    assert "Suspense Account" in result.message
    assert "doesn't recognise" in result.message
    assert not result.unrecognized_detail.empty
    assert list(result.unrecognized_detail["Account Name"]) == ["Suspense Account"]
    assert result.unrecognized_detail.iloc[0]["Amount"] == pytest.approx(5000.0)


def test_balance_sheet_check_ignores_a_zero_balance_unrecognised_account():
    # a zero-balance account with an odd category isn't contributing to
    # any gap - no point flagging it
    bs = pd.DataFrame([
        {"account_code": "1", "account_name": "Bank", "category": "Current Asset", "amount": 10000.0},
        {"account_code": "2", "account_name": "Share Capital", "category": "Equity", "amount": -10000.0},
        {"account_code": "3", "account_name": "Dormant Suspense", "category": "Unclassified", "amount": 0.0},
    ])
    result = financial_statements.build_bs_statement(bs, net_profit=0.0)
    assert result.status == "ok"
    assert result.unrecognized_detail.empty


def _fa_tb_row(code, name, account_type, debit, credit):
    return {"account_code": code, "account_name": name, "account_type": account_type,
            "debit": debit, "credit": credit, "balance": debit - credit}


def test_fixed_asset_category_grouping_handles_unpunctuated_names():
    # real Xero data uses no separator at all ("IT EQUIPMENT COST BROUGHT
    # FORWARD" rather than "IT EQUIPMENT - COST") - this is a regression
    # test for that, and for the NBV sign convention (accumulated
    # depreciation is a credit balance, so it must reduce NBV, not inflate it)
    tb_current = pd.DataFrame([
        _fa_tb_row("6350", "IT EQUIPMENT COST BROUGHT FORWARD", "Fixed Asset", 1612.19, 0),
        _fa_tb_row("6351", "IT EQUIPMENT COST OF ADDITIONS", "Fixed Asset", 1514.15, 0),
        _fa_tb_row("6360", "IT EQUIPMENT ACCUMULATED DEPRECIATION BROUGHT FORWARD", "Fixed Asset", 0, 89.57),
    ])
    tb_comparative = pd.DataFrame([
        _fa_tb_row("6350", "IT EQUIPMENT COST BROUGHT FORWARD", "Fixed Asset", 1612.19, 0),
        _fa_tb_row("6360", "IT EQUIPMENT ACCUMULATED DEPRECIATION BROUGHT FORWARD", "Fixed Asset", 0, 89.57),
    ])
    result = fixed_assets.category_level_rollforward(tb_current, tb_comparative, pd.DataFrame())

    assert len(result.detail) == 1  # all three rows grouped into one category
    row = result.detail.iloc[0]
    assert row["Category"] == "IT EQUIPMENT"
    assert row["Cost c/fwd (per TB)"] == pytest.approx(3126.34)
    assert row["Acc. depreciation c/fwd (per TB)"] == pytest.approx(89.57)  # positive, not -89.57
    assert row["NBV c/fwd"] == pytest.approx(3126.34 - 89.57)


def test_asset_register_depreciation_straight_line_and_reducing_balance():
    prior_register = pd.DataFrame([
        {"asset_id": "FA-1", "description": "Van", "category": "Motor Vehicles", "date_acquired": pd.Timestamp("2022-01-01"),
         "cost": 18000.0, "depreciation_method": "reducing_balance", "depreciation_rate": 25.0,
         "accumulated_depreciation_b_fwd": 10195.31, "disposed": False},
        {"asset_id": "FA-2", "description": "Laptops", "category": "Computer Equipment", "date_acquired": pd.Timestamp("2023-09-01"),
         "cost": 2400.0, "depreciation_method": "straight_line", "depreciation_rate": 33.33,
         "accumulated_depreciation_b_fwd": 799.20, "disposed": False},
    ])
    result = fixed_assets.asset_level_rollforward(prior_register, None, None, period_days=365)

    van = result.asset_schedule.set_index("Asset ID").loc["FA-1"]
    laptops = result.asset_schedule.set_index("Asset ID").loc["FA-2"]
    # reducing balance: (cost - acc dep b/fwd) * rate
    assert van["Depreciation Charge"] == pytest.approx((18000.0 - 10195.31) * 0.25, abs=0.01)
    # straight line: cost * rate
    assert laptops["Depreciation Charge"] == pytest.approx(2400.0 * 0.3333, abs=0.01)


def test_fixed_asset_register_generic_mapping():
    source = parsers.FileDataSource(SAMPLE_DIR / "fixed_asset_register_prior_year.csv")
    suggestion = mapping.suggest_mapping("fixed_asset_register", source.raw_columns())
    df = parsers.apply_mapping(source, "fixed_asset_register", suggestion)

    assert len(df) == 3
    assert df["cost"].sum() == pytest.approx(18000.0 + 2400.0 + 6000.0)
    van = df[df["asset_id"] == "FA-001"].iloc[0]
    assert van["depreciation_method"].strip().lower() == "reducing balance"
    assert pd.notna(van["date_acquired"])


def _fa_nominal_row(date, code, name, debit=0.0, credit=0.0, reference="", description="", contact="", source_type=""):
    return {"date": pd.Timestamp(date), "account_code": code, "account_name": name, "reference": reference,
            "description": description, "contact": contact, "source_type": source_type, "debit": debit, "credit": credit}


def test_new_addition_labelled_by_migration_keyword_in_description():
    # a real false positive found against real client data: an "Opening
    # Balance" section entry looks identical to a genuine purchase (both
    # are just a debit to a fixed asset cost code) - the description text
    # itself is the strongest available signal to tell them apart
    tb_current = pd.DataFrame([{"account_code": "1", "account_name": "COMPUTER EQUIPMENT COST", "account_type": "Fixed Asset", "balance": 1000.0}])
    prior_register = pd.DataFrame([{
        "asset_id": "FA-1", "description": "Old PC", "category": "Computer Equipment", "date_acquired": pd.Timestamp("2023-01-01"),
        "cost": 500.0, "depreciation_method": "straight_line", "depreciation_rate": 25.0,
        "accumulated_depreciation_b_fwd": 125.0, "disposed": False,
    }])
    nominal = pd.DataFrame([
        _fa_nominal_row("2025-06-01", "1", "COMPUTER EQUIPMENT COST", debit=1000.0, reference="OB-01",
                         description="Opening Balance conversion journal", source_type="Journal"),
    ])
    result = fixed_assets.asset_level_rollforward(prior_register, nominal, tb_current, period_days=365)
    assert result.new_additions.iloc[0]["Addition type"] == fixed_assets._POSSIBLE_MIGRATION


def test_new_addition_labelled_by_date_and_journal_source_near_period_start():
    tb_current = pd.DataFrame([{"account_code": "1", "account_name": "COMPUTER EQUIPMENT COST", "account_type": "Fixed Asset", "balance": 1000.0}])
    prior_register = pd.DataFrame([{
        "asset_id": "FA-1", "description": "Old PC", "category": "Computer Equipment", "date_acquired": pd.Timestamp("2023-01-01"),
        "cost": 500.0, "depreciation_method": "straight_line", "depreciation_rate": 25.0,
        "accumulated_depreciation_b_fwd": 125.0, "disposed": False,
    }])
    # no migration wording at all, but posted within days of the period
    # start on a Journal source - the structural fallback signal
    nominal = pd.DataFrame([
        _fa_nominal_row("2025-01-05", "1", "COMPUTER EQUIPMENT COST", debit=1000.0, reference="JNL-1",
                         description="Reclass entry", source_type="Journal"),
    ])
    result = fixed_assets.asset_level_rollforward(prior_register, nominal, tb_current, period_days=365, period_start=date(2025, 1, 1))
    assert result.new_additions.iloc[0]["Addition type"] == fixed_assets._POSSIBLE_MIGRATION


def test_genuine_mid_year_bill_purchase_not_flagged_as_migration():
    tb_current = pd.DataFrame([{"account_code": "1", "account_name": "COMPUTER EQUIPMENT COST", "account_type": "Fixed Asset", "balance": 1000.0}])
    prior_register = pd.DataFrame([{
        "asset_id": "FA-1", "description": "Old PC", "category": "Computer Equipment", "date_acquired": pd.Timestamp("2023-01-01"),
        "cost": 500.0, "depreciation_method": "straight_line", "depreciation_rate": 25.0,
        "accumulated_depreciation_b_fwd": 125.0, "disposed": False,
    }])
    nominal = pd.DataFrame([
        _fa_nominal_row("2025-06-15", "1", "COMPUTER EQUIPMENT COST", debit=1000.0, reference="INV-2025-042",
                         description="New laptop for John", contact="Dell Ltd", source_type="Bill"),
    ])
    result = fixed_assets.asset_level_rollforward(prior_register, nominal, tb_current, period_days=365, period_start=date(2025, 1, 1))
    assert result.new_additions.iloc[0]["Addition type"] == fixed_assets._LIKELY_GENUINE_ADDITION
    # purely advisory - never changes the existing status/count logic
    assert result.status == "review"
    assert len(result.new_additions) == 1


def test_journal_source_far_from_period_start_not_flagged_as_migration():
    # a Journal source alone isn't enough - it needs to ALSO be near the
    # period start, otherwise a genuine mid-year correcting journal would
    # be wrongly labelled a migration entry
    tb_current = pd.DataFrame([{"account_code": "1", "account_name": "COMPUTER EQUIPMENT COST", "account_type": "Fixed Asset", "balance": 1000.0}])
    prior_register = pd.DataFrame([{
        "asset_id": "FA-1", "description": "Old PC", "category": "Computer Equipment", "date_acquired": pd.Timestamp("2023-01-01"),
        "cost": 500.0, "depreciation_method": "straight_line", "depreciation_rate": 25.0,
        "accumulated_depreciation_b_fwd": 125.0, "disposed": False,
    }])
    nominal = pd.DataFrame([
        _fa_nominal_row("2025-08-20", "1", "COMPUTER EQUIPMENT COST", debit=1000.0, reference="JNL-99",
                         description="Correcting entry", source_type="Journal"),
    ])
    result = fixed_assets.asset_level_rollforward(prior_register, nominal, tb_current, period_days=365, period_start=date(2025, 1, 1))
    assert result.new_additions.iloc[0]["Addition type"] == fixed_assets._LIKELY_GENUINE_ADDITION


def test_far_additions_detail_also_carries_the_addition_type_label():
    tb_current = pd.DataFrame([{"account_code": "1", "account_name": "COMPUTER EQUIPMENT COST", "account_type": "Fixed Asset", "balance": 1000.0}])
    nominal = pd.DataFrame([
        _fa_nominal_row("2025-01-02", "1", "COMPUTER EQUIPMENT COST", debit=1000.0,
                         description="Opening Balance b/fwd", source_type="Journal"),
    ])
    detail = fixed_assets.far_additions_detail(tb_current, nominal)
    assert detail.iloc[0]["Addition type"] == fixed_assets._POSSIBLE_MIGRATION


def test_full_workbook_builds_and_saves(tmp_path, canonical_data):
    results = recon.run_all_recons(canonical_data)
    ca_results = control_accounts.build_all_rollforwards(
        canonical_data["tb_current"], canonical_data["tb_comparative"], canonical_data["nominal_current"],
        canonical_data["aged_debtors"], canonical_data["aged_creditors"],
    )
    mx_results = nominal_matrix.build_all_matrices(canonical_data["tb_current"], canonical_data["nominal_current"])
    ct_computation = corporation_tax.compute(accounting_profit=float(canonical_data["pl_current"]["amount"].sum()))
    fixed_asset_result = fixed_assets.category_level_rollforward(
        canonical_data["tb_current"], canonical_data["tb_comparative"], canonical_data["nominal_current"]
    )

    wb = build_workbook(
        "Brightwell Landscaping Supplies Limited",
        "Year ended 31 December 2025", "Year ended 31 December 2024",
        canonical_data, results, control_account_results=ca_results, matrix_results=mx_results,
        ct_computation=ct_computation, fixed_asset_result=fixed_asset_result,
    )
    out = tmp_path / "working_paper.xlsx"
    wb.save(out)

    assert out.exists()
    reopened = openpyxl.load_workbook(out)
    assert any("Corporation Tax" in s for s in reopened.sheetnames)
    assert "Index" in reopened.sheetnames
    assert any("TB Lead Schedule" in s for s in reopened.sheetnames)
    assert len(reopened.sheetnames) > 10


def test_full_workbook_includes_anomaly_and_compliance_checks(tmp_path, canonical_data):
    results = (
        recon.run_all_recons(canonical_data)
        + anomaly_detection.run_all_anomaly_checks(canonical_data["nominal_current"])
        + compliance_checks.run_all_compliance_checks(
            canonical_data["tb_current"], canonical_data["tb_comparative"], canonical_data["nominal_current"],
            current_year_profit=float(canonical_data["pl_current"]["amount"].sum()),
        )
    )
    wb = build_workbook(
        "Brightwell Landscaping Supplies Limited",
        "Year ended 31 December 2025", "Year ended 31 December 2024",
        canonical_data, results,
    )
    out = tmp_path / "working_paper_with_checks.xlsx"
    wb.save(out)

    reopened = openpyxl.load_workbook(out)
    assert any("Compliance Checklist" in s for s in reopened.sheetnames)
    checklist_ws = next(reopened[s] for s in reopened.sheetnames if "Compliance Checklist" in s)
    non_empty_rows = sum(1 for row in checklist_ws.iter_rows() if row[0].value)
    assert non_empty_rows > len(excel_builder.COMPLIANCE_CHECKLIST_ITEMS)  # title rows + every checklist line

    # the automated checks (DLA/dividend/petty cash/loan) show up as their
    # own recon-style sheets too, not just the static manual checklist
    assert any("DLA Review" in s for s in reopened.sheetnames)
